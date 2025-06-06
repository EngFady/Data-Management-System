
from django.forms import ValidationError
from django.shortcuts import render, get_object_or_404, redirect, HttpResponseRedirect, HttpResponse
from .models import Box, ProviderSegregation, ReimbursementSegregation, master_data, insurance_policy
from .forms import BoxForm, MasterDataForm,ProviderSegregationForm, InsurancePolicyForm,ReimbursementSegregationForm, bursegre,BoxSearchForm,bupsegre, SegreSearcForm, CSVUploadForm, ProviderSegreeDataFile, bub
from datetime import datetime
import csv
from django.db import IntegrityError, transaction
from decimal import Decimal
from django.utils.dateparse import parse_date, parse_datetime
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import authenticate, login
import csv
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .forms import bupsegre  
from .models import Box, ReimbursementSegregation
from django.views.generic import ListView

from django.utils import timezone
from datetime import timedelta
import datetime
from dateutil import parser
import io
from .forms import SegreStgCsvForm

class box_lists(ListView):
    template_name = "arcapp/box_list.html"
    model = Box
    context_object_name = 'boxes'
    def get_queryset(self):
        data = Box.objects.all()[:10]
        return data
    
    

# def box_list(request):
#     boxes = Box.objects.all()
#     return render(request, 'arcapp/box_list.html', {'boxes': boxes})





from django.contrib.auth.decorators import permission_required

################ box form ######################
@permission_required('arcapp.can_create_box', raise_exception=True)
def box_create(request):
    if request.method == 'POST':
        form = BoxForm(request.POST)
        if form.is_valid():
            try:
                box = form.save(commit=False)
                box.created_by = request.user
                box.save()  # Use box.save() instead of form.save()
                messages.success(request, f"Box {box.box_number} created successfully!")
                return redirect('box_list')
            except IntegrityError:
                messages.error(request, "Error: This box number already exists!")
            except ValidationError as e:
                messages.error(request, f"Validation Error: {e}")
        else:
            # Show actual form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.title()}: {error}")
    else:
        form = BoxForm()
    
    return render(request, 'arcapp/home_forms/box_form.html', {'form': form})




# def provider_segregation_create(request):
#     if request.method == 'POST':
#         form = ProviderSegregationForm(request.POST)
#         if form.is_valid():
#             form.save()
#             #return redirect('box_list')
#               # Redirect to any appropriate view
#             messages.success(request, 'the mentioned had been added successfully')
#             messages.success(request, "تم إدخال البيانات بنجاح")
#             print("data saved successfully")
#         else: messages.error(request,f"cannot add the mentioned provider due to {form.errors} ")
#     else:
#         form = ProviderSegregationForm()
#     return render(request, 'arcapp/home_forms/provider_segregation_form.html', {'form': form})






# @permission_required('arcapp.can_create_provider_segre', raise_exception=True)
# def provider_segregation_create(request):
#     if request.method == 'POST':
#         form = ProviderSegregationForm(request.POST)
#         if form.is_valid():
#             print("User:", request.user, "Is Authenticated:", request.user.is_authenticated)
#             providerz = form.save(commit=False)
#             providerz.created_by = request.user
#             providerz.save()

#             messages.success(request, 'The record has been added successfully')
#             messages.success(request, "تم إدخال البيانات بنجاح")
#             return redirect('provider_segregation_create')  # Redirect to same view for a fresh form
#         else:
#             for field, errors in form.errors.items():
#                 for error in errors:
#                     messages.error(request, f"Error in {field}: {error}")
#     else:
#         form = ProviderSegregationForm()
    
#     # Make sure form.media contains both CSS and JS
#     context = {
#         'form': form,
#     }
#     return render(request, 'arcapp/home_forms/provider_segregation_form.html', context)


from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.http import JsonResponse
from .forms import ProviderSegregationForm
from .models import insurance_policy, Box

@permission_required('arcapp.can_create_provider_segre', raise_exception=True)
def provider_segregation_create(request):
    if request.method == 'POST':
        form = ProviderSegregationForm(request.POST)
        if form.is_valid():
            print("User:", request.user, "Is Authenticated:", request.user.is_authenticated)
            
            
            box = form.cleaned_data['box_number']
            policy = form.cleaned_data['PolicyId']
            
            if box.insurance_company.id != policy.insurance_company.id:
                form.add_error('PolicyId', f'Policy ID belongs to {policy.insurance_company.name} but box belongs to {box.insurance_company.name}. They must match.')
                return render(request, 'arcapp/home_forms/provider_segregation_form.html', {'form': form})
            
            providerz = form.save(commit=False)
            providerz.created_by = request.user
            providerz.save()

            messages.success(request, 'The record has been added successfully')
            messages.success(request, "تم إدخال البيانات بنجاح")
            return redirect('provider_segregation_create')  
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")
    else:
        form = ProviderSegregationForm()
    
    context = {
        'form': form,
    }
    return render(request, 'arcapp/home_forms/provider_segregation_form.html', context)


def get_policy_info(request):
    policy_id = request.GET.get('policy_id')
    try:
        # Change from policy_id=policy_id to id=policy_id
        policy = insurance_policy.objects.get(id=policy_id)
        return JsonResponse({
            'client_name': policy.client_name,
            'client_id' : policy.client_id,
            'insurance_company_id': policy.insurance_company.id,
            'insurance_company_name': policy.insurance_company.name
        })
    except insurance_policy.DoesNotExist:
        return JsonResponse({'error': 'Policy not found'}, status=404)

# AJAX endpoint to get box details including insurance company
def get_box_details(request):
    box_id = request.GET.get('box_id')
    try:
        box = Box.objects.get(id=box_id)
        return JsonResponse({
            'box_number': box.box_number,
            'insurance_company_id': box.insurance_company.id,
            'insurance_company_name': box.insurance_company.name
        })
    except Box.DoesNotExist:
        return JsonResponse({'error': 'Box not found'}, status=404)

########### riembursement 2  the implementaion within django itself ###############
import requests
@permission_required('arcapp.can_create_rein_segre', raise_exception=True)
def reimbursement_segregation_create(request):
    if request.method == 'POST':
        form = ReimbursementSegregationForm(request.POST)
        if form.is_valid():
            spiderman = form.save(commit=False)
            spiderman.created_by = request.user
            form.save()
            messages.success(request, 'The record has been added successfully')
            messages.success(request, "تم إدخال البيانات بنجاح")
            # Optional: return to a clean form
            return redirect('reimbursement_segregation_create')  # Redirect to same view for a fresh form
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")
    else:
        form = ReimbursementSegregationForm()
    

    context = {
        'form': form,
    }
    return render(request, 'arcapp/home_forms/reimbursement_segregation_form.html', context)






# @permission_required('arcapp.can_create_rein_segre', raise_exception=True)
# def reimbursement_segregation_create(request):
#     if request.method == 'POST':
#         form = ReimbursementSegregationForm(request.POST)
#         if form.is_valid():
#             data = form.cleaned_data

#             # This dictionary now uses the EXACT field names from your forms.py
#             # and maps them to the camelCase keys for the .NET API.
#             reimbursement_data = {
#                 "boxNumberId": data.get('box_number').id,
#                 "claimCode":   data.get('Claim_Code'),
#                 "batchNumber": data.get('Batch_num'),
#                 #"batch_type" : data.get('Batch_type'),
#                 "batchType" :  data.get('Batch_type'),


#                 "englishName": data.get('English_name'),
#                 "arabicName":  data.get('Arab_name'),
#                 "payer":       data.get('Payer'),
#                 "policy":      data.get('Policy'),
#                 "hof":         data.get('Hof'),
#                 "auditUser":   data.get('Audit_user'),
#                 "auditDate":   data.get('Audit_date').isoformat(),
#                 #"auditDate":   data.get('Audit_date').isoformat()
#             }

#             # NOTE: The 'Batch_type' field is not included here because it's not
#             # in your .NET Reimbursement model. See note below.

#             payload = [reimbursement_data]

#             dotnet_api_url = 'http://localhost:5033/api/ReimbursementApi' # <-- IMPORTANT: Ensure this URL and port are correct

#             try:
#                 response = requests.post(dotnet_api_url, json=payload, verify=False)

#                 if response.status_code == 200:
#                     messages.success(request, 'The record has been sent and saved successfully!')
#                     messages.success(request, 'تم إرسال وحفظ البيانات بنجاح')
#                     return redirect('reimbursement_segregation_create')
#                 else:
#                     error_details = response.json().get('details', response.text)
#                     messages.error(request, f'API Error ({response.status_code}): Could not save the record.')
#                     messages.error(request, f'Details: {error_details}')

#             except requests.exceptions.RequestException as e:
#                 messages.error(request, f"Connection Error: Could not connect to the .NET service.")
#                 messages.error(request, f"Details: {e}")
        
#         else:
#             for field, errors in form.errors.items():
#                 for error in errors:
#                     messages.error(request, f"Error in {field}: {error}")
#     else:
#         form = ReimbursementSegregationForm()

#     context = {
#         'form': form,
#     }
#     return render(request, 'arcapp/home_forms/reimbursement_segregation_form.html', context)












########### riembursement segregation ######################3

# def reimbursement_segregation_create(request):
#     if request.method == 'POST':bulk
#         form = ReimbursementSegregationForm(request.POST)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "submitted successfully")
#             return redirect('reimbursement_segregation_create')  # Redirect to any appropriate view
#         else:
#             print(form.errors)
#             messages.error(request, "Error: Please check the form.")


#     else:
#         form = ReimbursementSegregationForm()   
#     return render(request, 'arcapp/home_forms/reimbursement_segregation_form.html', {'form': form})



################### box crud operation ##################


            ###### box search ######
@permission_required('arcapp.can_view_box', raise_exception=True)
def box_search(request):
    box = None
    error_message = None
    form = BoxSearchForm()

    if request.method == "POST" and "search" in request.POST:
        form = BoxSearchForm(request.POST)
        if form.is_valid():
            box_number = form.cleaned_data['box_number']
            try:
                box = get_object_or_404(Box, box_number=box_number)
            except Exception:
                error_message = f"Box with number {box_number} not found."
    
    return render(request, 'arcapp/crud/box_search.html', {
        'form': form,
        'box': box,
        'error_message': error_message,
    })



        ####### box update ############

from django.shortcuts import get_object_or_404
from django.contrib import messages
from .models import Box, BoxType, BoxLocation, InsuranceCompany  # Add these imports

@permission_required('arcapp.can_update_box', raise_exception=True)
def box_update(request):
    box = None
    success_message = None
    error_message = None

    if request.method == "POST":
        if "search" in request.POST:  
            current_box_number = request.POST.get("current_box_number")
            try:
                box = get_object_or_404(Box, box_number=current_box_number)
            except Exception as e:
                error_message = f"Box with number {current_box_number} not found."
                messages.error(request, error_message)
        elif "update" in request.POST:  
            box_id = request.POST.get("box_id")
            new_box_number = request.POST.get("new_box_number")
            new_box_type = request.POST.get("new_box_type")
            new_box_location = request.POST.get("new_box_location")
            #new_box_status = request.POST.get("new_box_status")
            new_insurance_company = request.POST.get("new_insurance_company")
            
            try:
                box = get_object_or_404(Box, id=box_id)
                
                if new_box_number:
                    box.box_number = new_box_number
                
                # Update to use normalized models
                if new_box_type:
                    box.box_type = get_object_or_404(BoxType, name=new_box_type)
                if new_box_location:
                    box.box_location = get_object_or_404(BoxLocation, name=new_box_location)
                # if new_box_status:
                #     box.box_status = get_object_or_404(BoxStatus, name = new_box_status)
                if new_insurance_company:
                    box.insurance_company = get_object_or_404(InsuranceCompany, name=new_insurance_company)
                
                box.Update_by = request.user
                box.save()
                success_message = "Box updated successfully."
                messages.success(request, success_message)
                
            except Exception as e:
                error_message = f"An error occurred: {str(e)}"
                messages.error(request, error_message)

    return render(request, 'arcapp/crud/box_update.html', {
        'box': box,
        'success_message': success_message,
        'error_message': error_message,
    })
    ######## box delete #########


from django.shortcuts import render, get_object_or_404
from django.http import Http404
from .models import Box  # Ensure you import your Box model

@permission_required('arcapp.can_delete_box', raise_exception=True)
def delete_box_by_number(request):
    success_message = None
    error_message = None
    box = None

    if request.method == "POST" and "delete" in request.POST:
        # Check initial submission (search by current_box_number)
        current_box_number = request.POST.get("current_box_number")
        box_number = request.POST.get("box_number")
        confirmation = request.POST.get("confirmation", "").strip()

        if current_box_number:
            # Initial step: search for the box
            try:
                box = Box.objects.get(box_number=current_box_number)
            except Box.DoesNotExist:
                error_message = f"Box with number '{current_box_number}' not found."
            except Exception as e:
                error_message = f"An error occurred: {str(e)}"
        elif box_number:
            # Confirmation step: validate and delete
            if confirmation != "CONFIRM":
                error_message = "Confirmation text is incorrect."
            else:
                try:
                    box = get_object_or_404(Box, box_number=box_number)
                    box.delete()
                    success_message = f"Box with number '{box_number}' deleted successfully."
                except Http404:
                    error_message = f"Box with number '{box_number}' not found."
                except Exception as e:
                    error_message = f"An error occurred while deleting the Box: {str(e)}"
        else:
            error_message = "Box number is missing in the request."

    return render(request, 'arcapp/crud/box_delete.html', {
        'success_message': success_message,
        'error_message': error_message,
        'box': box,  # Pass the box to show confirmation section
    })


############# provider crud operations ################


    #### provider search #######
@permission_required('arcapp.can_view_provider_segre', raise_exception=True)
def provider_search(request):
    provider = None
    box = None
    box_type = ''
    error_message = None
    form = SegreSearcForm()

    if request.method == "POST" and "search" in request.POST:
        form = SegreSearcForm(request.POST)
        if form.is_valid():
            ClaimCode = form.cleaned_data['claim_code']
            try:
                provider = get_object_or_404(ProviderSegregation, ClaimCode=ClaimCode)
                box_num = provider.box_number
                box = get_object_or_404(Box, box_number = box_num)
                
            except Exception:
                error_message = f"Box with number {ClaimCode} not found."
    
    return render(request, 'arcapp/crud/provider_crud_search.html', {
        'form': form,
        'provider': provider,
        'box' : box,
        'error_message': error_message,
    })


        #### provider segregation update ###########
# def provider_update(request):
#     provider_claim = None
#     success_message = None
#     error_message = None

#     if request.method == "POST":
#         if "search" in request.POST:  # Search for a provider claim by ClaimCode
#             current_claim_code = request.POST.get("current_claim_code")
#             try:
#                 provider_claim = ProviderSegregation.objects.get(ClaimCode=current_claim_code)
#             except ProviderSegregation.DoesNotExist:
#                 error_message = f"Claim with code {current_claim_code} not found."
#             except Exception as e:
#                 error_message = f"An error occurred: {str(e)}"
                
#         elif "update" in request.POST:  # Update the provider claim
#             claim_id = request.POST.get("claim_id")
#             updated_fields = {
#                 "box_number": request.POST.get("new_box_number"),
#                 "ClaimCode": request.POST.get("new_claim_code"),
#                 "PolicyId": request.POST.get("new_policy_id"),
#                 # "policyname": request.POST.get("new_policy_name"),
#                 "ClientId": request.POST.get("new_client_id"),
#                 "Member_name": request.POST.get("new_member_name"),
#                 "Provider_name": request.POST.get("new_provider_name"),
#                 "Receive_date": request.POST.get("new_receive_date"),
#                 "Issuance_date": request.POST.get("new_issuance_date"),
#                 "Segregation_date": request.POST.get("new_segregation_date"),
#                 "Segregation_by": request.POST.get("new_segregation_by"),
#                 "Audit_date": request.POST.get("new_audit_date"),
#                 "Audit_by": request.POST.get("new_audit_by"),
#                 "Box_status": request.POST.get("new_box_status"),
#                 "batchID": request.POST.get("new_batch_id"),
#                 'request_by' : request.POST.get("new_request_by"),
#                 'request_date' : request.POST.get("new_request_date"),                
#                 "Client_name": request.POST.get("new_client_name"),
#                 "claimscan": request.POST.get("new_claimscan"),
#                 "note": request.POST.get("new_note"),
#                 "DeductedAmount": request.POST.get("new_deducted_amount"),
#             }

#             try:
#                 provider_claim = get_object_or_404(ProviderSegregation, id=claim_id)

#                 # Update fields
#                 for field, value in updated_fields.items():
#                     if value:
#                         if field == "box_number":  
#                             try:
                                
#                                 cleaned_value = value.replace("Box ", "").strip()
#                                 provider_claim.box_number = Box.objects.get(box_number=cleaned_value)
#                             except Box.DoesNotExist:
#                                 error_message = f"Box with number {cleaned_value} does not exist."
#                                 break  
#                             except Exception as e:
#                                 error_message = f"An error occurred: {str(e)}"
#                                 break
#                         elif field == 'PolicyId':
#                             try:
                                
#                                 policy = insurance_policy.objects.get(policy_id=value)
#                                 provider_claim.PolicyId = policy
#                             except insurance_policy.DoesNotExist:
#                                 error_message = f"the policy not exist haha"
#                         else:
#                             setattr(provider_claim, field, value)

#                 if not error_message:
#                     provider_claim.Update_by = request.user
#                     provider_claim.save()
#                     success_message = "Claim updated successfully."
                    
#             except Exception as e:
#                 error_message = f"An error occurred: {str(e)}"

#     return render(request, 'arcapp/crud/provider_update.html', {
#         'provider_claim': provider_claim,
#         'success_message': success_message,
#         'error_message': error_message,
#     })






# from .models import ProviderSegregation, Box, insurance_policy
# Assuming your models are in models.py in the same app (arcapp)
from .models import ProviderSegregation, Box, insurance_policy, ClaimStatus, audit_by 


import re
@permission_required('arcapp.can_update_provider_segre', raise_exception=True)
def provider_update(request):
    provider_claim = None
    success_message = None
    error_message = None
    box_location = None
    insurance_company_name = None

    if request.method == "POST":
        if "search" in request.POST:
            current_claim_code = request.POST.get("current_claim_code")
            if not current_claim_code:
                error_message = "Please enter a Claim Code to search."
            else:
                try:
                    provider_claim = ProviderSegregation.objects.select_related(
                        'box_number', 'PolicyId', 'PolicyId__insurance_company',
                        'claim_status', 'Audit_by', 'return_by', 'Segregation_by' 
                    ).get(ClaimCode=current_claim_code)

                    if provider_claim.box_number:
                        box_location = provider_claim.box_number.box_location.name if provider_claim.box_number.box_location else "Location not set"
                    else:
                        box_location = "N/A (No box assigned)"

                    if provider_claim.PolicyId and provider_claim.PolicyId.insurance_company:
                        insurance_company_name = provider_claim.PolicyId.insurance_company.name
                    else:
                        insurance_company_name = "N/A (No policy or insurance company assigned)"

                except ProviderSegregation.DoesNotExist:
                    error_message = f"Claim with code '{current_claim_code}' not found."
                    provider_claim = None
                except Exception as e:
                    error_message = f"An error occurred during search: {str(e)}"
                    provider_claim = None

        elif "update" in request.POST:
            claim_id = request.POST.get("claim_id")
            if not claim_id:
                error_message = "Claim ID not provided for update."
            else:
                try:
                
                    provider_claim = get_object_or_404(
                        ProviderSegregation.objects.select_related(
                            'box_number', 'PolicyId', 'PolicyId__insurance_company',
                            'claim_status', 'Audit_by', 'return_by', 'Segregation_by', 'request_by'
                        ),
                        id=claim_id
                    )

                    
                    


                    
                    #provider_claim.ClaimCode = request.POST.get("new_claim_code", provider_claim.ClaimCode).strip()

                    provider_claim.Member_name = request.POST.get("new_member_name", provider_claim.Member_name).strip()
                    provider_claim.Provider_name = request.POST.get("new_provider_name", provider_claim.Provider_name).strip()
                    
                    provider_claim.batchID = request.POST.get("new_batch_id", provider_claim.batchID).strip()
                   #provider_claim.request_by = request.POST.get("new_request_by", provider_claim.request_by).strip()
                    provider_claim.note = request.POST.get("new_note", provider_claim.note).strip()
                    provider_claim.comment = request.POST.get("new_comment", provider_claim.comment).strip()

                
                    claimscan_val_str = request.POST.get("new_claimscan")
                    if claimscan_val_str is not None:
                        provider_claim.claimscan = claimscan_val_str.strip().lower() in ['true', 'on', '1', 'yes']
                    
                    
                    deducted_amount_str = request.POST.get("new_deducted_amount", "").strip()
                    if deducted_amount_str:
                        try:
                            provider_claim.DeductedAmount = Decimal(deducted_amount_str)
                        except InvalidOperation:
                            error_message = "Invalid value for Deducted Amount. Please enter a valid number."
                    elif provider_claim.DeductedAmount is not None : 
                            provider_claim.DeductedAmount = None

                    
                    date_fields_map = {
                        "Receive_date": "new_receive_date", "Issuance_date": "new_issuance_date",
                        "Segregation_date": "new_segregation_date", "Audit_date": "new_audit_date",
                        "request_date": "new_request_date", "retrieval_date": "new_retrieval_date",
                        "return_date": "new_return_date",
                    }
                    for model_field, form_field in date_fields_map.items():
                        date_val_str = request.POST.get(form_field, "").strip()
                        if date_val_str:
                            try:
                                setattr(provider_claim, model_field, date_val_str) 
                            except ValueError:
                                if not error_message: error_message = f"Invalid date format for {model_field.replace('_', ' ').title()}."
                        elif getattr(provider_claim, model_field) is not None:
                            setattr(provider_claim, model_field, None)

                    
                    
                    new_box_number_str = request.POST.get("new_box_number", "").strip()
                    if not error_message and new_box_number_str:
                        current_box_val = str(getattr(provider_claim.box_number, 'box_number', None))
                        if current_box_val != new_box_number_str:
                            try:
                                box_instance = Box.objects.get(box_number=new_box_number_str)
                                provider_claim.box_number = box_instance
                            except Box.DoesNotExist:
                                error_message = f"Box with number '{new_box_number_str}' does not exist."
                    elif not error_message and not new_box_number_str and provider_claim.box_number is not None:
                        if ProviderSegregation._meta.get_field('box_number').null:
                            provider_claim.box_number = None
                        else:
                            if not error_message: error_message = "Box Number cannot be empty as it's a required field."
                    new_claim_coding_str = request.POST.get("new_claim_code", "").strip()
                    #current_claim_code = str(getattr(provider_claim.ClaimCode, 'ClaimCode', None))
                    current_claim_code = provider_claim.ClaimCode
                    if current_claim_code !=  new_claim_coding_str:
                        #claim_code_pattern = re.compile(r"^\d+-d+$")
                        claim_code_pattern = re.compile(r"^\d+-\d+$")
                        if not claim_code_pattern.fullmatch(new_claim_coding_str):
                            print('haha')
                            
                            error_message = f"the claim code \"({new_claim_coding_str})\" not in AMC shapping for claims i.e  to be like 180113-22"
                        else : provider_claim.ClaimCode = new_claim_coding_str
                    
                            
        
                    new_policy_id_str = request.POST.get("new_policy_id", "").strip()
                    if not error_message and new_policy_id_str:
                        current_policy_val = str(getattr(provider_claim.PolicyId, 'policy_id', None))
                        if current_policy_val != new_policy_id_str:
                            try:
                                #policy = insurance_policy.objects.get(policy_id=new_policy_id_str)
                                policy = insurance_policy.objects.get(policy_id=new_policy_id_str)
                                box_num_claim = provider_claim.box_number
                                boxinc = Box.objects.get(box_number = str(box_num_claim))
                                #print(boxinc.insurance_company)
                                policinc = policy.insurance_company
                                #print(policinc)
                                if policinc == boxinc.insurance_company:
                                    #print('the 2 vals validd ')
                                    provider_claim.PolicyId = policy
                                    provider_claim.Client_name = policy.client_name 
                                    provider_claim.ClientId = policy.client_id
                                else: error_message = f"sorry to hear that but the box insurace company \"{boxinc.insurance_company}\" not the same pilicy insurance company (\"{policinc}\")" 
                            except insurance_policy.DoesNotExist:
                                error_message = f"Policy with ID '{new_policy_id_str}' not found."
                            except ValueError:
                                error_message = f"Policy ID '{new_policy_id_str}' is not a valid number."
                    elif not error_message and not new_policy_id_str and provider_claim.PolicyId is not None:
                        if ProviderSegregation._meta.get_field('PolicyId').null: 
                            provider_claim.PolicyId = None
                            provider_claim.Client_name = None
                            provider_claim.ClientId = None
                        else:
                            if not error_message: error_message = "Policy ID cannot be empty as it's a required field."
                    if new_policy_id_str:

                        
                            policy = insurance_policy.objects.get(policy_id=new_policy_id_str)
                            box_num_claim = provider_claim.box_number
                            boxinc = Box.objects.get(box_number = str(box_num_claim))
                            print(boxinc.insurance_company)
                            policinc = policy.insurance_company
                            print(policinc)
                            if policinc == boxinc.insurance_company:
                                print('the 2 vals validd ')
                                

                            else : print('no')
                        

                        
                    
                    new_claim_status_str = request.POST.get("new_claim_status_str", "").strip() 
                    if not error_message and new_claim_status_str:
                        current_status_val = getattr(provider_claim.claim_status, 'name', None)
                        if str(current_status_val) != new_claim_status_str:
                            try:
                                claim_status_instance = ClaimStatus.objects.get(name=new_claim_status_str)
                                provider_claim.claim_status = claim_status_instance
                            except ClaimStatus.DoesNotExist:
                                error_message = f"Claim Status with name '{new_claim_status_str}' does not exist."
                    elif not error_message and not new_claim_status_str and provider_claim.claim_status is not None:
                        if ProviderSegregation._meta.get_field('claim_status').null: 
                            provider_claim.claim_status = None
                        else:
                            if not error_message: error_message = "Claim Status cannot be empty as it's a required field."

                    
                    new_audit_by_str = request.POST.get("new_audit_by", "").strip() 
                    if not error_message and new_audit_by_str:
                        current_audit_by_name = getattr(provider_claim.Audit_by, 'name', None)
                        if str(current_audit_by_name) != new_audit_by_str:
                            try:
                                audit_by_instance = audit_by.objects.get(name=new_audit_by_str)
                                provider_claim.Audit_by = audit_by_instance
                            except audit_by.DoesNotExist:
                                error_message = f"Audit By user '{new_audit_by_str}' does not exist. Please add to master data."
                    elif not error_message and not new_audit_by_str and provider_claim.Audit_by is not None:
                        if ProviderSegregation._meta.get_field('Audit_by').null: #
                            provider_claim.Audit_by = None
                        else:
                            if not error_message: error_message = "Audit By cannot be empty as it's a required field."
                    
                    
                    new_return_by_str = request.POST.get("new_return_by", "").strip() 
                    if not error_message and new_return_by_str:
                        current_return_by_name = getattr(provider_claim.return_by, 'name', None)
                        if str(current_return_by_name) != new_return_by_str:
                            try:
                                return_by_instance = return_by_user.objects.get(name=new_return_by_str)
                                provider_claim.return_by = return_by_instance
                            except return_by_user.DoesNotExist:
                                error_message = f"Return By user '{new_return_by_str}' does not exist. Please add to master data."
                    elif not error_message and not new_return_by_str and provider_claim.return_by is not None:
                        
                        if ProviderSegregation._meta.get_field('return_by').null:
                            provider_claim.return_by = None
                       
                    new_request_by_str = request.POST.get("new_request_by", "").strip()
                    if not error_message and new_request_by_str:
                        current_request_by_usr = getattr(provider_claim.request_by, 'name', None)
                        if str(current_request_by_usr) != new_request_by_str:
                            try:
                                request_by_instance = request_by.objects.get(name = new_request_by_str)
                                provider_claim.request_by = request_by_instance
                            except request_by.DoesNotExist:
                                error_message = f"the current request user {new_request_by_str} not exist on master data hence please fill it into master data and then try again on adding the claim"
                            except Exception as e:
                                error_message = f"cannot insert the row due to {e}"
                        elif not error_message and not new_request_by_str and provider_claim.request_by is not None:
                        
                                if ProviderSegregation._meta.get_field('request_by').null:
                                    provider_claim.request_by = None
                    new_segregation_by_str = request.POST.get("new_segregation_by", "").strip() 
                    if not error_message and new_segregation_by_str:
                        current_segregation_by_name = getattr(provider_claim.Segregation_by, 'name', None)
                        if str(current_segregation_by_name) != new_segregation_by_str:
                            try:
                                segregation_by_instance = segregation_by.objects.get(name=new_segregation_by_str)
                                provider_claim.Segregation_by = segregation_by_instance 
                            except segregation_by.DoesNotExist:
                                error_message = f"Segregation By user '{new_segregation_by_str}' does not exist. Please add to master data."
                    elif not error_message and not new_segregation_by_str and provider_claim.Segregation_by is not None:
                
                        if ProviderSegregation._meta.get_field('Segregation_by').null:
                            provider_claim.Segregation_by = None
                     

                    if not error_message:
                        provider_claim.Update_by = request.user
                        provider_claim.save()
                        success_message = "Claim updated successfully."
                        
                        provider_claim = get_object_or_404(
                            ProviderSegregation.objects.select_related(
                                'box_number', 'PolicyId', 'PolicyId__insurance_company',
                                'claim_status', 'Audit_by', 'return_by', 'Segregation_by'
                            ),
                            id=claim_id
                        )
                
                except ProviderSegregation.DoesNotExist:
                    error_message = "Claim not found for update."
                except IntegrityError as ie:
                    error_message = f"Data integrity error: {str(ie)}. This might be due to a duplicate Claim Code if changed."
                except Exception as e:
                    if not error_message: 
                        error_message = f"An unexpected error occurred during update: {str(e)}"
            

            if provider_claim:
                if provider_claim.box_number and provider_claim.box_number.box_location:
                    box_location = provider_claim.box_number.box_location.name
                elif provider_claim.box_number:
                    box_location = "Location not set for box"
                else:
                    box_location = "N/A (No box assigned)"

                if provider_claim.PolicyId and provider_claim.PolicyId.insurance_company:
                    insurance_company_name = provider_claim.PolicyId.insurance_company.name
                else:
                    insurance_company_name = "N/A (No policy or insurance company assigned)"

    if error_message:
        success_message = None 

    context_data = {
        'provider_claim': provider_claim,
        'box_location': box_location,
        'insurance_company': insurance_company_name,
        'success_message': success_message,
        'error_message': error_message,
        'all_audit_by_options': audit_by.objects.all().order_by('name'),
        'all_return_by_options': return_by_user.objects.all().order_by('name'),
        'all_segregation_by_options': segregation_by.objects.all().order_by('name'), 
        'all_claim_status_options': ClaimStatus.objects.all().order_by('name'),
        'all_box_options': Box.objects.all().order_by('box_number'),
        'all_policy_options': insurance_policy.objects.select_related('insurance_company').all().order_by('policy_id'),
    }
    return render(request, 'arcapp/crud/provider_update.html', context_data)



# from django.shortcuts import render, get_object_or_404
# from django.contrib.auth.decorators import permission_required
# from .models import ProviderSegregation, Box, insurance_policy # Ensure these are imported

# @permission_required('arcapp.can_update_provider_segre', raise_exception=True)
# def provider_update(request):
#     provider_claim = None
#     success_message = None
#     error_message = None

#     if request.method == "POST":
#         if "search" in request.POST:
#             current_claim_code = request.POST.get("current_claim_code")
#             try:
#                 provider_claim = ProviderSegregation.objects.get(ClaimCode=current_claim_code)
#             except ProviderSegregation.DoesNotExist:
#                 error_message = f"Claim with code {current_claim_code} not found."
#             except Exception as e:
#                 error_message = f"An error occurred while searching: {str(e)}"
                
#         elif "update" in request.POST:
#             claim_id = request.POST.get("claim_id")
#             try:
#                 provider_claim = get_object_or_404(ProviderSegregation, id=claim_id)

#                 # Fields that are direct string/date/number inputs
#                 direct_update_fields = [
#                     "ClaimCode", "ClientId", "Member_name", "Provider_name", 
#                     "Receive_date", "Issuance_date", "Segregation_date", 
#                     "Segregation_by", "Audit_date", "Audit_by", "Box_status", 
#                     "batchID", 'request_by', 'request_date', "Client_name", 
#                     "claimscan", "note", "DeductedAmount"
#                 ]

#                 for field_name in direct_update_fields:
#                     # Use new_ prefix convention from template for POST data
#                     post_key = f"new_{field_name.lower().replace('claimcode', 'claim_code').replace('clientid', 'client_id').replace('member_name','member_name').replace('provider_name','provider_name').replace('receive_date','receive_date').replace('issuance_date','issuance_date').replace('segregation_date','segregation_date').replace('segregation_by','segregation_by').replace('audit_date','audit_date').replace('audit_by','audit_by').replace('box_status','box_status').replace('batchid','batch_id').replace('client_name','client_name').replace('claimscan','claimscan').replace('deductedamount','deducted_amount')}"
#                     if field_name == "ClaimCode": post_key = "new_claim_code"
#                     elif field_name == "ClientId": post_key = "new_client_id"
#                     elif field_name == "batchID": post_key = "new_batch_id"
#                     elif field_name == "DeductedAmount": post_key = "new_deducted_amount"


#                     value = request.POST.get(post_key)
#                     if value is not None: # Allow clearing fields by submitting empty string
#                          # Handle boolean for claimscan if it's a checkbox in spirit
#                         if field_name == "claimscan":
#                             # Assuming 'on' or 'true' for checked, or handle as per your form input
#                             # For a text input that might represent boolean:
#                             processed_value = value.lower() in ['true', 'on', 'yes', '1']
#                             setattr(provider_claim, field_name, processed_value)
#                         elif field_name in ["Receive_date", "Issuance_date", "Segregation_date", "Audit_date", "request_date"]:
#                             if value == '': # Handle empty date string
#                                 setattr(provider_claim, field_name, None)
#                             else:
#                                 setattr(provider_claim, field_name, value)
#                         elif field_name == "DeductedAmount":
#                             if value == '':
#                                 setattr(provider_claim, field_name, None)
#                             else:
#                                 try:
#                                     setattr(provider_claim, field_name, float(value))
#                                 except ValueError:
#                                     error_message = f"Invalid value for Deducted Amount: {value}"
#                                     break 
#                         else:
#                             setattr(provider_claim, field_name, value)
                
#                 if error_message: # If error from direct fields, stop before foreign keys
#                     pass
#                 else:
#                     # Handle Box Number (Foreign Key from ID)
#                     box_id = request.POST.get("new_box_number")
#                     if box_id:
#                         try:
#                             selected_box = Box.objects.get(id=box_id)
#                             provider_claim.box_number = selected_box
#                         except Box.DoesNotExist:
#                             error_message = f"Box with ID {box_id} does not exist."
#                         except ValueError:
#                             error_message = f"Invalid Box ID format: {box_id}."
#                     elif provider_claim.box_number is not None: # If box_id is empty, means user wants to clear it
#                         provider_claim.box_number = None


#                     # Handle Policy ID (Foreign Key from ID)
#                     if not error_message: # Proceed if no error from box
#                         policy_obj_id = request.POST.get("new_policy_id")
#                         if policy_obj_id:
#                             try:
#                                 selected_policy = insurance_policy.objects.get(id=policy_obj_id)
#                                 provider_claim.PolicyId = selected_policy
                                
#                                 # --- Server-side validation for insurance company match ---
#                                 if provider_claim.box_number and provider_claim.PolicyId:
#                                     if provider_claim.box_number.insurance_company != provider_claim.PolicyId.insurance_company:
#                                         error_message = (f"Update failed: Box belongs to "
#                                                        f"{provider_claim.box_number.insurance_company.name} but Policy "
#                                                        f"belongs to {provider_claim.PolicyId.insurance_company.name}. "
#                                                        f"They must match.")
#                                 # --- End server-side validation ---

#                             except insurance_policy.DoesNotExist:
#                                 error_message = f"Policy with ID {policy_obj_id} does not exist."
#                             except ValueError:
#                                 error_message = f"Invalid Policy ID format: {policy_obj_id}."
#                         elif provider_claim.PolicyId is not None: # If policy_obj_id is empty
#                              provider_claim.PolicyId = None


#                 if not error_message:
#                     provider_claim.Update_by = request.user # Assuming Update_by field exists
#                     provider_claim.save()
#                     success_message = "Claim updated successfully."
#                     # provider_claim = ProviderSegregation.objects.get(id=claim_id) # Re-fetch to show updated data
            
#             except ProviderSegregation.DoesNotExist:
#                 error_message = "Claim not found for update."
#             except Exception as e:
#                 error_message = f"An error occurred during update: {str(e)}"

#     return render(request, 'arcapp/crud/provider_update.html', {
#         'provider_claim': provider_claim,
#         'success_message': success_message,
#         'error_message': error_message,
#     })







        ##### provider deletion ###########
@permission_required('arcapp.can_delete_provider_segre', raise_exception=True)
def provider_delete(request):
    provider = None
    success_message = None
    error_message = None

    if request.method == "POST" and "delete" in request.POST:
        claim_code = request.POST.get("claim_code")
        try:
            provider = get_object_or_404(ProviderSegregation, ClaimCode=claim_code)
            provider.delete()
            success_message = "provider claim deleted successfully."
            provider = None
        except Exception as e:
            error_message = f"An error occurred while deleting: {str(e)}"

    return render(request, 'arcapp/crud/provider_delete.html', {
        'provider': provider,
        'success_message': success_message,
        'error_message': error_message,
    })

############## riembursement crud operation #################

        ##### riem search ######
@permission_required('arcapp.can_view_reim_segre', raise_exception=True)
def riem_search(request):
    riem = None
    box = None
    error_message = None
    form = SegreSearcForm()

    if request.method == "POST" and "search" in request.POST:
        form = SegreSearcForm(request.POST)
        if form.is_valid():
            claim_code = form.cleaned_data['claim_code']
            print(f"Searching for Claim_Code: {claim_code}")  

            try:
                riem = get_object_or_404(ReimbursementSegregation, Claim_Code=claim_code)
                box = get_object_or_404(Box, box_number = riem.box_number)
                print(f"Found: {riem}")  

            except Exception as e:
                error_message = f"Segregation with Claim Code {claim_code} not found."
                print(error_message)  
                print(e)  
    
    return render(request, 'arcapp/crud/reim_search.html',          
                   {
        'form': form,
        'riem': riem,
        'box' : box,
        'error_message': error_message,
    })

    ####### riem update ########
@permission_required('arcapp.can_update_reim_segre', raise_exception=True)
def reimbursement_update(request):
    reimbursement_claim = None
    success_message = None
    error_message = None

    if request.method == "POST":
        if "search" in request.POST:  
            current_claim_code = request.POST.get("current_claim_code")
            try:
                reimbursement_claim = ReimbursementSegregation.objects.get(Claim_Code=current_claim_code)
            except ReimbursementSegregation.DoesNotExist:
                error_message = f"Claim with code {current_claim_code} not found."
            except Exception as e:
                error_message = f"An error occurred: {str(e)}"
                
        elif "update" in request.POST:  
            claim_id = request.POST.get("claim_id")
            updated_fields = {
                "box_number": request.POST.get("new_box_number"),
                "Claim_Code": request.POST.get("new_claim_code"),
                "Batch_num": request.POST.get("new_batch_num"),
                "Batch_type": request.POST.get("new_batch_type"),
                "English_name": request.POST.get("new_english_name"),
                "Arab_name": request.POST.get("new_arab_name"),
                "Payer": request.POST.get("new_payer"),
                "Policy": request.POST.get("new_policy"),
                "Hof": request.POST.get("new_hof"),
                "Audit_user": request.POST.get("new_audit_user"),
                "Audit_date": request.POST.get("new_audit_date"),
            }

            try:
                reimbursement_claim = get_object_or_404(ReimbursementSegregation, id=claim_id)

                for field, value in updated_fields.items():
                    if value:
                        if field == "box_number":  
                            try:
                                
                                cleaned_value = value.replace("Box ", "").strip()
                                reimbursement_claim.box_number = Box.objects.get(box_number=cleaned_value)
                            except Box.DoesNotExist:
                                error_message = f"Box with number {cleaned_value} does not exist."
                                break  
                            except Exception as e:
                                error_message = f"An error occurred: {str(e)}"
                                break
                        else:
                            setattr(reimbursement_claim, field, value)

                if not error_message:
                    reimbursement_claim.Update_by = request.user
                    
                    reimbursement_claim.save()
                    success_message = "Claim updated successfully."
                    
            except Exception as e:
                error_message = f"An error occurred: {str(e)}"

    return render(request, 'arcapp/crud/reimbursement_update.html', {
        'reimbursement_claim': reimbursement_claim,
        'success_message': success_message,
        'error_message': error_message,
    })
    ####### riembursement delete ######## 
@permission_required('arcapp.can_delete_reim_segre', raise_exception=True)
def riem_delete(request):
    success_message = None
    error_message = None

    if request.method == "POST" and "delete" in request.POST:
        claim_code = request.POST.get("claim_code")
        try:

            reimbursement = get_object_or_404(ReimbursementSegregation, Claim_Code=claim_code)
            reimbursement.delete()
            success_message = f"Reimbursement with Claim Code '{claim_code}' deleted successfully."
        except Exception as e:
            error_message = f"An error occurred: {str(e)}"

    return render(request, 'arcapp/crud/riem_delete.html', {
        'success_message': success_message,
        'error_message': error_message,
    })


################## bulk insert boxes ##################

# def bulk_insert_box(request):
#     if request.method == 'POST' and request.FILES.get('csv_file'):
#         csv_file = request.FILES['csv_file']

#         # Ensure the uploaded file is a CSV
#         if not csv_file.name.endswith('.csv'):
#             return HttpResponse("Please upload a CSV file.", status=400)

#         try:
#             # Read the CSV content
#             data = csv_file.read().decode('utf-8').splitlines()
#             reader = csv.DictReader(data, delimiter=',')
#             csv_columns = reader.fieldnames

#             # Define required columns for the Box model
#             required_columns = ['box_number', 'box_type', 'box_location', 'insurance_company']
#             if not all(col in csv_columns for col in required_columns):
#                 return HttpResponse("CSV is missing required columns!", status=400)

#             with transaction.atomic():
#                 boxes = [
#                     Box(
#                         box_number=row['box_number'],
#                         box_type=row['box_type'],
#                         box_location=row['box_location'],
#                         insurance_company=row['insurance_company'] if row['insurance_company'] else None,
#                     )
#                     for row in reader
#                 ]
#                 # Bulk insert
#                 Box.objects.bulk_create(boxes)

#             return HttpResponse("Data has been uploaded successfully!")
#         except Exception as e:
#             return HttpResponse(f"An error occurred: {str(e)}", status=500)

#     return render(request, 'arcapp/bulk_insert/upload.html', {'form': CSVUploadForm()})

                    
from django.http import JsonResponse
from django.db import transaction
import csv

# def bulk_insert_box(request):
#     if request.method == 'POST' and request.FILES.get('csv_file'):
#         csv_file = request.FILES['csv_file']

#         # Ensure the uploaded file is a CSV
#         if not csv_file.name.endswith('.csv'):
#             return HttpResponse("Please upload a CSV file.", status=400)

#         try:
#             # Read the CSV content
#             data = csv_file.read().decode('utf-8').splitlines()
#             reader = csv.DictReader(data, delimiter=',')
#             csv_columns = reader.fieldnames

#             # Define required columns for the Box model
#             required_columns = ['box_number', 'box_type', 'box_location', 'insurance_company']
#             if not all(col in csv_columns for col in required_columns):
#                 return HttpResponse("CSV is missing required columns!", status=400)

#             error_details = []
#             valid_boxes = []
#             row_number = 1  # Start row tracking from 1

#             with transaction.atomic():
#                 for row in reader:
#                     row_number += 1  # Track row number
                    
#                     # Validate each column in the row
#                     for col in required_columns:
#                         if row[col] is None or row[col].strip() == "":
#                             error_details.append(f"Row {row_number}: Missing value in column '{col}'")

#                     try:
#                         box = Box(
#                             box_number=row['box_number'],
#                             box_type=row['box_type'],
#                             box_location=row['box_location'],
#                             insurance_company=row['insurance_company'] if row['insurance_company'] else None,
#                         )
#                         valid_boxes.append(box)
#                     except Exception as e:
#                         error_details.append(f"Row {row_number}: {str(e)}")

#                 # If there are errors, return a detailed error message
#                 if error_details:
#                     return JsonResponse({"errors": error_details}, status=400)

#                 # Bulk insert valid data
#                 Box.objects.bulk_create(valid_boxes)

#             return HttpResponse("Data has been uploaded successfully!")
#         except Exception as e:
#             return HttpResponse(f"An error occurred: {str(e)}", status=500)

#     return render(request, 'arcapp/bulk_insert/upload.html', {'form': CSVUploadForm()})






# def bulk_insert_box(request):
#     if request.method == 'POST' and request.FILES.get('csv_file'):
#         csv_file = request.FILES['csv_file']

#         if not csv_file.name.endswith('.csv'):
#             if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                 return JsonResponse({"success": False, "message": "Please upload a CSV file."}, status=400)
#             return render(request, 'arcapp/bulk_insert/upload.html', 
#                         {'error_message': "Please upload a CSV file."})

#         try:
#             data = csv_file.read().decode('utf-8').splitlines()
#             reader = csv.DictReader(data, delimiter=',')
#             csv_columns = reader.fieldnames

#             required_columns = ['box_number', 'box_type', 'box_location', 'insurance_company']
#             if not all(col in csv_columns for col in required_columns):
#                 if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                     return JsonResponse({"success": False, "message": "CSV is missing required columns!"}, status=400)
#                 return render(request, 'arcapp/bulk_insert/upload.html', 
#                             {'error_message': "CSV is missing required columns!"})

#             error_details = []
#             valid_boxes = []
#             row_number = 1  

#             with transaction.atomic():
#                 for row in reader:
#                     row_number += 1  
#                     errors_in_row = []

#                     for col in required_columns:
#                         if not row.get(col) or row[col].strip() == "":
#                             errors_in_row.append(f"Missing value in column '{col}'")
                    
#                     if errors_in_row:
#                         error_details.append(f"Row {row_number}: {', '.join(errors_in_row)}")
#                         continue  

#                     try:
#                         box = Box(
#                             box_number=row['box_number'].strip(),
#                             box_type=row['box_type'].strip(),
#                             box_location=row['box_location'].strip(),
#                             insurance_company=row['insurance_company'].strip() if row['insurance_company'] else None,
#                         )
                        
#                         box.full_clean()  
#                         valid_boxes.append(box)
#                     except ValidationError as e:
#                         for field, messages in e.message_dict.items():
#                             errors_in_row.extend(messages)
#                         error_details.append(f"Row with the number = {row_number}: {', '.join(errors_in_row)}")
#                     except Exception as e:
#                         error_details.append(f"Row {row_number}: {str(e)}")

#                 if error_details:
#                     if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                         return JsonResponse({"success": False, "errors": error_details}, status=400)
#                     return render(request, 'arcapp/bulk_insert/upload.html', 
#                                 {'error_details': error_details})

#                 Box.objects.bulk_create(valid_boxes)

#             if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                 return JsonResponse({"success": True, "message": f"Successfully uploaded {len(valid_boxes)} boxes!"})
#             return render(request, 'arcapp/bulk_insert/upload.html', 
#                         {'success_message': f"Successfully uploaded {len(valid_boxes)} boxes!"})
            
#         except Exception as e:
#             if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                 return JsonResponse({"success": False, "message": f"An error occurred: {str(e)}"}, status=500)
#             return render(request, 'arcapp/bulk_insert/upload.html', 
#                         {'error_message': f"An error occurred: {str(e)}"})

#     return render(request, 'arcapp/bulk_insert/upload.html')


@permission_required('arcapp.can_create_box', raise_exception=True)
def bulk_insert_box(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']

        if not csv_file.name.endswith('.csv'):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({"success": False, "message": "Please upload a CSV file."}, status=400)
            return render(request, 'arcapp/bulk_insert/upload.html', 
                        {'error_message': "Please upload a CSV file."})

        try:
            # Handle BOM and decode with proper encoding
            data = csv_file.read().decode('utf-8-sig').splitlines()  
            reader = csv.DictReader(data, delimiter=',')
            
            # Clean column names (remove BOM and whitespace)
            csv_columns = [col.strip().replace('\ufeff', '') for col in reader.fieldnames]  # ADDED
            
            required_columns = ['box_number', 'box_type', 'box_location', 'insurance_company']
            
            # Debug: Print actual columns received
            print("Detected CSV columns:", csv_columns)  # DEBUGGING
            
            if not all(col in csv_columns for col in required_columns):
                missing = [col for col in required_columns if col not in csv_columns]
                error_msg = f"Missing columns: {', '.join(missing)}"  # MORE SPECIFIC ERROR
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({"success": False, "message": error_msg}, status=400)
                return render(request, 'arcapp/bulk_insert/upload.html', 
                            {'error_message': error_msg})

            error_details = []
            valid_boxes = []
            row_number = 1  # Start counting from header row + 1

            with transaction.atomic():
                for row in reader:
                    row_number += 1
                    errors_in_row = []

                    # Clean all values (handle non-breaking spaces)
                    cleaned_row = {
                        key: value.strip().replace('\xa0', ' ').replace('\ufeff', '') 
                        for key, value in row.items()
                    }  # ADDED CLEANING

                    # Check for missing values using cleaned data
                    for col in required_columns:
                        if not cleaned_row.get(col) or cleaned_row[col] == "":
                            errors_in_row.append(f"Missing value in column '{col}'")
                    
                    if errors_in_row:
                        error_details.append(f"Row {row_number}: {', '.join(errors_in_row)}")
                        continue

                    try:
                        # Use cleaned_row instead of raw row
                        box_type_instance = BoxType.objects.get(
                            name=cleaned_row['box_type']
                        )
                        box_location_instance = BoxLocation.objects.get(
                            name=cleaned_row['box_location']
                        )
                        # box_status_instance = BoxStatus.objects.get(
                        #     name=cleaned_row['box_status']
                        # )
                        insurance_company_instance = InsuranceCompany.objects.get(
                            name=cleaned_row['insurance_company']
                        )
                    except BoxType.DoesNotExist as e:
                        errors_in_row.append(f"Box Type not found: {cleaned_row['box_type']}")
                    except BoxLocation.DoesNotExist as e:
                        errors_in_row.append(f"Box Location not found: {cleaned_row['box_location']}")
                    # except BoxStatus.DoesNotExist as e:
                    #     errors_in_row.append(f"Box Status not found: {cleaned_row['box_status']}")
                    except InsuranceCompany.DoesNotExist as e:
                        errors_in_row.append(f"Insurance Company not found: {cleaned_row['insurance_company']}")
                    except Exception as e:
                        errors_in_row.append(f"Unexpected error: {str(e)}")

                    if errors_in_row:
                        error_details.append(f"Row {row_number}: {', '.join(errors_in_row)}")
                        continue

                    try:
                        box = Box(
                            box_number=cleaned_row['box_number'],
                            box_type=box_type_instance,
                            box_location=box_location_instance,
                            #box_status=box_status_instance,
                            insurance_company=insurance_company_instance,
                        )
                        box.full_clean()
                        valid_boxes.append(box)
                    except ValidationError as e:
                        errors_in_row.extend(e.messages)
                        error_details.append(f"Row {row_number}: {', '.join(errors_in_row)}")
                    except Exception as e:
                        error_details.append(f"Row {row_number}: {str(e)}")

                if error_details:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({"success": False, "errors": error_details}, status=400)
                    return render(request, 'arcapp/bulk_insert/upload.html', 
                                {'error_details': error_details})

                Box.objects.bulk_create(valid_boxes)

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({"success": True, "message": f"Uploaded {len(valid_boxes)} boxes!"})
            return render(request, 'arcapp/bulk_insert/upload.html', 
                        {'success_message': f"Uploaded {len(valid_boxes)} boxes!"})
            
        except Exception as e:
            error_msg = f"Critical error: {str(e)}"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({"success": False, "message": error_msg}, status=500)
            return render(request, 'arcapp/bulk_insert/upload.html', 
                        {'error_message': error_msg})

    return render(request, 'arcapp/bulk_insert/upload.html')        
#################### bulk insert provider segre ##############


# def bulk_insert_claims(request):
#     if request.method == 'POST':
#         form = ProviderSegreeDataFile(request.POST, request.FILES)
#         if form.is_valid():
#             try:
#                 # Get the uploaded file from the form
#                 claim_file = request.FILES['csv_file']

#                 # Open the file in text mode with the correct encoding
#                 claim_file_data = claim_file.read().decode('utf-8').splitlines()

#                 # Use csv.DictReader to process the text data
#                 reader = csv.DictReader(claim_file_data)
#                 for row in reader:
#                     boxNumber = row.get('boxNumber')
#                     claim_code = row.get('ClaimCode')
#                     policy_id = row.get('PolicyId')
#                     policy_name = row.get('policyname')
#                     client_id = row.get('ClientId')
#                     member_name = row.get('Member_name')
#                     provider_name = row.get('Provider_name')
#                     receive_date = row.get('Receive_date')
#                     issuance_date = row.get('Issuance_date')
#                     segregation_date = row.get('Segregation_date')
#                     segregation_by = row.get('Segregation_by')
#                     audit_date = row.get('Audit_date')
#                     audit_by = row.get('Audit_by')
#                     box_status = row.get('Box_status')
#                     batch_id = row.get('batchID')
#                     client_name = row.get('Client_name')
#                     claim_scan = row.get('claimscan') == 'True'
#                     deducted_amount = row.get('DeductedAmount')
#                     note = row.get("note")

#                     try:
#                         box = Box.objects.get(box_number=boxNumber)
#                     except Box.DoesNotExist:
#                         return HttpResponse(f"Box with number {boxNumber} not found.")

#                     # Create ProviderSegregation instance
#                     ProviderSegregation.objects.create(
#                         box_number=box,
#                         ClaimCode=claim_code,
#                         PolicyId=policy_id,
#                         policyname=policy_name,
#                         ClientId=client_id,
#                         Member_name=member_name,
#                         Provider_name=provider_name,
#                         Receive_date=receive_date,
#                         Issuance_date=issuance_date,
#                         Segregation_date=segregation_date,
#                         Segregation_by=segregation_by,
#                         Audit_date=audit_date,
#                         Audit_by=audit_by,
#                         Box_status=box_status,
#                         batchID=batch_id,
#                         Client_name=client_name,
#                         claimscan=claim_scan,
#                         DeductedAmount=deducted_amount,
#                         note = note
#                     )
#                 return HttpResponse("Claims inserted successfully!")
#             except Exception as e:
#                 return HttpResponse(f"Error inserting claims: {e}")
#     else:
#         form = ProviderSegreeDataFile()
#     return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', {'form': form})


import csv
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.shortcuts import render
from .models import ProviderSegregation, Box
import datetime


from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import render
from dateutil import parser
from .models import Box, ProviderSegregation, insurance_policy



import csv
from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import permission_required
from django.db import transaction
from dateutil import parser
from .models import ProviderSegregation, Box, insurance_policy, ClaimStatus

@permission_required('arcapp.can_create_provider_segre', raise_exception=True)
def bulk_insert_claims(request):
    """
    Handles bulk insertion of ProviderSegregation claims from a CSV file.
    """
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    context = {} # Initialize context for rendering

    if request.method == 'POST':
        try:
            if 'csv_file' not in request.FILES:
                error_msg = "No file uploaded. Please select a CSV file."
                if is_ajax:
                    return JsonResponse({"success": False, "message": error_msg, "errors": [error_msg]}, status=400)
                context['error_message'] = error_msg
                return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', context)

            csv_file = request.FILES['csv_file']

            if not csv_file.name.lower().endswith('.csv'):
                error_msg = "Invalid file type. Please upload a CSV file."
                if is_ajax:
                    return JsonResponse({"success": False, "message": error_msg, "errors": [error_msg]}, status=400)
                context['error_message'] = error_msg
                return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', context)
            
            try:
                # Handle BOM and decode with proper encoding
                # It's generally better to read the file in binary mode and then decode
                # For InMemoryUploadedFile, .read() gives bytes
                file_content_bytes = csv_file.read()
                try:
                    decoded_file_content = file_content_bytes.decode('utf-8-sig')
                except UnicodeDecodeError:
                    decoded_file_content = file_content_bytes.decode('latin-1') # Fallback encoding
                
                decoded_file_lines = decoded_file_content.splitlines()

            except Exception as enc_e:
                error_msg = f"File encoding error: {str(enc_e)}. Please ensure the file is UTF-8 or Latin-1 encoded."
                if is_ajax: return JsonResponse({"success": False, "message": error_msg, "errors": [error_msg]}, status=400)
                context['error_message'] = error_msg
                return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', context)

            reader = csv.DictReader(decoded_file_lines)
            error_details = []
            created_count = 0
            
            # Normalize actual fieldnames from CSV header (remove BOM, strip whitespace)
            actual_fieldnames = [name.strip().replace('\ufeff', '') for name in reader.fieldnames] if reader.fieldnames else []
            
            # Define expected CSV column headers. These are the keys you'll use with row.get().
            # These should match the *normalized* headers from actual_fieldnames.
            # For simplicity, we'll assume the CSV headers match these keys after normalization.
            expected_csv_headers = [
                'boxNumber', 'ClaimCode', 'PolicyId', 'Member_name', 
                'Provider_name', 'Receive_date', 'Issuance_date', 'Segregation_date', 'Segregation_by',
                'Audit_date', 'Audit_by', 'batchID', 'request_by', 'request_date','claim_status',
                'claimscan', 'DeductedAmount', 'note', 
                'retrieval_date', 'return_date', 'return_by', 'comment' 
            ]
            
            # Check for missing critical headers
            # You can define a subset of absolutely critical headers if not all are mandatory for the file structure.
            critical_headers_for_file_structure = ['boxNumber', 'ClaimCode', 'PolicyId', 'request_by'] # Example
            missing_critical_headers = [header for header in critical_headers_for_file_structure if header not in actual_fieldnames]
            if missing_critical_headers:
                error_msg = f"CSV is missing critical header(s) required for processing: {', '.join(missing_critical_headers)}"
                if is_ajax:
                    return JsonResponse({"success": False, "message": error_msg, "errors": [error_msg]}, status=400)
                context['error_message'] = error_msg
                return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', context)

            claim_code_pattern = re.compile(r"^\d+-\d+$")
            with transaction.atomic():
                for row_number, row_data in enumerate(reader, start=2): # Start row_number from 2 (after header)
                    # Clean all keys (from header) and values from the current row
                    # Ensure keys used for row.get() match the normalized actual_fieldnames
                    row = {key.strip().replace('\ufeff', ''): str(value).strip() if value is not None else '' 
                           for key, value in row_data.items()}
                    
                    current_row_errors = [] # Store errors for the current row

                    try:
                        # --- Fetching Related Instances & Validating Mandatory Fields ---
                        box_number_val = row.get('boxNumber')
                        if not box_number_val: 
                            current_row_errors.append("CSV column 'boxNumber' is required.")
                        else:
                            box_instance = Box.objects.get(box_number=box_number_val)

                        policy_id_val = row.get('PolicyId')
                        if not policy_id_val: 
                            current_row_errors.append("CSV column 'PolicyId' is required.")
                        else:
                            try:
                                policy_instance = insurance_policy.objects.get(policy_id=int(policy_id_val))
                            except ValueError:
                                current_row_errors.append(f"PolicyId '{policy_id_val}' must be a valid integer.")
                        polic = insurance_policy.objects.get(policy_id=int(policy_id_val))
                        po_inc = polic.insurance_company
                        boxin = Box.objects.get(box_number = box_number_val)
                        box_inc = boxin.insurance_company
                        if po_inc != box_inc:
                            current_row_errors.append(f"the box - insurance company \"{box_inc}\" not the same as policy insurance Company \"{po_inc}\" hence please makeu sure from that the valid Insurance Company in both")

                        claim_status_name = row.get('claim_status')
                        if not claim_status_name: 
                            current_row_errors.append("CSV column 'claim_status' is required.")
                        else:
                            claim_status_instance = ClaimStatus.objects.get(name=claim_status_name)
                        
                        audit_by_name = row.get('Audit_by')
                        if not audit_by_name: 
                            current_row_errors.append("CSV column 'Audit_by' is required.")
                        else:
                            audit_by_instance = audit_by.objects.get(name=audit_by_name)
                        
                        # Optional ForeignKey: return_by
                        return_by_name = row.get('return_by')
                        return_by_instance = None
                        if return_by_name: 
                            return_by_instance = return_by_user.objects.get(name=return_by_name)

                        # Optional/Required ForeignKey: Segregation_by (adjust if mandatory)
                        segregation_by_name = row.get('Segregation_by')
                        segregation_by_instance = None 
                        if segregation_by_name: 
                            segregation_by_instance = segregation_by.objects.get(name=segregation_by_name)
                        # else: # If Segregation_by is mandatory from CSV
                        #     current_row_errors.append("CSV column 'Segregation_by' is required.")
                        
                        # Mandatory ForeignKey: request_by
                        request_by_name = row.get('request_by')
                        if not request_by_name: 
                            current_row_errors.append("CSV column 'request_by' is required.")
                        else:
                            request_by_instance = request_by.objects.get(name=request_by_name)

                        # --- Date Parsing Helper ---
                        def parse_date_field(date_str, field_name, is_required=False):
                            if not date_str:
                                if is_required:
                                    # current_row_errors.append(f"CSV column '{field_name}' is required.") # Handled by direct check below
                                    raise ValueError(f"Date for '{field_name}' is required and cannot be empty.")
                                return None # Allow optional date fields to be empty
                            try:
                                return parser.parse(date_str).date()
                            except (ValueError, TypeError) as e:
                                raise ValueError(f"Invalid date format for {field_name} ('{date_str}'). Expected YYYY-MM-DD or similar. Error: {e}")

                        # Mandatory Dates
                        receive_date = parse_date_field(row.get('Receive_date'), 'Receive_date', is_required=True)
                        issuance_date = parse_date_field(row.get('Issuance_date'), 'Issuance_date', is_required=True)
                        segregation_date = parse_date_field(row.get('Segregation_date'), 'Segregation_date', is_required=True)
                        audit_date = parse_date_field(row.get('Audit_date'), 'Audit_date', is_required=True)
                        request_date_val = parse_date_field(row.get('request_date'), 'request_date', is_required=True)

                        # Optional Dates
                        retrieval_date = parse_date_field(row.get('retrieval_date'), 'retrieval_date')
                        return_date = parse_date_field(row.get('return_date'), 'return_date')

                        # --- Other Fields ---
                        deducted_amount_str = row.get('DeductedAmount')
                        deducted_amount = None
                        if deducted_amount_str:
                            try:
                                deducted_amount = Decimal(deducted_amount_str)
                            except InvalidOperation:
                                current_row_errors.append(f"Invalid number for DeductedAmount: '{deducted_amount_str}'.")
                        
                        claim_code_val = row.get('ClaimCode')
                        if not claim_code_val: 
                            current_row_errors.append("CSV column 'ClaimCode' is required.")

                        elif not claim_code_pattern.fullmatch(claim_code_val):
                            current_row_errors.append(f"you are trying to enter {claim_code_val}invalid format for the claim code it must to be like 181013-227...")

                        claimscan_str = row.get('claimscan', 'false').lower() # Default to 'false' if missing
                        claimscan_bool = claimscan_str in ['true', '1', 'yes', 'on', 't']


                        # If any errors were collected during field processing for this row, add them and skip creation
                        if current_row_errors:
                            error_details.append(f"Row {row_number}: {'; '.join(current_row_errors)}")
                            continue # Skip to the next row

                        # --- Create ProviderSegregation Instance ---
                        # This part is reached only if all previous checks for this row passed
                        ProviderSegregation.objects.create(
                            box_number=box_instance,
                            ClaimCode=claim_code_val,
                            PolicyId=policy_instance,
                            ClientId=policy_instance.client_id if policy_instance else None,
                            Member_name=row.get('Member_name',''),
                            Provider_name=row.get('Provider_name',''),
                            Receive_date=receive_date,
                            Issuance_date=issuance_date,
                            Segregation_date=segregation_date,
                            Segregation_by=segregation_by_instance,
                            Audit_date=audit_date,
                            Audit_by=audit_by_instance,
                            batchID=row.get('batchID',''),
                            request_by=request_by_instance,
                            request_date=request_date_val,
                            Client_name=policy_instance.client_name if policy_instance else None,
                            claimscan=claimscan_bool,
                            DeductedAmount=deducted_amount,
                            note=row.get('note','') or None,
                            claim_status=claim_status_instance,
                            retrieval_date=retrieval_date,
                            return_date=return_date,
                            return_by=return_by_instance,
                            comment=row.get('comment','') or None,
                            created_by=request.user 
                        )
                        created_count += 1

                    # Specific DoesNotExist exceptions for related models
                    except Box.DoesNotExist:
                        error_details.append(f"Row {row_number}: Box '{row.get('boxNumber')}' not found in the database.")
                    except insurance_policy.DoesNotExist:
                        error_details.append(f"Row {row_number}: Policy with ID '{row.get('PolicyId')}' not found.")
                    except ClaimStatus.DoesNotExist:
                        error_details.append(f"Row {row_number}: Claim Status '{row.get('claim_status')}' not found in Master Data.")
                    except audit_by.DoesNotExist:
                        error_details.append(f"Row {row_number}: Audit By User '{row.get('Audit_by')}' not found in Master Data.")
                    except return_by_user.DoesNotExist:
                        error_details.append(f"Row {row_number}: Return By User '{row.get('return_by')}' not found in Master Data.")
                    except segregation_by.DoesNotExist:
                        error_details.append(f"Row {row_number}: Segregation User '{row.get('Segregation_by')}' not found in Master Data.")
                    except request_by.DoesNotExist:
                        error_details.append(f"Row {row_number}: Request By User '{row.get('request_by')}' not found in Master Data.")
                    
                    # General data validation errors (e.g., invalid date format not caught by helper, non-numeric PolicyId)
                    except ValueError as ve: # Catches ValueErrors raised manually or by int(), Decimal(), date parsing
                        error_details.append(f"Row {row_number}: Data validation error. {str(ve)}")
                    except InvalidOperation as ioe: # For Decimal conversion errors if not caught earlier
                        error_details.append(f"Row {row_number}: Invalid number for DeductedAmount. {str(ioe)}")
                    except IntegrityError as ie: # e.g., duplicate unique ClaimCode
                        error_details.append(f"Row {row_number} (ClaimCode: {row.get('ClaimCode','N/A')}): Database integrity error (likely duplicate ClaimCode): {str(ie)}")
                    except Exception as e:
                        # Catch any other unexpected error for this row
                        error_details.append(f"Row {row_number} (ClaimCode: {row.get('ClaimCode','N/A')}): An unexpected error occurred: {str(e)} (Type: {type(e).__name__})")
                
                if error_details: # If any row had errors during the loop, rollback the entire transaction
                    transaction.set_rollback(True) # Ensure atomicity; no records are created if any row fails
                    if is_ajax:
                        return JsonResponse({"success": False, "message": "Upload failed. Some rows had validation errors.", "errors": error_details, "created_count": 0}, status=400)
                    context['error_details'] = error_details
                    context['created_count'] = 0 # No records were actually committed
                    return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', context)

            # If loop completes without setting rollback (i.e., no errors in any row)
            success_msg = f"Successfully uploaded and created {created_count} claims!"
            if is_ajax:
                return JsonResponse({"success": True, "message": success_msg, "created_count": created_count})
            context['success_message'] = success_msg
            context['created_count'] = created_count
            return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', context)

        except Exception as e: # Catch broader exceptions like CSV parsing issues, file not found after initial check etc.
            error_msg = f"A critical error occurred during file processing or setup: {str(e)}"
            if is_ajax:
                return JsonResponse({"success": False, "message": error_msg, "errors": [error_msg]}, status=500)
            context['error_message'] = error_msg
            return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', context)

    return render(request, 'arcapp/bulk_insert/bulk_insert_claims.html', context)

##################################insert reimbursment bulk ##########################################



@permission_required('arcapp.can_create_rein_segre', raise_exception=True)
def reimbursement_insert_claims(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == 'POST':
        try:
            if 'csv_file' not in request.FILES:
                error_msg = "No file uploaded."
                if is_ajax:
                    return JsonResponse({"success": False, "message": error_msg}, status=400)
                return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

            csv_file = request.FILES['csv_file']

            if not csv_file.name.endswith('.csv'):
                error_msg = "Please upload a CSV file."
                if is_ajax:
                    return JsonResponse({"success": False, "message": error_msg}, status=400)
                return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

            data = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(data)
            error_details = []
            created_count = 0

            required_columns = ['boxNumber', 'Claim_Code', 'Batch_num', 'Batch_type', 
                               'English_name', 'Arab_name', 'Payer', 'Policy', 'Hof',
                               'Audit_user', 'audit_date']
            
            if not all(col in reader.fieldnames for col in required_columns):
                error_msg = "CSV is missing required columns."
                if is_ajax:
                    return JsonResponse({"success": False, "message": error_msg}, status=400)
                return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

            with transaction.atomic():
                for row_number, row in enumerate(reader, start=2):
                    try:
                        box_number = row.get('boxNumber')
                        try:
                            box = Box.objects.get(box_number=box_number)
                        except Box.DoesNotExist:
                            error_details.append(f"Row {row_number}: Box {box_number} not found")
                            continue

                        try:
                            audit_date = parser.parse(row.get('audit_date')).date()
                        except Exception as e:
                            error_details.append(f"Row {row_number}: Invalid audit date format")
                            continue

                        ReimbursementSegregation.objects.create(
                            box_number=box,
                            Claim_Code=row.get('Claim_Code'),
                            Batch_num=row.get('Batch_num'),
                            Batch_type=row.get('Batch_type'),
                            English_name=row.get('English_name'),
                            Arab_name=row.get('Arab_name'),
                            Payer=row.get('Payer'),
                            Policy=row.get('Policy'),
                            Hof=row.get('Hof'),
                            Audit_user=row.get('Audit_user'),
                            Audit_date=audit_date,
                        )
                        created_count += 1

                    except Exception as e:
                        error_details.append(f"Row {row_number}: {str(e)}")

                if error_details:
                    if is_ajax:
                        return JsonResponse({"success": False, "errors": error_details}, status=400)
                    return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_details': error_details})

            success_msg = f"Successfully uploaded {created_count} reimbursement claims!"
            if is_ajax:
                return JsonResponse({"success": True, "message": success_msg})
            return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'success_message': success_msg})

        except Exception as e:
            error_msg = f"An error occurred: {str(e)}"
            if is_ajax:
                return JsonResponse({"success": False, "message": error_msg}, status=500)
            return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

    return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html')






import json
import requests
from django.shortcuts import render
from django.http import JsonResponse
from django.db import transaction
from dateutil import parser

# Make sure to import your Box model
from .models import Box

# It's good practice to store the API endpoint in your settings.py
from django.conf import settings

# def reimbursement_insert_claims(request):
#     """
#     Handles the bulk upload of a CSV file containing reimbursement claims,
#     parses it, and sends the data to the .NET API endpoint for processing.
#     """
#     is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
#     # Get the API URL from Django settings for better configuration management.
#     # Example in settings.py: DOTNET_API_URL = "http://localhost:5000"
#     #api_url = f"{settings.DOTNET_API_URL}/api/ReimbursementApi"
#     api_url = 'http://localhost:5033/api/ReimbursementApi'
#     if request.method == 'POST':
#         if 'csv_file' not in request.FILES:
#             error_msg = "No file was uploaded."
#             return JsonResponse({"success": False, "message": error_msg}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#         csv_file = request.FILES['csv_file']
#         if not csv_file.name.endswith('.csv'):
#             error_msg = "Invalid file format. Please upload a CSV file."
#             return JsonResponse({"success": False, "message": error_msg}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#         try:
#             data = csv_file.read().decode('utf-8').splitlines()
#             reader = csv.DictReader(data)
            
#             # The column names in your CSV must match the keys used here.
#             required_columns = ['boxNumber', 'Claim_Code', 'Batch_num', 'Batch_type', 
#                                 'English_name', 'Arab_name', 'Payer', 'Policy', 'Hof',
#                                 'Audit_user', 'audit_date']

#             if not all(col in reader.fieldnames for col in required_columns):
#                 missing_cols = set(required_columns) - set(reader.fieldnames)
#                 error_msg = f"CSV is missing required columns: {', '.join(missing_cols)}"
#                 return JsonResponse({"success": False, "message": error_msg}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#             reimbursements_payload = []
#             error_details = []

#             for row_number, row in enumerate(reader, start=2):
#                 try:
#                     box_number_val = row.get('boxNumber')
#                     box_instance = Box.objects.get(box_number=box_number_val)
                    
#                     # Ensure the audit_date is in the correct format (ISO 8601 is robust)
#                     audit_date = parser.parse(row.get('audit_date')).isoformat()

#                     # Map CSV columns to the C# Reimbursement model properties.
#                     # Note the camelCase convention for JSON keys.
#                     reimbursement_data = {
#                         "boxNumberId": box_instance.id,
#                         "claimCode": row.get('Claim_Code'),
#                         "batchNumber": row.get('Batch_num'),
#                         "batchType": row.get('Batch_type'),
#                         "englishName": row.get('English_name'),
#                         "arabicName": row.get('Arab_name'),
#                         "payer": row.get('Payer'),
#                         "policy": row.get('Policy'),
#                         "hof": row.get('Hof'),
#                         "auditUser": row.get('Audit_user'),
#                         "auditDate": audit_date,
#                     }
#                     reimbursements_payload.append(reimbursement_data)

#                 except Box.DoesNotExist:
#                     error_details.append(f"Row {row_number}: Box with number '{box_number_val}' was not found.")
#                 except Exception as e:
#                     error_details.append(f"Row {row_number}: An error occurred while processing the row - {str(e)}")

#             if error_details:
#                 return JsonResponse({"success": False, "errors": error_details}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_details': error_details})

#             # Send the entire list of records to the .NET API
#             if not reimbursements_payload:
#                  error_msg = "No valid records to send to the API."
#                  return JsonResponse({"success": False, "message": error_msg}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#             headers = {'Content-Type': 'application/json'}
#             response = requests.post(api_url, data=json.dumps(reimbursements_payload), headers=headers, timeout=120)

#             # Check the response from the API
#             response.raise_for_status()  # Raises an HTTPError for bad responses (4xx or 5xx)

#             api_response = response.json()
#             success_msg = api_response.get("message", f"Successfully sent {len(reimbursements_payload)} records.")
            
#             if is_ajax:
#                 return JsonResponse({"success": True, "message": success_msg})
#             return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'success_message': success_msg})

#         except requests.exceptions.RequestException as e:
#             error_msg = f"API connection error: {str(e)}"
#             status_code = 503 # Service Unavailable
#         except Exception as e:
#             error_msg = f"An unexpected error occurred: {str(e)}"
#             status_code = 500

#         if is_ajax:
#             return JsonResponse({"success": False, "message": error_msg}, status=status_code)
#         return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#     return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html')




############## phase 2 processing django/dotnet ###############


# @permission_required('arcapp.can_create_rein_segre', raise_exception=True)
# def reimbursement_insert_claims(request):
#     """
#     Handles the bulk upload of a CSV file containing reimbursement claims,
#     parses it, and sends the data to the .NET API endpoint for processing.
#     """
#     is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
#     # Get the API URL from Django settings for better configuration management.
#     api_url = 'http://localhost:5033/api/ReimbursementApi'
    
#     if request.method == 'POST':
#         if 'csv_file' not in request.FILES:
#             error_msg = "No file was uploaded."
#             return JsonResponse({"success": False, "message": error_msg}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#         csv_file = request.FILES['csv_file']
#         if not csv_file.name.endswith('.csv'):
#             error_msg = "Invalid file format. Please upload a CSV file."
#             return JsonResponse({"success": False, "message": error_msg}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#         try:
#             data = csv_file.read().decode('utf-8').splitlines()
#             reader = csv.DictReader(data)
            
#             required_columns = ['boxNumber', 'Claim_Code', 'Batch_num', 'Batch_type', 
#                                 'English_name', 'Arab_name', 'Payer', 'Policy', 'Hof',
#                                 'Audit_user', 'audit_date']

#             if not all(col in reader.fieldnames for col in required_columns):
#                 missing_cols = set(required_columns) - set(reader.fieldnames)
#                 error_msg = f"CSV is missing required columns: {', '.join(missing_cols)}"
#                 return JsonResponse({"success": False, "message": error_msg}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#             reimbursements_payload = []
#             error_details = []

#             for row_number, row in enumerate(reader, start=2):
#                 try:
#                     box_number_val = row.get('boxNumber')
#                     box_instance = Box.objects.get(box_number=box_number_val)
                    
#                     audit_date = parser.parse(row.get('audit_date')).isoformat()

#                     reimbursement_data = {
#                         "boxNumberId": box_instance.id,
#                         "claimCode": row.get('Claim_Code'),
#                         "batchNumber": row.get('Batch_num'),
#                         "batchType": row.get('Batch_type'),
#                         "englishName": row.get('English_name'),
#                         "arabicName": row.get('Arab_name'),
#                         "payer": row.get('Payer'),
#                         "policy": row.get('Policy'),
#                         "hof": row.get('Hof'),
#                         "auditUser": row.get('Audit_user'),
#                         "auditDate": audit_date,
#                     }
#                     reimbursements_payload.append(reimbursement_data)

#                 except Box.DoesNotExist:
#                     error_details.append(f"Row {row_number}: Box with number '{box_number_val}' was not found.")
#                 except Exception as e:
#                     error_details.append(f"Row {row_number}: An error occurred while processing the row - {str(e)}")

#             if error_details:
#                 return JsonResponse({"success": False, "errors": error_details}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_details': error_details})

#             if not reimbursements_payload:
#                  error_msg = "No valid records to send to the API."
#                  return JsonResponse({"success": False, "message": error_msg}, status=400) if is_ajax else render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#             # IMPROVED: Process in batches and with better connection handling
#             batch_size = 100  # Process 100 records at a time
#             total_processed = 0
            
#             for i in range(0, len(reimbursements_payload), batch_size):
#                 batch = reimbursements_payload[i:i + batch_size]
                
#                 # Create a new session for each batch to avoid connection issues
#                 session = requests.Session()
#                 session.headers.update({'Content-Type': 'application/json'})
                
#                 # Configure session with connection pooling and retry strategy
#                 adapter = requests.adapters.HTTPAdapter(
#                     pool_connections=1,
#                     pool_maxsize=1,
#                     max_retries=requests.adapters.Retry(
#                         total=3,
#                         backoff_factor=1,
#                         status_forcelist=[500, 502, 503, 504]
#                     )
#                 )
#                 session.mount('http://', adapter)
#                 session.mount('https://', adapter)
                
#                 try:
#                     # Increased timeout and added keep-alive handling
#                     response = session.post(
#                         api_url, 
#                         data=json.dumps(batch), 
#                         timeout=(30, 300),  # (connection timeout, read timeout)
#                         headers={
#                             'Content-Type': 'application/json',
#                             'Connection': 'close'  # Prevent keep-alive issues
#                         }
#                     )
                    
#                     response.raise_for_status()
#                     total_processed += len(batch)
                    
#                 except requests.exceptions.Timeout:
#                     error_msg = f"Request timed out while processing batch {i//batch_size + 1}. Processed {total_processed} records successfully."
#                     break
#                 except requests.exceptions.ConnectionError as e:
#                     error_msg = f"Connection error while processing batch {i//batch_size + 1}: {str(e)}. Processed {total_processed} records successfully."
#                     break
#                 except requests.exceptions.RequestException as e:
#                     error_msg = f"API error while processing batch {i//batch_size + 1}: {str(e)}. Processed {total_processed} records successfully."
#                     break
#                 finally:
#                     session.close()
            
#             else:
#                 # This executes if the loop completed without breaking
#                 success_msg = f"Successfully processed all {total_processed} records in {len(reimbursements_payload)//batch_size + 1} batches."
                
#                 if is_ajax:
#                     return JsonResponse({"success": True, "message": success_msg})
#                 return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'success_message': success_msg})

#         except requests.exceptions.RequestException as e:
#             error_msg = f"API connection error: {str(e)}"
#             status_code = 503
#         except Exception as e:
#             error_msg = f"An unexpected error occurred: {str(e)}"
#             status_code = 500

#         if is_ajax:
#             return JsonResponse({"success": False, "message": error_msg}, status=status_code)
#         return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html', {'error_message': error_msg})

#     return render(request, 'arcapp/bulk_insert/reimbursement_insert_bulk.html')

from .utils.bulk_update_box import BoxProcessor
from django.http import JsonResponse

@permission_required('arcapp.can_update_box', raise_exception=True)
def bulk_update_boxes_csv(request):
    if request.method == "POST":
        form = bub(request.POST, request.FILES)
        if form.is_valid():
            processor = BoxProcessor()
            updated_count, errors = processor.process_csv(request.FILES["csv_file"])
            
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    "success": updated_count > 0 and not errors,
                    "updated_count": updated_count,
                    "errors": errors if errors else None
                })
            
            
            if updated_count:
                messages.success(request, f"Updated {updated_count} boxes successfully")
            if errors:
                for error in errors:
                    messages.error(request, error)
            
            return redirect("bulk_update_boxes")
    
    return render(request, "arcapp/bulk_updates/butlk_update_boxes.html", {
        "form": bub()
    })

############# clean code bulk update provider segregation #######################


from django.shortcuts import render, redirect
from .forms import bupsegre

# def bulk_update_provider_segregation(request):
#     if request.method == "POST":
#         form = bupsegre(request.POST, request.FILES)
#         if form.is_valid():
#             csv_file = request.FILES["csv_file"]
#             updated_count, errors = process_provider_segregation_csv(csv_file, request)

#             if updated_count > 0:
#                 messages.success(request, f"Updated {updated_count} records")
#             if errors:
#                 for error in errors:
#                     messages.error(request, error)
            
#             return redirect("bulk_update_provider_segregation")
#     else:
#         form = bupsegre()

#     return render(request, "arcapp/bulk_updates/bulk_update_provider_segregation.html", {"form": form})


from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib import messages
# from .utils.provider_bulk_update_utils import process_provider_segregation_csv


# @require_http_methods(["GET", "POST"])
# def bulk_update_provider_segregation(request):
#     is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
#     if request.method == 'POST':
#         try:
#             if 'csv_file' not in request.FILES:
#                 error_msg = "No file uploaded."
#                 if is_ajax:
#                     return JsonResponse({"success": False, "message": error_msg}, status=400)
#                 messages.error(request, error_msg)
#                 return redirect("bulk_update_provider_segregation")

#             csv_file = request.FILES['csv_file']
            
#             if not csv_file.name.endswith('.csv'):
#                 error_msg = "Please upload a CSV file."
#                 if is_ajax:
#                     return JsonResponse({"success": False, "message": error_msg}, status=400)
#                 messages.error(request, error_msg)
#                 return redirect("bulk_update_provider_segregation")

#             updated_count, errors = process_provider_segregation_csv(csv_file, request)
            
#             if errors:
#                 if is_ajax:
#                     return JsonResponse({
#                         "success": False,
#                         "message": f"Completed with {len(errors)} errors",
#                         "errors": errors
#                     }, status=207)
#                 for error in errors:
#                     messages.error(request, error)
            
#             success_msg = f"Successfully updated {updated_count} records"
#             if is_ajax:
#                 return JsonResponse({
#                     "success": True,
#                     "message": success_msg,
#                     "updated_count": updated_count
#                 })
#             messages.success(request, success_msg)
            
#         except Exception as e:
#             error_msg = f"An error occurred: {str(e)}"
#             if is_ajax:
#                 return JsonResponse({"success": False, "message": error_msg}, status=500)
#             messages.error(request, error_msg)
        
#         return redirect("bulk_update_provider_segregation")
    
#     return render(request, "arcapp/bulk_updates/bulk_update_provider_segregation.html")

def process_provider_segregation_csv(csv_file, request):
    updated_count = 0
    errors = []
    
    try:
        
        decoded_file = csv_file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded_file)
        
        
        headers = reader.fieldnames
        if 'ClaimCode' not in headers:
            return 0, ["CSV file must contain a 'ClaimCode' column"]
        
        for row_num, row in enumerate(reader, start=2):  
            try:
                
                claim_code = row.get('ClaimCode', '').strip()
                if not claim_code:
                    errors.append(f"Row {row_num}: Missing ClaimCode")
                    continue
                
                
                try:
                    
                    record = ProviderSegregation.objects.get(ClaimCode=claim_code)
                except ProviderSegregation.DoesNotExist:
                    errors.append(f"Row {row_num}: Record with ClaimCode '{claim_code}' not found")
                    continue
                
                
                fields_updated = False
                
                
                for field_name, value in row.items():
                    
                    if field_name == 'ClaimCode':
                        continue
                    
                    
                    if value is None or value.strip() == '':
                        continue
                    
                    
                    if hasattr(record, field_name):
                        try:
                            
                            field_type = type(getattr(record, field_name))
                            
                            
                            clean_value = value.strip()
                            if field_type == bool:
                                
                                clean_value = clean_value.lower() in ['true', 'yes', '1', 't', 'y']
                            elif field_type == int:
                                clean_value = int(clean_value)
                            elif field_type == float:
                                clean_value = float(clean_value)
                            elif field_type == datetime.date:
                                
                                try:
                                    clean_value = datetime.datetime.strptime(clean_value, '%d/%m/%Y').date()
                                except ValueError:
                                    try:
                                        clean_value = datetime.datetime.strptime(clean_value, '%Y-%m-%d').date()
                                    except ValueError:
                                        raise ValueError(f"Invalid date format for {field_name}")
                            
                            
                            setattr(record, field_name, clean_value)
                            fields_updated = True
                            
                        except Exception as e:
                            errors.append(f"Row {row_num}: Error setting {field_name}={value}: {str(e)}")
                    else:
                        errors.append(f"Row {row_num}: Unknown field '{field_name}'")
                
                
                if fields_updated:
                    record.save()
                    updated_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
    
    except Exception as e:
        errors.append(f"File processing error: {str(e)}")
    
    return updated_count, errors





import csv
from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import permission_required
from django.db import transaction
from dateutil import parser
from .models import ProviderSegregation, Box, insurance_policy, ClaimStatus


from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import permission_required
from django.db import transaction, IntegrityError
from decimal import Decimal, InvalidOperation
from dateutil import parser
import csv
from .models import ProviderSegregation, Box, insurance_policy, ClaimStatus, audit_by, return_by_user, segregation_by, request_by

@permission_required('arcapp.can_update_provider_segre', raise_exception=True)
def bulk_update_provider(request):
    if request.method == 'POST':
        try:
            if 'csv_file' not in request.FILES:
                return JsonResponse({"success": False, "message": "No file uploaded."}, status=400)

            csv_file = request.FILES['csv_file']
            if not csv_file.name.lower().endswith(('.csv', '.xlsx', '.xls')):
                return JsonResponse({"success": False, "message": "Please upload a CSV or Excel file."}, status=400)

            try:
                decoded_file = csv_file.read().decode('utf-8-sig').splitlines()
            except UnicodeDecodeError:
                try:
                    csv_file.seek(0)
                    decoded_file = csv_file.read().decode('latin-1').splitlines()
                except Exception as enc_e:
                    return JsonResponse({"success": False, "message": f"File encoding error: {str(enc_e)}. Please use UTF-8 or Latin-1 encoding."}, status=400)

            reader = csv.DictReader(decoded_file)
            
            actual_fieldnames = reader.fieldnames if reader.fieldnames else []
            if 'ClaimCode' not in actual_fieldnames:
                return JsonResponse({
                    "success": False, 
                    "message": "Invalid file format. The CSV must contain a 'ClaimCode' column."
                }, status=400)

            errors = []
            updated_count = 0
            processed_rows = 0

            with transaction.atomic():
                for row_number, row in enumerate(reader, start=2):
                    processed_rows += 1
                    claim_code = row.get('ClaimCode', '').strip()
                    
                    if not claim_code:
                        errors.append(f"Row {row_number}: 'ClaimCode' is required but was not found or is empty.")
                        continue

                    try:
                        provider = ProviderSegregation.objects.select_related('box_number', 'PolicyId').get(ClaimCode=claim_code)
                        has_updates_for_row = False

                        # --- Start of new validation logic ---
                        new_box_number_str = row.get('box_number', '').strip()
                        new_policy_id_str = row.get('PolicyId', '').strip()

                        if new_box_number_str:
                            try:
                                new_box = Box.objects.get(box_number=new_box_number_str)
                                if provider.PolicyId and new_box.insurance_company != provider.PolicyId.insurance_company:
                                    errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Insurance company mismatch. The new box's insurance company does not match the policy's insurance company.")
                                    continue
                                if provider.box_number != new_box:
                                    provider.box_number = new_box
                                    has_updates_for_row = True
                            except Box.DoesNotExist:
                                errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Box with number '{new_box_number_str}' not found.")
                                continue
                        
                        if new_policy_id_str:
                            try:
                                new_policy = insurance_policy.objects.get(policy_id=int(new_policy_id_str))
                                if provider.box_number and new_policy.insurance_company != provider.box_number.insurance_company:
                                     errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Insurance company mismatch. The new policy's insurance company does not match the box's insurance company.")
                                     continue
                                if provider.PolicyId != new_policy:
                                    provider.PolicyId = new_policy
                                    provider.Client_name = new_policy.client_name if hasattr(new_policy, 'client_name') else None
                                    provider.ClientId = new_policy.client_id if hasattr(new_policy, 'client_id') else None
                                    has_updates_for_row = True
                            except insurance_policy.DoesNotExist:
                                errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Policy with ID '{new_policy_id_str}' not found.")
                                continue
                            except (ValueError, TypeError):
                                errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Invalid format for PolicyId '{new_policy_id_str}'. Must be an integer.")
                                continue

                        # --- End of new validation logic ---


                        fk_fields_to_update = {
                            'claim_status': (ClaimStatus, 'name'),
                            'Audit_by': (audit_by, 'name'),
                            'return_by': (return_by_user, 'name'),
                            'Segregation_by' : (segregation_by, 'name'),
                            'request_by' : (request_by, 'name')
                        }

                        for model_field_name, (RelatedModel, lookup_field) in fk_fields_to_update.items():
                            if model_field_name in row:
                                fk_val_str = row.get(model_field_name, '').strip()
                                current_fk_instance = getattr(provider, model_field_name)
                                
                                if fk_val_str: 
                                    new_fk_instance = None
                                    try:
                                        new_fk_instance = RelatedModel.objects.get(**{lookup_field: fk_val_str})
                                        
                                        if current_fk_instance != new_fk_instance:
                                            setattr(provider, model_field_name, new_fk_instance)
                                            has_updates_for_row = True
                                    except RelatedModel.DoesNotExist:
                                        errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Related {RelatedModel.__name__} with {lookup_field} '{fk_val_str}' not found for field '{model_field_name}'.")
                                        if not ProviderSegregation._meta.get_field(model_field_name).null:
                                            raise ValueError(f"Required related object for {model_field_name} not found.")
                                        continue 
                                    except ValueError:
                                        errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Invalid ID format for {model_field_name}: '{fk_val_str}'.")
                                        if not ProviderSegregation._meta.get_field(model_field_name).null:
                                            raise ValueError(f"Invalid ID for required {model_field_name}.")
                                        continue
                        
                        date_fields = ['Receive_date', 'Issuance_date', 'Segregation_date', 'Audit_date', 'request_date', 'retrieval_date', 'return_date']
                        for field_name in date_fields:
                            if field_name in row:
                                date_str = row.get(field_name, '').strip()
                                current_db_value = getattr(provider, field_name)
                                model_field_object = ProviderSegregation._meta.get_field(field_name)

                                if date_str: 
                                    try:
                                        parsed_date = parser.parse(date_str).date()
                                        if current_db_value != parsed_date:
                                            setattr(provider, field_name, parsed_date)
                                            has_updates_for_row = True
                                    except (ValueError, TypeError):
                                        errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Invalid date format for {field_name}: '{date_str}'. Expected YYYY-MM-DD.")
                                        if not model_field_object.null: 
                                            raise ValueError(f"Invalid date format for required field {field_name} ('{date_str}').")
                                        continue
                        
                        simple_text_fields = ['Member_name', 'Provider_name', 'batchID', 'note', 'comment']
                        for field in simple_text_fields:
                            if field in row:
                                raw_value = row.get(field) 
                                if raw_value is not None : 
                                    value = raw_value.strip()
                                    current_field_value = getattr(provider, field)
                                    if value: 
                                        if str(current_field_value if current_field_value is not None else "") != value:
                                            setattr(provider, field, value)
                                            has_updates_for_row = True

                        if 'DeductedAmount' in row:
                            deducted_amount_str = row.get('DeductedAmount', None)
                            if deducted_amount_str is not None : 
                                deducted_amount_val_clean = deducted_amount_str.strip()
                                if deducted_amount_val_clean: 
                                    try:
                                        new_deducted_amount = Decimal(deducted_amount_val_clean)
                                        if provider.DeductedAmount != new_deducted_amount:
                                            provider.DeductedAmount = new_deducted_amount
                                            has_updates_for_row = True
                                    except InvalidOperation:
                                        errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Invalid number format for DeductedAmount: '{deducted_amount_str}'.")
                                        if not ProviderSegregation._meta.get_field('DeductedAmount').null:
                                            raise ValueError("Invalid format for required DeductedAmount.")
                                        continue

                        if 'claimscan' in row:
                            claimscan_str = row.get('claimscan', None)
                            if claimscan_str is not None and claimscan_str.strip() != "": 
                                new_claimscan_bool = str(claimscan_str).strip().lower() in ['true', '1', 'yes', 'on']
                                if provider.claimscan != new_claimscan_bool:
                                    provider.claimscan = new_claimscan_bool
                                    has_updates_for_row = True
                        
                        if has_updates_for_row:
                            provider.Update_by = request.user if request.user.is_authenticated else None
                            provider.save()
                            updated_count += 1

                    except ProviderSegregation.DoesNotExist:
                        errors.append(f"Row {row_number}: ClaimCode '{claim_code}' not found in the database.")
                    except IntegrityError as ie:
                        errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Integrity error (e.g. unique constraint): {str(ie)}")
                    except ValueError as ve: 
                        errors.append(f"Row {row_number} (ClaimCode: {claim_code}): Data processing error: {str(ve)}")
                    except Exception as e:
                        errors.append(f"Row {row_number} (ClaimCode: {claim_code}): An unexpected error occurred: {str(e)} (Type: {type(e).__name__})")

                if errors: 
                    transaction.set_rollback(True) 
                    return JsonResponse({
                        'success': False,
                        'message': f'Update failed. {len(errors)} error(s) occurred out of {processed_rows} processed row(s). No records were updated.',
                        'errors': errors, 
                        'updated_count': 0,
                        'processed_rows': processed_rows
                    }, status=400)

            return JsonResponse({
                'success': True,
                'message': f'Process complete. Successfully updated {updated_count} out of {processed_rows} processed records.',
                'errors': errors, 
                'updated_count': updated_count,
                'processed_rows': processed_rows
            })

        except Exception as e: 
            return JsonResponse({'success': False, 'message': f'A critical error occurred: {str(e)}'}, status=500)

    return render(request, 'arcapp/bulk_updates/bulk_update_provider_segregation.html')

############### bulk update reimbursement segre ###################


from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from .forms import bupsegre
from .utils.bulk_upate_reimbursement import process_reimbursement_csv

@permission_required('arcapp.can_update_reim_segre', raise_exception=True)
def bulk_update_reimbursement_segregation(request):
    if request.method == "POST":
        form = bupsegre(request.POST, request.FILES)
        if form.is_valid():
            updated_count, errors = process_reimbursement_csv(request.FILES["csv_file"])
            
            # Check if request is AJAX
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            if is_ajax:
                if errors:
                    return JsonResponse({
                        'success': False,
                        'message': f"Encountered {len(errors)} errors during processing.",
                        'errors': errors
                    })
                else:
                    return JsonResponse({
                        'success': True,
                        'message': f"Successfully processed {updated_count} records."
                    })
            else:
                # Handle traditional form submission
                if updated_count:
                    messages.success(request, f"Successfully processed {updated_count} records")
                if errors:
                    messages.error(request, f"Encountered {len(errors)} errors:")
                    for error in errors:
                        messages.error(request, error)
                
                return redirect("bulk_update_reim_segre")
    
    return render(request, "arcapp/bulk_updates/bulk_update_reim_segre.html", {
        "form": bupsegre()
    })
################ bulk delete boxes ####################

from .forms import bdb
# def bulk_delete_boxes(request):
#     if request.method == 'POST':
#         form = bdb(request.POST, request.FILES)
#         if form.is_valid():
#             file = form.cleaned_data['file']
#             try:

#                 box_numbers = []
#                 csv_file = file.read().decode('utf-8').splitlines()
#                 csv_reader = csv.DictReader(csv_file)
                

#                 if 'box_number' not in csv_reader.fieldnames:
#                     form.add_error('file', "CSV file must contain a 'box_number' column.")
#                     return render(request, 'arcapp/bulk_delete/bulk_delete_box.html', {'form': form})
                
#                 for row in csv_reader:
#                     box_number = row.get('box_number')
#                     print(box_number)
#                     if box_number:
#                         box_numbers.append(box_number.strip())
                
#                 deleted_count, _ = Box.objects.filter(box_number__in=box_numbers).delete()
                
#                 return JsonResponse({'message': f'Successfully deleted {deleted_count} boxes.'})
            
#             except Exception as e:
#                 form.add_error('file', f"An error occurred while processing the file: {e}")
#                 print(e)

#         return render(request, 'arcapp/bulk_delete/bulk_delete_box.html', {'form': form})
    

#     form = bdb()
#     return render(request, 'arcapp/bulk_delete/bulk_delete_box.html', {'form': form})



@permission_required('arcapp.can_delete_box', raise_exception=True)
def bulk_delete_boxes(request):
    if request.method == 'POST':
        form = bdb(request.POST, request.FILES)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                box_numbers = []
                csv_file = file.read().decode('utf-8').splitlines()
                csv_reader = csv.DictReader(csv_file)
                
                # Validate the CSV structure
                if 'box_number' not in csv_reader.fieldnames:
                    error_msg = "CSV file must contain a 'box_number' column."
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': error_msg,
                            'errors': [error_msg]
                        })
                    else:
                        form.add_error('file', error_msg)
                        return render(request, 'arcapp/bulk_delete/bulk_delete_box.html', {'form': form})
                
                # Collect box numbers
                errors = []
                for row in csv_reader:
                    box_number = row.get('box_number')
                    if box_number:
                        box_numbers.append(box_number.strip())
                    else:
                        errors.append(f"Missing box number in row {csv_reader.line_num}")
                
                # Perform deletion
                if box_numbers:
                    deleted_count, deletion_details = Box.objects.filter(box_number__in=box_numbers).delete()
                    
                    success_msg = f'Successfully deleted {deleted_count} boxes.'
                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'message': success_msg,
                            'deleted_count': deleted_count
                        })
                    else:
                        messages.success(request, success_msg)
                else:
                    error_msg = "No valid box numbers found in the CSV file."
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': error_msg,
                            'errors': errors
                        })
                    else:
                        form.add_error('file', error_msg)
            
            except Exception as e:
                error_msg = f"An error occurred while processing the file: {str(e)}"
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': error_msg,
                        'errors': [error_msg]
                    })
                else:
                    form.add_error('file', error_msg)
                    print(e)

        else:
            # Form validation errors
            if is_ajax:
                errors = []
                for field, field_errors in form.errors.items():
                    for error in field_errors:
                        errors.append(f"{field}: {error}")
                
                return JsonResponse({
                    'success': False,
                    'message': 'Form validation failed',
                    'errors': errors
                })
        
        # Fall through to render for non-AJAX requests with errors
        if not is_ajax:
            return render(request, 'arcapp/bulk_delete/bulk_delete_box.html', {'form': form})

    # GET request
    form = bdb()
    return render(request, 'arcapp/bulk_delete/bulk_delete_box.html', {'form': form})













############ bulk delete provider segre ###############
@permission_required('arcapp.can_delete_provider_segre', raise_exception=True)
def bulk_delete_provider_segre(request):
    if request.method == "POST" and request.FILES['csv_file']:
        
        csv_file = request.FILES['csv_file']
        
        try:
            
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            csv_reader = csv.reader(decoded_file)

           
            claim_codes_to_delete = [row[0] for row in csv_reader if row]

            
            deleted_count, _ = ProviderSegregation.objects.filter(ClaimCode__in=claim_codes_to_delete).delete()

            
            messages.success(request, f"{deleted_count} records were successfully deleted.")
            return render(request, 'arcapp/bulk_delete/bulk_delete_provider_segre.html')  
            
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return render(request, 'arcapp/bulk_delete/bulk_delete_provider_segre.html')  
    else:
        return render(request, 'arcapp/bulk_delete/bulk_delete_provider_segre.html')


############ bulk delete riem segre ###############

from .forms import bdr
@permission_required('arcapp.can_delete_reim_segre', raise_exception=True)
def bulk_delete_riem_segre(request):
    if request.method == "POST" and request.FILES['csv_file']:

        csv_file = request.FILES['csv_file']
        
        try:

            decoded_file = csv_file.read().decode('utf-8').splitlines()
            csv_reader = csv.reader(decoded_file)


            claim_codes_to_delete = [row[0] for row in csv_reader if row]


            deleted_count, _ = ReimbursementSegregation.objects.filter(Claim_Code__in=claim_codes_to_delete).delete()


            messages.success(request, f"{deleted_count} records were successfully deleted.")
            return render(request, 'arcapp/bulk_delete/bulk_delete_riem_segre.html')  
            
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return render(request, 'arcapp/bulk_delete/bulk_delete_riem_segre.html')  
    else:
        return render(request, 'arcapp/bulk_delete/bulk_delete_riem_segre.html')

              
############### top deducted amount ################

def top_deductions(request):

    top_deductions = (
        ProviderSegregation.objects.order_by('-DeductedAmount')
        .values('ClaimCode', 'DeductedAmount')[:10]
    )


    claim_codes = [item['ClaimCode'] for item in top_deductions]

    deducted_amounts = [float(item['DeductedAmount']) if item['DeductedAmount'] is not None else 0.0 for item in top_deductions]


    return render(request, 'arcapp/data_visualization/top_deductions.html', {
        'claim_codes': claim_codes,
        'deducted_amounts': deducted_amounts,
    })
from django.shortcuts import render
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
from .models import Box, ProviderSegregation, ReimbursementSegregation
import json
from collections import Counter
from datetime import timedelta
from django.utils import timezone
from django.db.models.functions import ExtractYear


def home_dashboard(request):
    top_deductions = (
        ProviderSegregation.objects.order_by('-DeductedAmount')
        .values('ClaimCode', 'DeductedAmount')[:10]
    )
    claim_codes = [item['ClaimCode'] for item in top_deductions]
    deducted_amounts = [float(item['DeductedAmount']) if item['DeductedAmount'] is not None else 0.0 for item in top_deductions]
    





    provider_count = ProviderSegregation.objects.count()
    reimbursement_count = ReimbursementSegregation.objects.count()
    segregation_counts = [provider_count, reimbursement_count]
    







    top_boxes_provider = ProviderSegregation.objects.values('box_number__box_number').annotate(
        count=Count('ClaimCode')
    )
    top_boxes_reimbursement = ReimbursementSegregation.objects.values('box_number__box_number').annotate(
        count=Count('Claim_Code')
    )
    




    box_claim_counts = {}
    for item in top_boxes_provider:
        box_num = item['box_number__box_number']
        box_claim_counts[box_num] = box_claim_counts.get(box_num, 0) + item['count']
    
    for item in top_boxes_reimbursement:
        box_num = item['box_number__box_number']
        box_claim_counts[box_num] = box_claim_counts.get(box_num, 0) + item['count']
    


    top_20_boxes = sorted(box_claim_counts.items(), key=lambda x: x[1], reverse=True)[:20]
    top_boxes = [item[0] for item in top_20_boxes]
    top_boxes_counts = [item[1] for item in top_20_boxes]
    
    # 4. Box Creation Trend
    # Get boxes with their creation dates
    boxes = Box.objects.all().order_by('data_entrance_date')
    
    if boxes.exists():
        # Group boxes by month
        first_date = boxes.first().data_entrance_date.date().replace(day=1)
        last_date = timezone.now().date().replace(day=1)
        
        # Generate all months between first and last date
        months = []
        current_date = first_date
        while current_date <= last_date:
            months.append(current_date.strftime('%b %Y'))
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
        
        # Count boxes created per month
        box_counts_by_month = {}
        for box in boxes:
            month_str = box.data_entrance_date.strftime('%b %Y')
            box_counts_by_month[month_str] = box_counts_by_month.get(month_str, 0) + 1
        
        # Create cumulative counts
        cumulative_count = 0
        cumulative_counts = []
        
        for month in months:
            cumulative_count += box_counts_by_month.get(month, 0)
            cumulative_counts.append(cumulative_count)
        
        box_creation_dates = months
        cumulative_box_counts = cumulative_counts
    else:
        box_creation_dates = []
        cumulative_box_counts = []
    
    # 5. Boxes by Type
    box_types_data = Box.objects.values('box_type').annotate(count=Count('id'))
    box_types = [item['box_type'] for item in box_types_data]
    box_type_counts = [item['count'] for item in box_types_data]
    
    return render(request, 'arcapp/data_visualization/data_dashboard.html', {
        'claim_codes': json.dumps(claim_codes),
        'deducted_amounts': json.dumps(deducted_amounts),
        'segregation_counts': json.dumps(segregation_counts),
        'top_boxes': json.dumps(top_boxes),
        'top_boxes_counts': json.dumps(top_boxes_counts),
        'box_creation_dates': json.dumps(box_creation_dates),
        'cumulative_box_counts': json.dumps(cumulative_box_counts),
        'box_types': json.dumps(box_types),
        'box_type_counts': json.dumps(box_type_counts),
        'user_name': request.user.first_name if request.user.is_authenticated else 'User',

    })


##################### data visualization 2 ######################
def data_visuslization(request):
    # 1. Top Deductions Chart
    top_deductions = (
        ProviderSegregation.objects.order_by('-DeductedAmount')
        .values('ClaimCode', 'DeductedAmount')[:10]
    )
    claim_codes = [item['ClaimCode'] for item in top_deductions]
    deducted_amounts = [float(item['DeductedAmount']) if item['DeductedAmount'] is not None else 0.0 for item in top_deductions]
    
    # 2. Segregation Distribution
    provider_count = ProviderSegregation.objects.count()
    reimbursement_count = ReimbursementSegregation.objects.count()
    segregation_counts = [provider_count, reimbursement_count]
    
    # 3. Top 20 Boxes by Claim Count
    top_boxes_provider = ProviderSegregation.objects.values('box_number__box_number').annotate(
        count=Count('ClaimCode')
    )
    top_boxes_reimbursement = ReimbursementSegregation.objects.values('box_number__box_number').annotate(
        count=Count('Claim_Code')
    )
    
    # Combine and sum counts from both models
    box_claim_counts = {}
    for item in top_boxes_provider:
        box_num = item['box_number__box_number']
        box_claim_counts[box_num] = box_claim_counts.get(box_num, 0) + item['count']
    
    for item in top_boxes_reimbursement:
        box_num = item['box_number__box_number']
        box_claim_counts[box_num] = box_claim_counts.get(box_num, 0) + item['count']
    
    # Sort by count (descending) and take top 20
    top_20_boxes = sorted(box_claim_counts.items(), key=lambda x: x[1], reverse=True)[:20]
    top_boxes = [item[0] for item in top_20_boxes]
    top_boxes_counts = [item[1] for item in top_20_boxes]
    
    # 4. Box Creation Trend
    # Get boxes with their creation dates
    boxes = Box.objects.all().order_by('data_entrance_date')
    
    if boxes.exists():
        # Group boxes by month
        first_date = boxes.first().data_entrance_date.date().replace(day=1)
        last_date = timezone.now().date().replace(day=1)
        
        # Generate all months between first and last date
        months = []
        current_date = first_date
        while current_date <= last_date:
            months.append(current_date.strftime('%b %Y'))
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
        
        # Count boxes created per month
        box_counts_by_month = {}
        for box in boxes:
            month_str = box.data_entrance_date.strftime('%b %Y')
            box_counts_by_month[month_str] = box_counts_by_month.get(month_str, 0) + 1
        
        # Create cumulative counts
        cumulative_count = 0
        cumulative_counts = []
        
        for month in months:
            cumulative_count += box_counts_by_month.get(month, 0)
            cumulative_counts.append(cumulative_count)
        
        box_creation_dates = months
        cumulative_box_counts = cumulative_counts
    else:
        box_creation_dates = []
        cumulative_box_counts = []
    
    # 5. Boxes by Type
    box_types_data = Box.objects.values('box_type').annotate(count=Count('id'))
    box_types = [item['box_type'] for item in box_types_data]
    box_type_counts = [item['count'] for item in box_types_data]
    
    # NEW VISUALIZATIONS
    
    # 1. Top Clients by Claims Count
    top_clients = ProviderSegregation.objects.values('Client_name').annotate(
        count=Count('ClaimCode')
    ).order_by('-count')[:10]
    client_names = [item['Client_name'] for item in top_clients]
    client_claim_counts = [item['count'] for item in top_clients]
    
    # 2. Top Providers by Claims Count
    top_providers = ProviderSegregation.objects.values('Provider_name').annotate(
        count=Count('ClaimCode')
    ).order_by('-count')[:10]
    provider_names = [item['Provider_name'] for item in top_providers]
    provider_claim_counts = [item['count'] for item in top_providers]
    
    # 3. Claims by Year (Receive Date)
    claims_by_year = ProviderSegregation.objects.annotate(
        year=ExtractYear('Receive_date')
    ).values('year').annotate(
        count=Count('ClaimCode')
    ).order_by('year')
    
    claim_years = [item['year'] for item in claims_by_year]
    yearly_claim_counts = [item['count'] for item in claims_by_year]
    
    # 4. Claims by Segregation User
    claims_by_segregator = ProviderSegregation.objects.values('Segregation_by').annotate(
        count=Count('ClaimCode')
    ).order_by('-count')
    
    segregation_users = [item['Segregation_by'] for item in claims_by_segregator]
    segregation_counts_by_user = [item['count'] for item in claims_by_segregator]
    
    # 5. Audits by User
    audits_by_user = ProviderSegregation.objects.values('Audit_by').annotate(
        count=Count('ClaimCode')
    ).order_by('-count')
    
    audit_users = [item['Audit_by'] for item in audits_by_user]
    audit_counts_by_user = [item['count'] for item in audits_by_user]
    
    return render(request, 'arcapp/data_visualization/data_vis2.html', {
        'claim_codes': json.dumps(claim_codes),
        'deducted_amounts': json.dumps(deducted_amounts),
        'segregation_counts': json.dumps(segregation_counts),
        'top_boxes': json.dumps(top_boxes),
        'top_boxes_counts': json.dumps(top_boxes_counts),
        'box_creation_dates': json.dumps(box_creation_dates),
        'cumulative_box_counts': json.dumps(cumulative_box_counts),
        'box_types': json.dumps(box_types),
        'box_type_counts': json.dumps(box_type_counts),
        # New data for additional visualizations
        'client_names': json.dumps(client_names),
        'client_claim_counts': json.dumps(client_claim_counts),
        'provider_names': json.dumps(provider_names),
        'provider_claim_counts': json.dumps(provider_claim_counts),
        'claim_years': json.dumps(claim_years),
        'yearly_claim_counts': json.dumps(yearly_claim_counts),
        'segregation_users': json.dumps(segregation_users),
        'segregation_counts_by_user': json.dumps(segregation_counts_by_user),
        'audit_users': json.dumps(audit_users),
        'audit_counts_by_user': json.dumps(audit_counts_by_user),
    })








########################### claim retrieval ######################

import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import now
from .models import ProviderSegregation
from django.core.exceptions import ValidationError

def claim_retrieval(request):
    if request.method == "POST":
        
        batch_id = request.POST.get("batch_id")
        audit_date_start = request.POST.get("audit_date_start")
        audit_date_end = request.POST.get("audit_date_end")
        audit_by = request.POST.get("audit_by")

        
        claims = ProviderSegregation.objects.all().order_by('-Audit_date')

       
        try:
            if batch_id:
                claims = claims.filter(batchID__icontains=batch_id)
            
            if audit_date_start and audit_date_end:
                claims = claims.filter(
                    Audit_date__gte=audit_date_start,
                    Audit_date__lte=audit_date_end
                )
            elif audit_date_start or audit_date_end:
                raise ValidationError("Both start and end dates are required for date filtering")
            
            if audit_by:
                claims = claims.filter(Audit_by__icontains=audit_by)

        except ValidationError as e:
            return render(request, "arcapp/claim_retrival/claim_retrieval.html", {
                'error_message': str(e)
            })

       
        if not claims.exists():
            return render(request, "arcapp/claim_retrival/claim_retrieval.html", {
                'error_message': "No claims found matching the specified criteria"
            })

        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Claims Data"

        headers = [
            "Box Number", "Claim Code", "Policy ID", "Client ID",
            "Member Name", "Provider Name", "Receive Date", "Issuance Date",
            "Segregation Date", "Segregated By", "Audit Date", "Audited By", "Batch ID", "request_by", "request_date", "Client Name",
            "Claim Scanned", "Deducted Amount", "Claim Status", "note", "retrieval date", 
            "return date", "return by", "comment"
        ]


        ws.append(headers)

        for claim in claims:
            row = [
                str(claim.box_number),
                claim.ClaimCode,
                str(claim.PolicyId),
                #claim.policyname,
                claim.ClientId,
                claim.Member_name,
                claim.Provider_name,
                claim.Receive_date.strftime('%Y-%m-%d') if claim.Receive_date else 'N/A',
                claim.Issuance_date.strftime('%Y-%m-%d') if claim.Issuance_date else 'N/A',
                claim.Segregation_date.strftime('%Y-%m-%d') if claim.Segregation_date else 'N/A',
                claim.Segregation_by,
                claim.Audit_date.strftime('%Y-%m-%d') if claim.Audit_date else 'N/A',
                claim.Audit_by,
                #claim.Box_status,
                claim.batchID,
                claim.request_by,
                claim.request_date.strftime('%Y-%m-%d') if claim.request_date else 'N/A',
                claim.Client_name,
                'Yes' if claim.claimscan else 'No',
                float(claim.DeductedAmount) if claim.DeductedAmount else 0.0,
                str(claim.claim_status),
                claim.note,
                claim.retrieval_date,
                claim.return_date,
                claim.return_by,
                str(claim.comment)
                #claim.data_entrance_date.strftime('%Y-%m-%d %H:%M') if claim.data_entrance_date else 'N/A',
            ]
            ws.append(row)

        
        filename_parts = []
        if batch_id: filename_parts.append(f"batch_{batch_id}")
        if audit_date_start: filename_parts.append(f"from_{audit_date_start}")
        if audit_date_end: filename_parts.append(f"to_{audit_date_end}")
        if audit_by: filename_parts.append(f"auditor_{audit_by}")
        filename = f"claims_{'_'.join(filename_parts)}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    return render(request, "arcapp/claim_retrival/claim_retrieval.html")







########### claim_retrieval update option ##############
import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.utils import timezone 
from django.db import transaction 
from django.contrib.auth.decorators import login_required 
from datetime import datetime
from .models import ProviderSegregation, ClaimStatus 
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages 
from django.db import transaction
from django.utils import timezone 
from datetime import datetime
from decimal import Decimal, InvalidOperation
import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.db import transaction
from datetime import datetime 
from decimal import Decimal, InvalidOperation
from django.contrib import messages 
from .models import ProviderSegregation, ClaimStatus, Box, insurance_policy, User as AuthUser



# def claim_retrieval_workload(request):
#     context = {
#         'errors_details': [],
#         'processed_records_details': [],
#         'warnings': [],
#     }
#     if request.method == 'POST':
#         context.update({
#             'batch_id': request.POST.get('batch_id', ''),
#             'audit_date_start': request.POST.get('audit_date_start', ''),
#             'audit_date_end': request.POST.get('audit_date_end', ''),
#             'audit_by': request.POST.get('audit_by', ''),
#         })

#     UPDATABLE_HEADERS_MAP = {
#         "Box Number": ("box_number", "fk_box"),
#         "Policy ID": ("PolicyId", "fk_policy"),
#         "Client ID (Claim)": ("ClientId", "string"),
#         "Member Name": ("Member_name", "string"),
#         "Provider Name": ("Provider_name", "string"),
#         "Receive Date": ("Receive_date", "date"),
#         "Issuance Date": ("Issuance_date", "date"),
#         "Segregation Date": ("Segregation_date", "date"),
#         "Segregated By": ("Segregation_by", "string"),
#         "Audit Date": ("Audit_date", "date"),
#         "Audited By": ("Audit_by", "fk_audit_by"),
#         "Batch ID": ("batchID", "string"),
#         "Requested By": ("request_by", "string"),
#         "Request Date": ("request_date", "date"),
#         "Client Name (Claim)": ("Client_name", "string"),
#         "Claim Scanned": ("claimscan", "boolean"),
#         "Deducted Amount": ("DeductedAmount", "decimal"),
#         "Note": ("note", "string_nullable"),
#         "Retrieval Date": ("retrieval_date", "date_nullable"),
#         "Return By": ("return_by", "string_nullable"),
#         "Return By" : ("return_by", "fk_return_by"),
#         "Return Date": ("return_date", "date_nullable"),
#         "Comment": ("comment", "string_nullable"),
#         "Claim Status": ("claim_status", "fk_status"),
#     }
#     ALL_EXPORT_HEADERS = ["Claim Code"] + list(UPDATABLE_HEADERS_MAP.keys())

#     if request.method == "POST":
#         action = request.POST.get("action")
#         if action == "download":
#             batch_id = request.POST.get("batch_id")
#             audit_date_start_str = request.POST.get("audit_date_start")
#             audit_date_end_str = request.POST.get("audit_date_end")
#             audit_by_name_str = request.POST.get("audit_by")

#             claims = ProviderSegregation.objects.select_related(
#                 'box_number', 'PolicyId', 'claim_status', 'Audit_by', 'Update_by', 'created_by'
#             ).all().order_by('-Audit_date')

#             try:
#                 if not batch_id and not (audit_date_start_str and audit_date_end_str) and not audit_by_name_str:
#                     messages.error(request, "Please provide at least Batch ID or Audit Date Range or Auditor Name to download the report.")
#                     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#                 if batch_id:
#                     claims = claims.filter(batchID__icontains=batch_id)
                
#                 if audit_date_start_str and audit_date_end_str:
#                     try:
#                         audit_date_start = datetime.strptime(audit_date_start_str, '%Y-%m-%d').date()
#                         audit_date_end = datetime.strptime(audit_date_end_str, '%Y-%m-%d').date()
#                         claims = claims.filter(Audit_date__gte=audit_date_start, Audit_date__lte=audit_date_end)
#                     except ValueError:
#                         messages.error(request, "Invalid date format for audit dates. Please use YYYY-MM-DD.")
#                         return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
#                 elif audit_date_start_str or audit_date_end_str:
#                     messages.error(request, "Both start and end audit dates are required for date filtering.")
#                     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
                
#                 if audit_by_name_str:
#                     claims = claims.filter(Audit_by__name__icontains=audit_by_name_str)

#             except Exception as e:
#                 messages.error(request, f"Error during filtering: {str(e)}")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             if not claims.exists():
#                 messages.info(request, "No claims found matching the specified criteria for download.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             filename_parts = ["claims_report"]
#             if batch_id: filename_parts.append(f"batch_{batch_id}")
#             if audit_date_start_str: filename_parts.append(f"from_{audit_date_start_str}")
#             if audit_date_end_str: filename_parts.append(f"to_{audit_date_end_str}")
#             if audit_by_name_str: filename_parts.append(f"auditor_{audit_by_name_str}")
#             filename = "_".join(filename_parts) + ".csv"
            
#             response = HttpResponse(content_type='text/csv')
#             response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
#             writer = csv.writer(response)
#             writer.writerow(ALL_EXPORT_HEADERS)
#             for claim in claims:
#                 row_data = [
#                     str(claim.ClaimCode) if claim.ClaimCode else "",
#                     str(claim.box_number.box_number) if claim.box_number else "",
#                     str(claim.PolicyId.policy_id) if claim.PolicyId else "",
#                     str(claim.ClientId) if claim.ClientId else "",
#                     str(claim.Member_name) if claim.Member_name else "",
#                     str(claim.Provider_name) if claim.Provider_name else "",
#                     claim.Receive_date.strftime('%Y-%m-%d') if claim.Receive_date else "",
#                     claim.Issuance_date.strftime('%Y-%m-%d') if claim.Issuance_date else "",
#                     claim.Segregation_date.strftime('%Y-%m-%d') if claim.Segregation_date else "",
#                     str(claim.Segregation_by) if claim.Segregation_by else "",
#                     claim.Audit_date.strftime('%Y-%m-%d') if claim.Audit_date else "",
#                     str(claim.Audit_by.name) if claim.Audit_by else "",

#                     str(claim.batchID) if claim.batchID else "",
#                     str(claim.request_by) if claim.request_by else "",
#                     claim.request_date.strftime('%Y-%m-%d') if claim.request_date else "",
#                     str(claim.Client_name) if claim.Client_name else "",
#                     ('Yes' if claim.claimscan else ('No' if claim.claimscan is False else "")) if claim.claimscan is not None else "",
#                     str(claim.DeductedAmount) if claim.DeductedAmount is not None else "",
#                     str(claim.note) if claim.note else "",
#                     claim.retrieval_date.strftime('%Y-%m-%d') if claim.retrieval_date else "",


#                     str(claim.return_by.name) if claim.return_by else "",
#                     #str(claim.Audit_by.name) if claim.Audit_by else "",


#                     claim.return_date.strftime('%Y-%m-%d') if claim.return_date else "",
#                     str(claim.comment) if claim.comment else "",
#                     str(claim.claim_status.name) if claim.claim_status else "",
#                 ]
#                 writer.writerow(row_data)
#             return response

#         elif action == "import_and_update":
#             if 'update_sheet' not in request.FILES:
#                 messages.error(request, "No file selected for import.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
            
#             file = request.FILES['update_sheet']
#             if not file.name.lower().endswith('.csv'):
#                 messages.error(request, "Invalid file type. Please upload a .csv file.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             data_to_update_from_csv = []
#             try:
#                 decoded_file = file.read().decode('utf-8-sig')
#                 io_string = io.StringIO(decoded_file)
#                 csv_reader = csv.reader(io_string)
                
#                 header_row_values = next(csv_reader, None)
#                 if not header_row_values:
#                     messages.error(request, "Uploaded CSV file is empty or has no header row.")
#                     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#                 if "Claim Code" not in header_row_values:
#                     messages.error(request, "Missing mandatory header in uploaded CSV: 'Claim Code'. Please use the downloaded template.")
#                     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#                 for expected_h in UPDATABLE_HEADERS_MAP.keys():
#                     if expected_h not in header_row_values:
#                         context.setdefault('warnings', []).append(f"Optional header '{expected_h}' not found in uploaded CSV. This column will not be processed for updates.")
                
#                 col_indices = {header: idx for idx, header in enumerate(header_row_values) if header in ALL_EXPORT_HEADERS}

#                 for row_num, row_values in enumerate(csv_reader, start=2):
#                     if not any(field.strip() for field in row_values): continue
                    
#                     record = {}
#                     try:
#                         claim_code_idx = col_indices.get("Claim Code")
#                         if claim_code_idx is None or claim_code_idx >= len(row_values):
#                             context.setdefault('warnings', []).append(f"Skipped row {row_num}: Claim Code column not found or row is too short.")
#                             continue
#                         claim_code_val = row_values[claim_code_idx]
#                     except IndexError:
#                         context.setdefault('warnings', []).append(f"Skipped row {row_num}: Error accessing Claim Code (row might be malformed).")
#                         continue
                    
#                     if not claim_code_val or not str(claim_code_val).strip():
#                         context.setdefault('warnings', []).append(f"Skipped row {row_num}: Missing Claim Code value.")
#                         continue
                    
#                     record['Claim Code'] = str(claim_code_val).strip()
#                     has_data_for_update_in_row = False
#                     for excel_header, (model_field, field_type) in UPDATABLE_HEADERS_MAP.items():
#                         header_idx = col_indices.get(excel_header)
#                         if header_idx is not None and header_idx < len(row_values):
#                             cell_value = row_values[header_idx]
#                             record[excel_header] = cell_value
#                             if cell_value is not None and str(cell_value).strip() != "":
#                                 has_data_for_update_in_row = True
                    
#                     if not has_data_for_update_in_row:
#                         context.setdefault('warnings', []).append(f"Skipped row {row_num} for Claim Code '{record['Claim Code']}': No updatable data found in other columns.")
#                         continue
#                     data_to_update_from_csv.append(record)

#             except UnicodeDecodeError:
#                 messages.error(request, "Error decoding file. Please ensure it is a valid UTF-8 encoded CSV.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
#             except csv.Error as e:
#                 messages.error(request, f"Error processing CSV file structure: {e}")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
#             except Exception as e:
#                 messages.error(request, f"An unexpected error occurred during file processing: {e} (Type: {type(e).__name__})")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             if not data_to_update_from_csv:
#                 if not context.get('warnings'):
#                      messages.info(request, "No valid data rows found in the uploaded CSV to process.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             success_count = 0
#             error_count = 0
            
#             current_user = request.user if request.user.is_authenticated else None
#             auth_user_instance = None
#             if current_user:
#                 try:
#                     auth_user_instance = AuthUser.objects.get(pk=current_user.pk)
#                 except AuthUser.DoesNotExist:
#                     auth_user_instance = None
#                     context.setdefault('warnings', []).append("Could not associate updating user: Logged in user not found in custom AuthUser table.")
            
#             for record_data in data_to_update_from_csv:
#                 claim_code = record_data.get('Claim Code')
#                 try:
#                     with transaction.atomic():
#                         claim = ProviderSegregation.objects.get(ClaimCode=claim_code)
#                         updated_fields_log = []
                        
#                         for excel_header, (model_field_name, field_type) in UPDATABLE_HEADERS_MAP.items():
#                             if excel_header not in record_data: continue
                            
#                             new_value_raw = record_data.get(excel_header)
#                             current_value = getattr(claim, model_field_name, None)
#                             parsed_new_value = None
#                             update_field = False
                            
#                             new_value_clean_str = str(new_value_raw).strip() if new_value_raw is not None else None

#                             if field_type == "string":
#                                 parsed_new_value = new_value_clean_str if new_value_clean_str is not None else ""
#                                 if parsed_new_value != (str(current_value) if current_value is not None else ""):
#                                     update_field = True
#                             elif field_type == "string_nullable":
#                                 parsed_new_value = new_value_clean_str if new_value_clean_str else None
#                                 if parsed_new_value != current_value:
#                                     update_field = True
#                             elif field_type == "date" or field_type == "date_nullable":
#                                 if new_value_clean_str:
#                                     try:
#                                         parsed_new_value = datetime.strptime(new_value_clean_str.split(" ")[0], '%Y-%m-%d').date()
#                                     except ValueError:
#                                         raise ValueError(f"Invalid date format for '{excel_header}' ('{new_value_raw}'). Expected YYYY-MM-DD.")
#                                 elif field_type == "date_nullable":
#                                     parsed_new_value = None
#                                 else:
#                                     if not ProviderSegregation._meta.get_field(model_field_name).null:
#                                         raise ValueError(f"'{excel_header}' cannot be empty (required date field).")
#                                     parsed_new_value = None
#                                 if parsed_new_value != current_value:
#                                     update_field = True
#                             elif field_type == "boolean":
#                                 if new_value_clean_str is not None:
#                                     if new_value_clean_str.lower() in ['yes', 'true', '1']: parsed_new_value = True
#                                     elif new_value_clean_str.lower() in ['no', 'false', '0']: parsed_new_value = False
#                                     elif new_value_clean_str == '' and ProviderSegregation._meta.get_field(model_field_name).null: parsed_new_value = None
#                                     else: raise ValueError(f"Invalid boolean value for '{excel_header}' ('{new_value_raw}'). Expected Yes/No, True/False, or empty for None if applicable.")
#                                 else: # new_value_raw is None
#                                     if ProviderSegregation._meta.get_field(model_field_name).null: parsed_new_value = None
#                                     else: parsed_new_value = current_value # No change if None not allowed and no value given

#                                 if parsed_new_value != current_value:
#                                     update_field = True
#                             elif field_type == "decimal":
#                                 if new_value_clean_str:
#                                     try:
#                                         parsed_new_value = Decimal(new_value_clean_str)
#                                     except InvalidOperation:
#                                         raise ValueError(f"Invalid decimal value for '{excel_header}' ('{new_value_raw}').")
#                                 elif ProviderSegregation._meta.get_field(model_field_name).null:
#                                     parsed_new_value = None
#                                 else:
#                                      raise ValueError(f"'{excel_header}' is a mandatory decimal field and cannot be empty.")
#                                 if parsed_new_value != current_value:
#                                     update_field = True
                            
#                             elif field_type in ["fk_box", "fk_policy", "fk_status", "fk_audit_by", "fk_return_by"]:
#                                 current_fk_instance = getattr(claim, model_field_name, None)
#                                 if not new_value_clean_str:
#                                     if not ProviderSegregation._meta.get_field(model_field_name).null:
#                                         raise ValueError(f"'{excel_header}' is a mandatory field and cannot be empty.")
#                                     parsed_new_value = None
#                                 else:
#                                     try:
#                                         if field_type == "fk_box":
#                                             parsed_new_value = Box.objects.get(box_number=new_value_clean_str)
#                                         elif field_type == "fk_policy":
#                                             parsed_new_value = insurance_policy.objects.get(policy_id=int(new_value_clean_str))
#                                         elif field_type == "fk_status":
#                                             parsed_new_value = ClaimStatus.objects.get(name=new_value_clean_str)
#                                         elif field_type == "fk_audit_by":
#                                             parsed_new_value = audit_by.objects.get(name=new_value_clean_str)
                                        
#                                         elif field_type == "fk_return_by":
#                                             parsed_new_value = return_by_user.objects.get(name=new_value_clean_str)


#                                     except (Box.DoesNotExist, insurance_policy.DoesNotExist, ClaimStatus.DoesNotExist, audit_by.DoesNotExist):
#                                         raise ValueError(f"Related object for '{excel_header}' with value '{new_value_clean_str}' not found.")
#                                     except ValueError as verr: # For int conversion for policy_id
#                                         raise ValueError(f"Invalid ID format for '{excel_header}' ('{new_value_clean_str}'): {str(verr)}")

#                                 if parsed_new_value != current_fk_instance:
#                                     update_field = True

#                             if update_field:
#                                 setattr(claim, model_field_name, parsed_new_value)
#                                 new_val_display = str(parsed_new_value.name if hasattr(parsed_new_value, 'name') else \
#                                                   (parsed_new_value.box_number if hasattr(parsed_new_value, 'box_number') else \
#                                                   (parsed_new_value.policy_id if hasattr(parsed_new_value, 'policy_id') else parsed_new_value))) \
#                                                   if parsed_new_value is not None else "None/Empty"
#                                 old_val_display = str(current_value.name if hasattr(current_value, 'name') else \
#                                                   (current_value.box_number if hasattr(current_value, 'box_number') else \
#                                                   (current_value.policy_id if hasattr(current_value, 'policy_id') else current_value))) \
#                                                   if current_value is not None else "None/Empty"
#                                 updated_fields_log.append(f"{excel_header} to '{new_val_display}' (was '{old_val_display}')")
                                
#                                 if model_field_name == "PolicyId" and parsed_new_value:
#                                     claim.Client_name = parsed_new_value.client_name
#                                     claim.ClientId = parsed_new_value.client_id
#                                     updated_fields_log.append(f"Client Name updated to '{claim.Client_name}' from policy")
#                                     updated_fields_log.append(f"Client ID updated to '{claim.ClientId}' from policy")


#                         if updated_fields_log:
#                             if auth_user_instance:
#                                 claim.Update_by = auth_user_instance
                            
#                             if hasattr(claim, 'updated_at'): # Check if model has updated_at
#                                 model_updated_at_field = claim._meta.get_field('updated_at')
#                                 if not getattr(model_updated_at_field, 'auto_now', False) and \
#                                    not getattr(model_updated_at_field, 'auto_now_add', False): # only set if not auto
#                                     claim.updated_at = timezone.now()

#                             claim.save()
#                             success_count += 1
#                             context['processed_records_details'].append(f"Claim '{claim_code}': Successfully updated ({', '.join(updated_fields_log)}).")
#                         else:
#                             context['processed_records_details'].append(f"Claim '{claim_code}': No changes detected or applied.")
                
#                 except ProviderSegregation.DoesNotExist:
#                     error_count += 1
#                     context['errors_details'].append(f"Claim '{claim_code}': Not found in database.")
#                 except ValueError as ve:
#                     error_count += 1
#                     context['errors_details'].append(f"Claim '{claim_code}': Data validation error - {str(ve)}")
#                 except IntegrityError as ie:
#                     error_count += 1
#                     context['errors_details'].append(f"Claim '{claim_code}': Data integrity error (e.g. duplicate value for a unique field) - {str(ie)}")
#                 except Exception as e:
#                     error_count += 1
#                     context['errors_details'].append(f"Claim '{claim_code}': An unexpected error occurred: {e} (Type: {type(e).__name__})")
            
#             total_records_in_csv_for_update = len(data_to_update_from_csv)
#             if success_count > 0:
#                 messages.success(request, f"Import and Update process complete. Successfully updated {success_count} of {total_records_in_csv_for_update} processed claim(s).")
#             if error_count > 0:
#                 messages.error(request, f"Import and Update process encountered issues. Failed to update {error_count} of {total_records_in_csv_for_update} processed claim(s). See details on page.")
#             if success_count == 0 and error_count == 0 and total_records_in_csv_for_update > 0:
#                  messages.info(request, "Import and Update process complete. No claims required changes based on the CSV data, or all data matched existing records.")
            
#             if request.session.get('imported_data_ready'): del request.session['imported_data_ready']
#             if request.session.get('claims_to_update_data'): del request.session['claims_to_update_data']
            
#             return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)




import re
import csv
import io
from datetime import datetime # Make sure this is imported
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.utils import timezone 


from .models import ProviderSegregation 
from .models import Box 
from .models import insurance_policy 
from .models import ClaimStatus
from .models import audit_by
from .models import return_by_user 
from .models import segregation_by 


# def claim_retrieval_workload(request):
#     context = {
#         'errors_details': [],
#         'processed_records_details': [],
#         'warnings': [],
#     }
#     if request.method == 'POST':
#         context.update({
#             'batch_id': request.POST.get('batch_id', ''),
#             'audit_date_start': request.POST.get('audit_date_start', ''),
#             'audit_date_end': request.POST.get('audit_date_end', ''),
#             'audit_by': request.POST.get('audit_by', ''),
#         })

#     UPDATABLE_HEADERS_MAP = {
#         "Box Number": ("box_number", "fk_box"),
#         "Policy ID": ("PolicyId", "fk_policy"),
#         "Client ID (Claim)": ("ClientId", "string"),
#         "Member Name": ("Member_name", "string"),
#         "Provider Name": ("Provider_name", "string"),
#         "Receive Date": ("Receive_date", "date"),
#         "Issuance Date": ("Issuance_date", "date"),
#         "Segregation Date": ("Segregation_date", "date"),
#         "Segregated By": ("Segregation_by", "fk_segregation_by"),  
#         "Audit Date": ("Audit_date", "date"),
#         "Audited By": ("Audit_by", "fk_audit_by"),
#         "Batch ID": ("batchID", "string"),
#         "Requested By": ("request_by", "string"),
#         "Request Date": ("request_date", "date"),
#         "Client Name (Claim)": ("Client_name", "string"),
#         "Claim Scanned": ("claimscan", "boolean"),
#         "Deducted Amount": ("DeductedAmount", "decimal"),
#         "Note": ("note", "string_nullable"),
#         "Retrieval Date": ("retrieval_date", "date_nullable"),
#         "Return By": ("return_by", "fk_return_by"),
#         "Return Date": ("return_date", "date_nullable"),
#         "Comment": ("comment", "string_nullable"),
#         "Claim Status": ("claim_status", "fk_status"),
#     }
#     ALL_EXPORT_HEADERS = ["Claim Code"] + list(UPDATABLE_HEADERS_MAP.keys())

#     if request.method == "POST":
#         action = request.POST.get("action")
#         if action == "download":
#             batch_id = request.POST.get("batch_id")
#             audit_date_start_str = request.POST.get("audit_date_start")
#             audit_date_end_str = request.POST.get("audit_date_end")
#             audit_by_name_str = request.POST.get("audit_by")

#             claims = ProviderSegregation.objects.select_related(
#                 'box_number', 'PolicyId', 'claim_status', 'Audit_by', 
#                 'Update_by', 'created_by', 'Segregation_by'  
#             ).all().order_by('-Audit_date')

#             try:
#                 if not batch_id and not (audit_date_start_str and audit_date_end_str) and not audit_by_name_str:
#                     messages.error(request, "Please provide at least Batch ID or Audit Date Range or Auditor Name to download the report.")
#                     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#                 if batch_id:
#                     claims = claims.filter(batchID__icontains=batch_id)
                
#                 if audit_date_start_str and audit_date_end_str:
#                     try:
#                         audit_date_start = datetime.strptime(audit_date_start_str, '%Y-%m-%d').date()
#                         audit_date_end = datetime.strptime(audit_date_end_str, '%Y-%m-%d').date()
#                         claims = claims.filter(Audit_date__gte=audit_date_start, Audit_date__lte=audit_date_end)
#                     except ValueError:
#                         messages.error(request, "Invalid date format for audit dates. Please use YYYY-MM-DD.")
#                         return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
#                 elif audit_date_start_str or audit_date_end_str:
#                     messages.error(request, "Both start and end audit dates are required for date filtering.")
#                     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
                
#                 if audit_by_name_str:
#                     claims = claims.filter(Audit_by__name__icontains=audit_by_name_str)

#             except Exception as e:
#                 messages.error(request, f"Error during filtering: {str(e)}")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             if not claims.exists():
#                 messages.info(request, "No claims found matching the specified criteria for download.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             filename_parts = ["claims_report"]
#             if batch_id: filename_parts.append(f"batch_{batch_id}")
#             if audit_date_start_str: filename_parts.append(f"from_{audit_date_start_str}")
#             if audit_date_end_str: filename_parts.append(f"to_{audit_date_end_str}")
#             if audit_by_name_str: filename_parts.append(f"auditor_{audit_by_name_str}")
#             filename = "_".join(filename_parts) + ".csv"
            
#             response = HttpResponse(content_type='text/csv')
#             response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
#             writer = csv.writer(response)
#             writer.writerow(ALL_EXPORT_HEADERS)
#             for claim in claims:
#                 row_data = [
#                     str(claim.ClaimCode) if claim.ClaimCode else "",
#                     str(claim.box_number.box_number) if claim.box_number else "",
#                     str(claim.PolicyId.policy_id) if claim.PolicyId else "",
#                     str(claim.ClientId) if claim.ClientId else "",
#                     str(claim.Member_name) if claim.Member_name else "",
#                     str(claim.Provider_name) if claim.Provider_name else "",
#                     claim.Receive_date.strftime('%Y-%m-%d') if claim.Receive_date else "",
#                     claim.Issuance_date.strftime('%Y-%m-%d') if claim.Issuance_date else "",
#                     claim.Segregation_date.strftime('%Y-%m-%d') if claim.Segregation_date else "",
                    
#                     str(claim.Segregation_by.name) if claim.Segregation_by else "",
#                     claim.Audit_date.strftime('%Y-%m-%d') if claim.Audit_date else "",
#                     str(claim.Audit_by.name) if claim.Audit_by else "",
#                     str(claim.batchID) if claim.batchID else "",
#                     str(claim.request_by) if claim.request_by else "",
#                     claim.request_date.strftime('%Y-%m-%d') if claim.request_date else "",
#                     str(claim.Client_name) if claim.Client_name else "",
#                     ('Yes' if claim.claimscan else ('No' if claim.claimscan is False else "")) if claim.claimscan is not None else "",
#                     str(claim.DeductedAmount) if claim.DeductedAmount is not None else "",
#                     str(claim.note) if claim.note else "",
#                     claim.retrieval_date.strftime('%Y-%m-%d') if claim.retrieval_date else "",
#                     str(claim.return_by.name) if claim.return_by else "",
#                     claim.return_date.strftime('%Y-%m-%d') if claim.return_date else "",
#                     str(claim.comment) if claim.comment else "",
#                     str(claim.claim_status.name) if claim.claim_status else "",
#                 ]
#                 writer.writerow(row_data)
#             return response

#         elif action == "import_and_update":
#             if 'update_sheet' not in request.FILES:
#                 messages.error(request, "No file selected for import.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
            
#             file = request.FILES['update_sheet']
#             if not file.name.lower().endswith('.csv'):
#                 messages.error(request, "Invalid file type. Please upload a .csv file.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             data_to_update_from_csv = []
#             try:
#                 decoded_file = file.read().decode('utf-8-sig')
#                 io_string = io.StringIO(decoded_file)
#                 csv_reader = csv.reader(io_string)
                
#                 header_row_values = next(csv_reader, None)
#                 if not header_row_values:
#                     messages.error(request, "Uploaded CSV file is empty or has no header row.")
#                     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#                 if "Claim Code" not in header_row_values:
#                     messages.error(request, "Missing mandatory header in uploaded CSV: 'Claim Code'. Please use the downloaded template.")
#                     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#                 for expected_h in UPDATABLE_HEADERS_MAP.keys():
#                     if expected_h not in header_row_values:
#                         context.setdefault('warnings', []).append(f"Optional header '{expected_h}' not found in uploaded CSV. This column will not be processed for updates.")
                
#                 col_indices = {header: idx for idx, header in enumerate(header_row_values) if header in ALL_EXPORT_HEADERS}

#                 for row_num, row_values in enumerate(csv_reader, start=2):
#                     if not any(field.strip() for field in row_values): continue
                    
#                     record = {}
#                     try:
#                         claim_code_idx = col_indices.get("Claim Code")
#                         if claim_code_idx is None or claim_code_idx >= len(row_values):
#                             context.setdefault('warnings', []).append(f"Skipped row {row_num}: Claim Code column not found or row is too short.")
#                             continue
#                         claim_code_val = row_values[claim_code_idx]
#                     except IndexError:
#                         context.setdefault('warnings', []).append(f"Skipped row {row_num}: Error accessing Claim Code (row might be malformed).")
#                         continue
                    
#                     if not claim_code_val or not str(claim_code_val).strip():
#                         context.setdefault('warnings', []).append(f"Skipped row {row_num}: Missing Claim Code value.")
#                         continue
                    
#                     record['Claim Code'] = str(claim_code_val).strip()
#                     has_data_for_update_in_row = False
#                     for excel_header, (model_field, field_type) in UPDATABLE_HEADERS_MAP.items():
#                         header_idx = col_indices.get(excel_header)
#                         if header_idx is not None and header_idx < len(row_values):
#                             cell_value = row_values[header_idx]
#                             record[excel_header] = cell_value
#                             if cell_value is not None and str(cell_value).strip() != "":
#                                 has_data_for_update_in_row = True
                    
#                     if not has_data_for_update_in_row:
#                         context.setdefault('warnings', []).append(f"Skipped row {row_num} for Claim Code '{record['Claim Code']}': No updatable data found in other columns.")
#                         continue
#                     data_to_update_from_csv.append(record)

#             except UnicodeDecodeError:
#                 messages.error(request, "Error decoding file. Please ensure it is a valid UTF-8 encoded CSV.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
#             except csv.Error as e:
#                 messages.error(request, f"Error processing CSV file structure: {e}")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
#             except Exception as e:
#                 messages.error(request, f"An unexpected error occurred during file processing: {e} (Type: {type(e).__name__})")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             if not data_to_update_from_csv:
#                 if not context.get('warnings'):
#                      messages.info(request, "No valid data rows found in the uploaded CSV to process.")
#                 return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#             success_count = 0
#             error_count = 0
            
#             current_user = request.user if request.user.is_authenticated else None
#             auth_user_instance = None
#             if current_user:
#                 try:
#                     auth_user_instance = AuthUser.objects.get(pk=current_user.pk)
#                 except AuthUser.DoesNotExist:
#                     auth_user_instance = None
#                     context.setdefault('warnings', []).append("Could not associate updating user: Logged in user not found in custom AuthUser table.")
            
#             for record_data in data_to_update_from_csv:
#                 claim_code = record_data.get('Claim Code')
#                 try:
#                     with transaction.atomic():
#                         claim = ProviderSegregation.objects.get(ClaimCode=claim_code)
#                         updated_fields_log = []
                        
#                         for excel_header, (model_field_name, field_type) in UPDATABLE_HEADERS_MAP.items():
#                             if excel_header not in record_data: continue
                            
#                             new_value_raw = record_data.get(excel_header)
#                             current_value = getattr(claim, model_field_name, None)
#                             parsed_new_value = None
#                             update_field = False
                            
#                             new_value_clean_str = str(new_value_raw).strip() if new_value_raw is not None else None

#                             if field_type == "string":
#                                 parsed_new_value = new_value_clean_str if new_value_clean_str is not None else ""
#                                 if parsed_new_value != (str(current_value) if current_value is not None else ""):
#                                     update_field = True
#                             elif field_type == "string_nullable":
#                                 parsed_new_value = new_value_clean_str if new_value_clean_str else None
#                                 if parsed_new_value != current_value:
#                                     update_field = True
#                             elif field_type == "date" or field_type == "date_nullable":
                                
#                                 if new_value_clean_str:
#                                     try:
                                        
#                                         if not re.match(r'^\d{4}-\d{2}-\d{2}$', new_value_clean_str):
#                                             raise ValueError("Date must be in YYYY-MM-DD format")
#                                         parsed_new_value = datetime.strptime(new_value_clean_str, '%Y-%m-%d').date()
#                                     except ValueError as ve:
#                                         raise ValueError(f"Invalid date format for '{excel_header}' ('{new_value_raw}'). {str(ve)}")
#                                 elif field_type == "date_nullable":
#                                     parsed_new_value = None
#                                 else:
                                    
#                                     if not ProviderSegregation._meta.get_field(model_field_name).null:
#                                         raise ValueError(f"'{excel_header}' cannot be empty (required date field).")
#                                     parsed_new_value = None
                                    
#                                 if parsed_new_value != current_value:
#                                     update_field = True
#                             elif field_type == "boolean":
#                                 if new_value_clean_str is not None:
#                                     if new_value_clean_str.lower() in ['yes', 'true', '1']: parsed_new_value = True
#                                     elif new_value_clean_str.lower() in ['no', 'false', '0']: parsed_new_value = False
#                                     elif new_value_clean_str == '' and ProviderSegregation._meta.get_field(model_field_name).null: parsed_new_value = None
#                                     else: raise ValueError(f"Invalid boolean value for '{excel_header}' ('{new_value_raw}'). Expected Yes/No, True/False, or empty for None if applicable.")
#                                 else: 
#                                     if ProviderSegregation._meta.get_field(model_field_name).null: parsed_new_value = None
#                                     else: parsed_new_value = current_value 

#                                 if parsed_new_value != current_value:
#                                     update_field = True
#                             elif field_type == "decimal":
#                                 if new_value_clean_str:
#                                     try:
#                                         parsed_new_value = Decimal(new_value_clean_str)
#                                     except InvalidOperation:
#                                         raise ValueError(f"Invalid decimal value for '{excel_header}' ('{new_value_raw}').")
#                                 elif ProviderSegregation._meta.get_field(model_field_name).null:
#                                     parsed_new_value = None
#                                 else:
#                                      raise ValueError(f"'{excel_header}' is a mandatory decimal field and cannot be empty.")
#                                 if parsed_new_value != current_value:
#                                     update_field = True
                            
#                             elif field_type in ["fk_box", "fk_policy", "fk_status", "fk_audit_by", "fk_return_by", "fk_segregation_by"]:
#                                 current_fk_instance = getattr(claim, model_field_name, None)
#                                 if not new_value_clean_str:
#                                     if not ProviderSegregation._meta.get_field(model_field_name).null:
#                                         raise ValueError(f"'{excel_header}' is a mandatory field and cannot be empty.")
#                                     parsed_new_value = None
#                                 else:
#                                     try:
#                                         if field_type == "fk_box":
#                                             parsed_new_value = Box.objects.get(box_number=new_value_clean_str)
#                                         elif field_type == "fk_policy":
#                                             parsed_new_value = insurance_policy.objects.get(policy_id=int(new_value_clean_str))
#                                         elif field_type == "fk_status":
#                                             parsed_new_value = ClaimStatus.objects.get(name=new_value_clean_str)
#                                         elif field_type == "fk_audit_by":
#                                             parsed_new_value = audit_by.objects.get(name=new_value_clean_str)
#                                         elif field_type == "fk_return_by":
#                                             parsed_new_value = return_by_user.objects.get(name=new_value_clean_str)
#                                         elif field_type == "fk_segregation_by":  # New FK handling
#                                             parsed_new_value = segregation_by.objects.get(name=new_value_clean_str)
#                                     except (Box.DoesNotExist, insurance_policy.DoesNotExist, 
#                                             ClaimStatus.DoesNotExist, audit_by.DoesNotExist,
#                                             return_by_user.DoesNotExist, segregation_by.DoesNotExist):
#                                         raise ValueError(f"Related object for '{excel_header}' with value '{new_value_clean_str}' not found.")
#                                     except ValueError as verr: 
#                                         raise ValueError(f"Invalid ID format for '{excel_header}' ('{new_value_clean_str}'): {str(verr)}")

#                                 if parsed_new_value != current_fk_instance:
#                                     update_field = True

#                             if update_field:
#                                 setattr(claim, model_field_name, parsed_new_value)
#                                 new_val_display = str(parsed_new_value.name if hasattr(parsed_new_value, 'name') else \
#                                                   (parsed_new_value.box_number if hasattr(parsed_new_value, 'box_number') else \
#                                                   (parsed_new_value.policy_id if hasattr(parsed_new_value, 'policy_id') else parsed_new_value))) \
#                                                   if parsed_new_value is not None else "None/Empty"
#                                 old_val_display = str(current_value.name if hasattr(current_value, 'name') else \
#                                                   (current_value.box_number if hasattr(current_value, 'box_number') else \
#                                                   (current_value.policy_id if hasattr(current_value, 'policy_id') else current_value))) \
#                                                   if current_value is not None else "None/Empty"
#                                 updated_fields_log.append(f"{excel_header} to '{new_val_display}' (was '{old_val_display}')")
                                
#                                 if model_field_name == "PolicyId" and parsed_new_value:
#                                     claim.Client_name = parsed_new_value.client_name
#                                     claim.ClientId = parsed_new_value.client_id
#                                     updated_fields_log.append(f"Client Name updated to '{claim.Client_name}' from policy")
#                                     updated_fields_log.append(f"Client ID updated to '{claim.ClientId}' from policy")

#                         if updated_fields_log:
#                             if auth_user_instance:
#                                 claim.Update_by = auth_user_instance
                            
#                             if hasattr(claim, 'updated_at'): 
#                                 model_updated_at_field = claim._meta.get_field('updated_at')
#                                 if not getattr(model_updated_at_field, 'auto_now', False) and \
#                                    not getattr(model_updated_at_field, 'auto_now_add', False): # only set if not auto
#                                     claim.updated_at = timezone.now()

#                             claim.save()
#                             success_count += 1
#                             context['processed_records_details'].append(f"Claim '{claim_code}': Successfully updated ({', '.join(updated_fields_log)}).")
#                         else:
#                             context['processed_records_details'].append(f"Claim '{claim_code}': No changes detected or applied.")
                
#                 except ProviderSegregation.DoesNotExist:
#                     error_count += 1
#                     context['errors_details'].append(f"Claim '{claim_code}': Not found in database.")
#                 except ValueError as ve:
#                     error_count += 1
#                     context['errors_details'].append(f"Claim '{claim_code}': Data validation error - {str(ve)}")
#                 except IntegrityError as ie:
#                     error_count += 1
#                     context['errors_details'].append(f"Claim '{claim_code}': Data integrity error (e.g. duplicate value for a unique field) - {str(ie)}")
#                 except Exception as e:
#                     error_count += 1
#                     context['errors_details'].append(f"Claim '{claim_code}': An unexpected error occurred: {e} (Type: {type(e).__name__})")
            
#             total_records_in_csv_for_update = len(data_to_update_from_csv)
#             if success_count > 0:
#                 messages.success(request, f"Import and Update process complete. Successfully updated {success_count} of {total_records_in_csv_for_update} processed claim(s).")
#             if error_count > 0:
#                 messages.error(request, f"Import and Update process encountered issues. Failed to update {error_count} of {total_records_in_csv_for_update} processed claim(s). See details on page.")
#             if success_count == 0 and error_count == 0 and total_records_in_csv_for_update > 0:
#                  messages.info(request, "Import and Update process complete. No claims required changes based on the CSV data, or all data matched existing records.")
            
#             if request.session.get('imported_data_ready'): del request.session['imported_data_ready']
#             if request.session.get('claims_to_update_data'): del request.session['claims_to_update_data']
            
#             return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

#     return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)






































import io
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.db import transaction, IntegrityError
from django.core.exceptions import ValidationError
from .models import (
    ProviderSegregation, Box, insurance_policy, ClaimStatus, audit_by, 
    return_by_user, segregation_by, request_by, InsuranceCompany
)

# I've added all necessary imports for the function to work standalone.

def claim_retrieval_workload(request):
    context = {
        'errors_details': [],
        'processed_records_details': [],
        'warnings': [],
        'batch_id': request.POST.get('batch_id', '') if request.method == 'POST' else '',
        'audit_date_start': request.POST.get('audit_date_start', '') if request.method == 'POST' else '',
        'audit_date_end': request.POST.get('audit_date_end', '') if request.method == 'POST' else '',
        'audit_by': request.POST.get('audit_by', '') if request.method == 'POST' else '',
    }

    UPDATABLE_HEADERS_MAP = {
        "Box Number": ("box_number", "fk_box"),
        "Policy ID": ("PolicyId", "fk_policy"),
        "Client ID (Claim)": ("ClientId", "string"),
        "Member Name": ("Member_name", "string"),
        "Provider Name": ("Provider_name", "string"),
        "Receive Date": ("Receive_date", "date"),
        "Issuance Date": ("Issuance_date", "date"),
        "Segregation Date": ("Segregation_date", "date"),
        "Segregated By": ("Segregation_by", "fk_segregation_by"),
        "Audit Date": ("Audit_date", "date"),
        "Audited By": ("Audit_by", "fk_audit_by"),
        "Batch ID": ("batchID", "string"),
        "Requested By": ("request_by", "fk_request_by"),
        "Request Date": ("request_date", "date"),
        "Client Name (Claim)": ("Client_name", "string"),
        "Claim Scanned": ("claimscan", "boolean"),
        "Deducted Amount": ("DeductedAmount", "decimal"),
        "Note": ("note", "string_nullable"),
        "Retrieval Date": ("retrieval_date", "date_nullable"),
        "Return By": ("return_by", "fk_return_by"),
        "Return Date": ("return_date", "date_nullable"),
        "Comment": ("comment", "string_nullable"),
        "Claim Status": ("claim_status", "fk_status"),
    }
    ALL_EXPORT_HEADERS = ["Claim Code"] + list(UPDATABLE_HEADERS_MAP.keys())

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "download":
            batch_id = request.POST.get("batch_id")
            audit_date_start_str = request.POST.get("audit_date_start")
            audit_date_end_str = request.POST.get("audit_date_end")
            audit_by_name_str = request.POST.get("audit_by")

            claims_query = ProviderSegregation.objects.select_related(
                'box_number', 'PolicyId', 'claim_status', 'Audit_by', 
                'Update_by', 'created_by', 'Segregation_by', 'request_by', 'return_by'
            ).all().order_by('-Audit_date')

            try:
                if not any([batch_id, audit_date_start_str, audit_date_end_str, audit_by_name_str]):
                    messages.error(request, "Please provide at least Batch ID, Audit Date Range, or Auditor Name to download.")
                    return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

                if batch_id:
                    claims_query = claims_query.filter(batchID__icontains=batch_id)
                
                if audit_date_start_str and audit_date_end_str:
                    try:
                        audit_date_start = datetime.strptime(audit_date_start_str, '%Y-%m-%d').date()
                        audit_date_end = datetime.strptime(audit_date_end_str, '%Y-%m-%d').date()
                        if audit_date_start > audit_date_end:
                                messages.error(request, "Audit start date cannot be after audit end date.")
                                return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
                        claims_query = claims_query.filter(Audit_date__range=[audit_date_start, audit_date_end])
                    except ValueError:
                        messages.error(request, "Invalid date format for audit dates. Please use YYYY-MM-DD.")
                        return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
                elif audit_date_start_str or audit_date_end_str:
                    messages.error(request, "Both start and end audit dates are required for date range filtering.")
                    return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
                
                if audit_by_name_str:
                    claims_query = claims_query.filter(Audit_by__name__icontains=audit_by_name_str)

            except Exception as e:
                messages.error(request, f"Error during filtering claims: {str(e)}")
                return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

            if not claims_query.exists():
                messages.info(request, "No claims found matching the specified criteria for download.")
                return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

            filename = f"claims_report_{timezone.now().strftime('%Y%m%d%H%M%S')}.csv"
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            writer = csv.writer(response)
            writer.writerow(ALL_EXPORT_HEADERS)

            for claim in claims_query:
                def get_related_attr(related_obj, attr='name'):
                    return getattr(related_obj, attr, "") if related_obj else ""

                row_data = [
                    str(claim.ClaimCode or ""),
                    get_related_attr(claim.box_number, 'box_number'),
                    get_related_attr(claim.PolicyId, 'policy_id'),
                    str(claim.ClientId or ""),
                    str(claim.Member_name or ""),
                    str(claim.Provider_name or ""),
                    claim.Receive_date.strftime('%d/%m/%Y') if claim.Receive_date else "",
                    claim.Issuance_date.strftime('%d/%m/%Y') if claim.Issuance_date else "",
                    claim.Segregation_date.strftime('%d/%m/%Y') if claim.Segregation_date else "",
                    get_related_attr(claim.Segregation_by), 
                    claim.Audit_date.strftime('%d/%m/%Y') if claim.Audit_date else "",
                    get_related_attr(claim.Audit_by),
                    str(claim.batchID or ""),
                    get_related_attr(claim.request_by),
                    claim.request_date.strftime('%d/%m/%Y') if claim.request_date else "",
                    str(claim.Client_name or ""), 
                    'Yes' if claim.claimscan else 'No' if claim.claimscan is False else "",
                    str(claim.DeductedAmount or ""),
                    str(claim.note or ""),
                    claim.retrieval_date.strftime('%d/%m/%Y') if claim.retrieval_date else "",
                    get_related_attr(claim.return_by),
                    claim.return_date.strftime('%d/%m/%Y') if claim.return_date else "",
                    str(claim.comment or ""),
                    get_related_attr(claim.claim_status),
                ]
                writer.writerow(row_data)
            return response

        elif action == "import_and_update":
            # ... (file upload checks remain the same) ...
            if 'update_sheet' not in request.FILES:
                messages.error(request, "No file selected for import.")
                return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
            
            file = request.FILES['update_sheet']
            if not file.name.lower().endswith('.csv'):
                messages.error(request, "Invalid file type. Please upload a .csv file.")
                return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
            
            # --- Start: Processing the uploaded CSV ---
            data_to_update_from_csv = []
            try:
                decoded_file_content = file.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file_content)
                csv_reader = csv.DictReader(io_string)
                
                csv_headers = [h.strip().replace('\ufeff', '') for h in csv_reader.fieldnames] if csv_reader.fieldnames else []
                if "Claim Code" not in csv_headers:
                    messages.error(request, "Missing mandatory header in uploaded CSV: 'Claim Code'.")
                    return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)

                for row_num, raw_row_dict in enumerate(csv_reader, start=2):
                    processed_row_dict = {k.strip().replace('\ufeff', ''): str(v).strip() for k, v in raw_row_dict.items() if v is not None}
                    if not any(processed_row_dict.values()): continue
                    
                    claim_code_val = processed_row_dict.get("Claim Code", "")
                    if not claim_code_val:
                        context.setdefault('warnings', []).append(f"Skipped row {row_num}: Missing 'Claim Code' value.")
                        continue
                    
                    data_to_update_from_csv.append(processed_row_dict)

            except Exception as e:
                messages.error(request, f"An unexpected error occurred during file processing: {e}")
                return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
            
            # --- Start: Database Update Logic ---
            updated_records_count = 0
            processed_claim_codes = set()
            
            for record_data_dict in data_to_update_from_csv:
                claim_code_from_csv = record_data_dict.get('Claim Code')
                if claim_code_from_csv in processed_claim_codes:
                    context.setdefault('warnings', []).append(f"Duplicate Claim Code '{claim_code_from_csv}' in CSV. Processing first instance only.")
                    continue
                processed_claim_codes.add(claim_code_from_csv)

                try:
                    with transaction.atomic():
                        # MODIFICATION: Use select_related for efficiency and to get insurance companies for validation
                        claim_to_update = ProviderSegregation.objects.select_related(
                            'box_number__insurance_company', 
                            'PolicyId__insurance_company'
                        ).get(ClaimCode=claim_code_from_csv)
                        
                        fields_changed_in_this_record_log = []
                        
                        for csv_header, (model_field_name, field_type_tag) in UPDATABLE_HEADERS_MAP.items():
                            new_value_str_from_csv = record_data_dict.get(csv_header, None)
                            if new_value_str_from_csv is None: continue

                            current_db_value = getattr(claim_to_update, model_field_name, None)
                            parsed_new_value_for_field = None
                            
                            try:
                                # --- Start of parsing logic for each field type ---
                                if field_type_tag == "fk_box":
                                    parsed_new_value_for_field = Box.objects.select_related('insurance_company').get(box_number=new_value_str_from_csv)
                                    # --- NEW VALIDATION LOGIC FOR BOX ---
                                    if claim_to_update.PolicyId and parsed_new_value_for_field.insurance_company != claim_to_update.PolicyId.insurance_company:
                                        context['errors_details'].append(f"Claim '{claim_code_from_csv}': Box update failed. The new box's insurance company ('{parsed_new_value_for_field.insurance_company.name}') does not match the policy's insurance company ('{claim_to_update.PolicyId.insurance_company.name}').")
                                        continue # Skip this field update

                                elif field_type_tag == "fk_policy":
                                    parsed_new_value_for_field = insurance_policy.objects.select_related('insurance_company').get(policy_id=int(new_value_str_from_csv))
                                    # --- NEW VALIDATION LOGIC FOR POLICY ---
                                    if claim_to_update.box_number and parsed_new_value_for_field.insurance_company != claim_to_update.box_number.insurance_company:
                                        context['errors_details'].append(f"Claim '{claim_code_from_csv}': Policy update failed. The new policy's insurance company ('{parsed_new_value_for_field.insurance_company.name}') does not match the box's insurance company ('{claim_to_update.box_number.insurance_company.name}').")
                                        continue # Skip this field update

                                # Other elifs for parsing different field types (date, boolean, decimal, other FKs)
                                elif field_type_tag in ["date", "date_nullable"]:
                                    parsed_new_value_for_field = datetime.strptime(new_value_str_from_csv, '%d/%m/%Y').date()
                                elif field_type_tag == "boolean":
                                    parsed_new_value_for_field = new_value_str_from_csv.lower() in ['yes', 'true', '1']
                                elif field_type_tag == "decimal":
                                    parsed_new_value_for_field = Decimal(new_value_str_from_csv)
                                elif field_type_tag == "fk_status":
                                    parsed_new_value_for_field = ClaimStatus.objects.get(name=new_value_str_from_csv)
                                elif field_type_tag == "fk_audit_by":
                                    parsed_new_value_for_field = audit_by.objects.get(name=new_value_str_from_csv)
                                elif field_type_tag == "fk_return_by":
                                    parsed_new_value_for_field = return_by_user.objects.get(name=new_value_str_from_csv)
                                elif field_type_tag == "fk_segregation_by":
                                    parsed_new_value_for_field = segregation_by.objects.get(name=new_value_str_from_csv)
                                elif field_type_tag == "fk_request_by":
                                    parsed_new_value_for_field = request_by.objects.get(name=new_value_str_from_csv)
                                else: # string and string_nullable
                                     parsed_new_value_for_field = new_value_str_from_csv
                                # --- End of parsing logic ---

                                if getattr(current_db_value, 'pk', current_db_value) != getattr(parsed_new_value_for_field, 'pk', parsed_new_value_for_field):
                                    setattr(claim_to_update, model_field_name, parsed_new_value_for_field)
                                    fields_changed_in_this_record_log.append(f"{csv_header} updated")
                                    
                                    if model_field_name == "PolicyId" and parsed_new_value_for_field:
                                        claim_to_update.Client_name = parsed_new_value_for_field.client_name
                                        claim_to_update.ClientId = parsed_new_value_for_field.client_id
                                        fields_changed_in_this_record_log.append("Client info auto-updated")

                            except (Box.DoesNotExist, insurance_policy.DoesNotExist, ClaimStatus.DoesNotExist, audit_by.DoesNotExist, return_by_user.DoesNotExist, segregation_by.DoesNotExist, request_by.DoesNotExist) as e_fk_nf:
                                context['errors_details'].append(f"Claim '{claim_code_from_csv}', Field '{csv_header}': Related object '{new_value_str_from_csv}' not found. Field not updated.")
                            except (ValueError, InvalidOperation, TypeError) as e_parse:
                                context['errors_details'].append(f"Claim '{claim_code_from_csv}', Field '{csv_header}': Error parsing value '{new_value_str_from_csv}'. Field not updated.")

                        if fields_changed_in_this_record_log:
                            claim_to_update.Update_by = request.user if request.user.is_authenticated else None
                            claim_to_update.full_clean()
                            claim_to_update.save()
                            updated_records_count += 1
                            context['processed_records_details'].append(f"Claim '{claim_code_from_csv}': Successfully updated ({', '.join(fields_changed_in_this_record_log)}).")

                except ProviderSegregation.DoesNotExist:
                    context['errors_details'].append(f"Claim '{claim_code_from_csv}': Not found in database. Row skipped.")
                except ValidationError as ve:
                    error_msg_dict = '; '.join([f'{f}: {", ".join(m)}' for f, m in ve.message_dict.items()])
                    context['errors_details'].append(f"Claim '{claim_code_from_csv}': Validation error: {error_msg_dict}")
                except Exception as e_row:
                    context['errors_details'].append(f"Claim '{claim_code_from_csv}': An unexpected error occurred: {e_row}")

            # --- Final messaging to user ---
            if updated_records_count > 0:
                messages.success(request, f"Update complete. Applied changes to {updated_records_count} claim(s).")
            if context['errors_details']:
                messages.error(request, f"Process encountered {len(context['errors_details'])} error(s). Please review details below.")
            if not updated_records_count and not context['errors_details']:
                 messages.info(request, "Process complete. No claims required changes based on the CSV data.")

            return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)
    
    return render(request, "arcapp/claim_retrival/claim_ret_workloads.html", context)











################ login ####################

# def login_view(request):
#     if request.method == "POST":
#         form = AuthenticationForm(request, data=request.POST)
#         if form.is_valid():
#             username = form.cleaned_data.get('username')
#             password = form.cleaned_data.get('password')
#             user = authenticate(request, username=username, password=password)
#             if user is not None:
#                 login(request, user)  # Log the user in
#                 return redirect("box_list")  # Redirect to a page after successful login
#             else:
#                 form.add_error(None, "Invalid username or password")  # Show error if authentication fails
#         else:
#             # If form is invalid, render the form again with errors
#             return render(request, "testapp/users/login.html", {"form": form})
#     else:
#         form = AuthenticationForm()
#         return render(request, "arcapp/users/login.html", {"form": form})














########### vinicus junior ##########
# def login_view(request):

#     request.session.flush()
    
#     if request.method == "POST":
#         form = AuthenticationForm(request, data=request.POST)
#         if form.is_valid():
#             username = form.cleaned_data.get('username')
#             password = form.cleaned_data.get('password')
#             user = authenticate(request, username=username, password=password)
#             if user is not None:
#                 login(request, user)
                

#                 request.session['last_activity'] = timezone.now().isoformat()
                

#                 request.session.set_expiry(0)
                
#                 return redirect("box_list")
#             else:
#                 form.add_error(None, "Invalid username or password")
#         else: print(f"form errors : {form.errors}")        
#     else:
#         form = AuthenticationForm()
#         print("checked")
#     return render(request, "arcapp/users/login.html", {"form": form})


#### enforce user 










# def login_view(request):
#     if request.method == "POST":
#         username = request.POST.get('login-username')
#         password = request.POST.get('login-password')
        
#         # Get the user without logging them in yet
#         from django.contrib.auth import get_user_model
#         User = get_user_model()
        
#         try:
#             user = User.objects.get(username=username)
#             # Check if it's a first login by looking at last_login
#             is_first_login = user.last_login is None
            
#             # Now authenticate and login
#             user = authenticate(request, username=username, password=password)
#             if user is not None:
#                 login(request, user)
                
#                 if is_first_login:
#                     return redirect('password_change')
#                 else:
#                     return redirect("box_list")
#             else:
#                 messages.error(request, "Invalid password!")
#         except User.DoesNotExist:
#             messages.error(request, "Invalid username!")
            
#         return render(request, 'arcapp/users/login.html', {})
            
#     return render(request, 'arcapp/users/login.html', {})














def login_view(request):
    form_errors = {}  # Dictionary to store field-specific errors
    
    if request.method == "POST":
        username = request.POST.get('login-username')
        password = request.POST.get('login-password')
        
        # Get the user without logging them in yet
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        try:
            user = User.objects.get(username=username)
            # Check if it's a first login by looking at last_login
            is_first_login = user.last_login is None
            
            # Now authenticate and login
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                
                if not user.password_changed:  # Check the BooleanField
                    return redirect('password_change')
            
                return redirect("home_kpi")
            else:
                form_errors['login-password'] = ["Invalid password!"]
                messages.error(request, "Invalid password!")
        except User.DoesNotExist:
            form_errors['login-username'] = ["Invalid username!"]
            messages.error(request, "Invalid username!")
            
        return render(request, 'arcapp/users/login.html', {
            'form_errors': form_errors,
            'non_field_errors': []  # For general errors not tied to a specific field
        })
            
    return render(request, 'arcapp/users/login.html', {
        'form_errors': {},
        'non_field_errors': []
    })









from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import logout
from django.contrib import messages
from django.shortcuts import render, redirect


from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import logout
from django.contrib import messages
from django.shortcuts import render, redirect


def custom_password_change_view(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Set password_changed to True
            user.password_changed = True
            user.save()
            logout(request)
            messages.success(request, "Password updated. Please log in again.")
            return redirect('login')
        else:
            messages.error(request, "Error updating password.")
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'arcapp/users/password_change.html', {'form': form})

























# import csv
# from django.http import HttpResponse

def reimbursement_insert_sample_sheet(request):

    sample_data = [
        {
            'boxNumber': '123456',
            'Claim_Code': 'CLAIM001',
            'Batch_num': 'BATCH001',
            'Batch_type': 'TYPE_A',
            'English_name': 'John Doe',
            'Arab_name': 'جون دو',
            'Payer': 'Insurance Co.',
            'Policy': 'POLICY001',
            'Hof': 'Head of Family',
            'Audit_user': 'Auditor Name',
            'audit_date': '2025-01-22',  
            
        }
    ]


    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_sheet.csv"'


    writer = csv.DictWriter(response, fieldnames=sample_data[0].keys())
    writer.writeheader()
    writer.writerows(sample_data)

    return response

def provider_insert_by_bulk_sample_sheet(request):
																	

    sample_data = [
        {
            'boxNumber' : '256411',
            'ClaimCode' : 'CC-5016',
            'PolicyId' : 'P-95903',
            # 'policyname' : 'Policy-3',
            #'ClientId' : 'C-7647',
            'Member_name' : 'Member-28',
            'Provider_name' : 'Provider-42',
            'Receive_date' : '2024-07-10',
            'Issuance_date' : '2024-07-10',
            'Segregation_date' : '2024-07-10',
            'Segregation_by' : 'SegBy-17',
            'Audit_date' : '2024-07-10',
            'Audit_by' : 'AuditBy-9',
            #'Box_status' : 'Pending',
            'batchID' : 'B-2553',
            'request_by' : 'UNKOWN',
            'request_date' : "2025-05-01",
            #'Client_name' : 'Client-50',
            'claim_status' : 'test',
            'claimscan' : True,
            'DeductedAmount' : 5000,
            'note' : "note",
            "retrieval_date" : "2025-05-01",
            "return_date" : "2025-05-01",
            "return_by" : 'jocker',
            "comment" : "comment",

        }
    ]


    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_sheet.csv"'


    writer = csv.DictWriter(response, fieldnames=sample_data[0].keys())
    writer.writeheader()
    writer.writerows(sample_data)

    return response
def box_insert_by_bulk_sample_sheet(request):
    sample_data = [
        {
            'box_number' : '123456',
            'box_type' : 'type',
            'box_location' : 'INFOFORT',
            #'box_status' : 'retrieval',
            'insurance_company'	: 'AMC'

        }
    ]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="box_sheet.csv"'


    writer = csv.DictWriter(response, fieldnames=sample_data[0].keys())
    writer.writeheader()
    writer.writerows(sample_data)

    return response

def box_update_by_bulk_sample_sheet(request):
    sample_data = [
        {
            'box_number' : '123456',
            'box_type' : 'type',
            'box_location' : 'InfoFort',
           # 'box_status' : 'retrieval',
            'insurance_company'	: 'AMC'

        }
    ]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="box_sheet.csv"'


    writer = csv.DictWriter(response, fieldnames=sample_data[0].keys())
    writer.writeheader()
    writer.writerows(sample_data)

    return response

def provider_update_by_bulk_sample_sheet(request):
    """
    Generates a sample CSV file for bulk updating provider segregation records.
    The first column MUST be 'ClaimCode'. Other columns are optional.
    """
    # NOTE: The order of keys here determines the column order in the CSV.
    # ClaimCode MUST be the first column.
    sample_data = [
        {
            'ClaimCode' : '16661-1518', # This is the identifier for the update
            'box_number' : '179690',
            'PolicyId' : 'N/A', # Use 'N/A' or a valid ID to update, or leave blank to ignore
            'Member_name' : 'venom_updated',
            'Provider_name' : 'new_provider',
            'Receive_date' : '01/02/2025',
            'Issuance_date' : '01/02/2025',
            'Segregation_date' : '01/02/2025',
            'Segregation_by' : 'thor',
            'Audit_date' : '15/02/2025',
            'Audit_by' : 'auditor_new',
            'batchID' : '54321',
            'request_by' : 'ironman',
            'request_date' : '01/02/2025',
            'claim_status' : 'send', # Corrected header from 'Claim Status'
            'claimscan' : 'True',
            'DeductedAmount' : '750.50',
            'note' : 'This record has been updated via bulk upload.',
            'retrieval_date' : '01/03/2025',
            'return_date' : '15/03/2025',
            'return_by' : 'captain_america',
            'comment' : 'Updated comment.',
        }
    ]

    
    fieldnames = [
        'ClaimCode', 'box_number', 'PolicyId', 'Member_name', 'Provider_name', 
        'Receive_date', 'Issuance_date', 'Segregation_date', 'Segregation_by', 
        'Audit_date', 'Audit_by', 'batchID', 'request_by', 'request_date', 
        'claim_status', 'claimscan', 'DeductedAmount', 'note', 'retrieval_date', 
        'return_date', 'return_by', 'comment'
    ]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bulk_update_provider_sample.csv"'

    writer = csv.DictWriter(response, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(sample_data)

    return response

def reimbursement_update_sample_sheet(request):


    sample_data = [
        {
            'box_number': '179690',
            'Claim_Code': '16661-1518',
            'Batch_num': '12345',
            'Batch_type': 'Type A',
            'English_name': 'John Doe',
            'Arab_name': 'جون دو',
            'Payer': 'ABC Insurance',
            'Policy': 'Policy 123',
            'Hof': 'Hospital A',
            'Audit_user': 'admin',
            'Audit_date': '2023-11-22' 
        }
    ]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reimbursement_sheet.csv"'


    writer = csv.DictWriter(response, fieldnames=sample_data[0].keys())
    writer.writeheader()
    writer.writerows(sample_data)

    return response


###### reimbursemnt delete by bulk sample sheet #######

def generate_sample_bulk_delete_csv(request):

    claim_codes = [
        "174172-1",
        "174182-1",
        "174182-2",
        "174182-3",
        "174182-4",
        "174182-5",
    ]


    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bulk_delete_claims.csv"'


    writer = csv.writer(response)


    writer.writerow(["Claim Code"])


    for claim_code in claim_codes:
        writer.writerow([claim_code])

    return response


def generate_sample_bulk_delete_boxes_csv(request):

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bulk_delete_boxes_sample.csv"'


    writer = csv.writer(response)


    writer.writerow(['box_number'])


    sample_data = [
        ['12345'],
        ['67890'],
        ['11223'],
        ['44556'],
        ['78901'],
    ]


    for row in sample_data:
        writer.writerow(row)

    return response

########## bulk delete providers sample sheet ###############
import csv
from django.http import HttpResponse

def generate_sample_bulk_delete_provider_segre_csv(request):

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bulk_delete_provider_segre_sample.csv"'


    writer = csv.writer(response)


    writer.writerow(['claim_code'])


    sample_data = [
        ['174172-1'],
        ['174182-2'],
        ['174182-3'],
        ['174182-4'],
        ['174182-5'],
    ]


    for row in sample_data:
        writer.writerow(row)


    return response
def deduct(request):
    return render(request, "arcapp/deductions.html")



from django.http import JsonResponse
from .models import Box

def box_autocomplete(request):
    term = request.GET.get('term', '')
    if not term:
        return JsonResponse({'results': []})
    
    boxes = Box.objects.filter(box_number__icontains=term)[:50]  # Limit results
    results = [{'id': box.id, 'text': box.box_number} for box in boxes]
    return JsonResponse({'results': results})









############ master data module ##############


# def manage_master_data(request):
#     success = False
#     error_message = None

#     if request.method == 'POST':
#         form = MasterDataForm(request.POST)
#         if form.is_valid():
#             field_type = form.cleaned_data['field_type']
#             field_value = form.cleaned_data['field_value']
            
#             try:
#                 master_data.objects.create(**{field_type: field_value}, created_by=request.user)
#                 success = True
#                 # Clear form on success
#                 form = MasterDataForm()  
#             except IntegrityError:
#                 error_message = "⚠️ This value already exists! Please enter a unique value."
#             except Exception as e:
#                 error_message = f"⚠️ Error: {str(e)}"
#         else:
#             error_message = "⚠️ Please correct the errors below."
#     else:
#         form = MasterDataForm()

#     return render(request, 'arcapp/master_data_form.html', {
#         'form': form,
#         'success': success,
#         'error_message': error_message,
#     })

from django.shortcuts import render, redirect
from django.db import IntegrityError
from .models import BoxType, BoxLocation,request_by, InsuranceCompany, ClaimStatus, audit_by, return_by_user, segregation_by
from .forms import MasterDataForm, RequestByUsrForm,BoxTypeForm, BoxLocationForm, InsuranceCompanyForm, ClaimStatusForm, AuditByForm, ReurnByUsrForm,  SegregationByUsrForm

def manage_master_data(request):
    success = False
    error_message = None
    
    if request.method == 'POST':
        form = MasterDataForm(request.POST)
        if form.is_valid():
            field_type = form.cleaned_data['field_type']
            field_value = form.cleaned_data['field_value']
            
            try:
                # Create records in the appropriate model based on field_type
                if field_type == 'box_type':
                    BoxType.objects.create(name=field_value, created_by=request.user)
                elif field_type == 'box_location':
                    BoxLocation.objects.create(name=field_value, created_by=request.user)
                # elif field_type == 'box_status':
                #     BoxStatus.objects.create(name=field_value, created_by=request.user)
                elif field_type == 'claim_status':
                    ClaimStatus.objects.create(name = field_value, created_by = request.user)
                elif field_type == 'insurance_company':
                    InsuranceCompany.objects.create(name=field_value, created_by=request.user)
                elif field_type == 'audit_by':
                    audit_by.objects.create(name=field_value, created_by=request.user)
                elif field_type == 'return_by_usr':
                    return_by_user.objects.create(name=field_value, created_by=request.user)
                elif field_type == 'segregation_by_usr':
                    segregation_by.objects.create(name=field_value, created_by=request.user)
                elif field_type == 'request_by_user':
                    request_by.objects.create(name = field_value, created_by = request.user)

                    
                
                success = True
                # Clear form on success
                form = MasterDataForm()
            except IntegrityError:
                error_message = "⚠️ This value already exists! Please enter a unique value."
            except Exception as e:
                error_message = f"⚠️ Error: {str(e)}"
        else:
            error_message = "⚠️ Please correct the errors below."
    else:
        form = MasterDataForm()
    
    # Get existing values for display
    box_types = BoxType.objects.all().order_by('name')[:3]
    box_locations = BoxLocation.objects.all().order_by('name')[:3]
    claim_status = ClaimStatus.objects.all().order_by('name')[:3]
    #box_status = BoxStatus.objects.all().order_by('name')[:3]
    insurance_companies = InsuranceCompany.objects.all().order_by('name')
    audit_usr = audit_by.objects.all().order_by('name')[:3]
    return_by_usr = return_by_user.objects.all().order_by('name')[:3]
    segregation_by_usr = segregation_by.objects.all().order_by('name')[:3]
    request_by_usr = request_by.objects.all().order_by('name')[:3]
    
    return render(request, 'arcapp/master_data_form.html', {
        'form': form,
        'success': success,
        'error_message': error_message,
        'box_types': box_types,
        'box_locations': box_locations,
        'claim_status' : claim_status,
        'audit_by' : audit_usr,
        'return_by_usr' : return_by_usr,
        'sgeregation_by_usr' : segregation_by_usr,
        'request_by_usr' : request_by_usr,
        #'box_status' : box_status,
        'insurance_companies': insurance_companies,


    })






#######  update or elete master data  ############


from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404
from django.db import IntegrityError
from .models import BoxType, BoxLocation, InsuranceCompany
from .forms import BoxTypeForm, BoxLocationForm, InsuranceCompanyForm

def master_data_list(request):
    box_types = BoxType.objects.all().order_by('name')
    box_locations = BoxLocation.objects.all().order_by('name')
    claim_status = ClaimStatus.objects.all().order_by('name')
    #box_status = BoxStatus.objects.all().order_by('name')
    insurance_companies = InsuranceCompany.objects.all().order_by('name')
    audit_user = audit_by.objects.all().order_by('name')
    return_by_usr = return_by_user.objects.all().order_by('name')
    segregation_by_usr = segregation_by.objects.all().order_by('name')
    request_by_usr = request_by.objects.all().order_by('name')

    
    return render(request, 'arcapp/mater_data_management/master_data_list.html', {
        'box_types': box_types,
        'box_locations': box_locations,
        #'box_status' : box_status,
        'claim_status' : claim_status,
        'insurance_companies': insurance_companies,
        'audit_by' : audit_user,
        'return_by_usr' : return_by_usr,
        'segregation_by_usr' : segregation_by_usr,
        'request_by_usr' : request_by_usr,

    })

def update_master_data(request, model_type, id):
    model_map = {
        'box_type': (BoxType, BoxTypeForm),
        'box_location': (BoxLocation, BoxLocationForm),
        'claim_status' : (ClaimStatus, ClaimStatusForm),
        #'box_status': (BoxStatus, BoxStatusForm),
        'insurance_company': (InsuranceCompany, InsuranceCompanyForm),
        'audit_by' : (audit_by, AuditByForm),
        'return_by_usr' : (return_by_user, ReurnByUsrForm),
        'segregation_by_usr' : (segregation_by, SegregationByUsrForm),
        'request_by_usr' : (request_by, RequestByUsrForm),
        
    }
    
    if model_type not in model_map:
        raise Http404("Invalid model type")
    
    ModelClass, FormClass = model_map[model_type]
    instance = get_object_or_404(ModelClass, id=id)
    
    if request.method == 'POST':
        form = FormClass(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('master_data_list')
    else:
        form = FormClass(instance=instance)
    
    return render(request, 'arcapp/mater_data_management/update_master_data.html', {
        'form': form,
        'model_type': model_type,
        'instance': instance
    })

def delete_master_data(request, model_type, id):
    model_map = {
        'box_type': BoxType,
        'box_location': BoxLocation,
        #'box_status' : BoxStatus,
        'claim_status' : ClaimStatus,
        'insurance_company': InsuranceCompany,
        'audit_by': audit_by,
        'return_by_usr' : (return_by_user, ReurnByUsrForm),
        'segregation_by_usr' : (segregation_by, SegregationByUsrForm),
        'request_by_usr' : (request_by, RequestByUsrForm),

    }
    
    if model_type not in model_map:
        raise Http404("Invalid model type")
    
    ModelClass = model_map[model_type]
    instance = get_object_or_404(ModelClass, id=id)
    
    if request.method == 'POST':
        instance.delete()
        return redirect('master_data_list')
    
    return render(request, 'arcapp/mater_data_management/confirm_delete.html', {
        'instance': instance,
        'model_type': model_type
    })


############ insurance policy ##############
 
# def manage_insurance_policy(request):
#     if request.method == 'POST':
#         form = InsurancePolicyForm(request.POST)
#         if form.is_valid():
#             try:
#                 # inspolic = form.save()
#                 # inspolic.created_by = request.user


#                 inspolic = form.save(commit=False)
#                 inspolic.created_by = request.user
#                 inspolic.save()

#                 messages.success(request, "✅ Policy created successfully!")
#                 return redirect('manage_insurance_policy')
#             except Exception as e:
#                 messages.error(request, f"⚠️ Error: {str(e)}")
#         else:
#             messages.error(request, "⚠️ Please correct the errors below")
#     else:
#         form = InsurancePolicyForm()
    
#     return render(request, 'arcapp/insurance_policy.html', {
#         'form': form
#     })



@permission_required('arcapp.can_create_insurance', raise_exception=True)
def manage_insurance_policy(request):
    if request.method == 'POST':
        form = InsurancePolicyForm(request.POST)
        if form.is_valid():
            try:
                inspolic = form.save(commit=False)
                inspolic.created_by = request.user
                inspolic.save()

                messages.success(request, "✅ Policy created successfully!")
                return redirect('manage_insurance_policy')
            except Exception as e:
                messages.error(request, f"⚠️ Error: {str(e)}")
        else:
            messages.error(request, "⚠️ Please correct the errors below")
    else:
        form = InsurancePolicyForm()
    
    return render(request, 'arcapp/insurance_policy.html', {
        'form': form
    })






######## update the insurance ########


@permission_required('arcapp.can_update_insurance', raise_exception=True)
# def update_insurance_policy(request):
#     policy = None
#     success_message = None
#     error_message = None

#     if request.method == "POST":
#         if 'search' in request.POST:
#             current_policy_id = request.POST.get("current_policy_id")
#             try:
#                 policy = get_object_or_404(insurance_policy, policy_id=current_policy_id)
#             except Exception as e:
#                 error_message = f"Policy with ID {current_policy_id} not found."
#                 messages.error(request, error_message)
                
#         elif 'update' in request.POST:
#             policy_id = request.POST.get("policy_id")
#             new_policy_id = request.POST.get("new_policy_id")
#             new_client_name = request.POST.get("new_client_name")
#             new_insurance_company = request.POST.get("new_insurance_company")

#             try:
#                 policy = get_object_or_404(insurance_policy, id=policy_id)
                
#                 if new_policy_id:
#                     policy.policy_id = new_policy_id
                
#                 if new_client_name:
#                     policy.client_name = new_client_name
                
#                 if new_insurance_company:
#                     policy.insurance_company = get_object_or_404(InsuranceCompany, name=new_insurance_company)
                
#                 policy.Update_by = request.user
#                 policy.save()
#                 success_message = "Policy updated successfully!"
#                 messages.success(request, success_message)
                
#             except Exception as e:
#                 error_message = f"Error updating policy: {str(e)}"
#                 messages.error(request, error_message)

#     return render(request, 'arcapp/crud/insurance_update.html', {
#         'policy': policy,
#         'success_message': messages.get_messages(request),
#         'error_message': messages.get_messages(request),
#     })
                      

#@permission_required('arcapp.can_update_insurance', raise_exception=True)
def update_insurance_policy(request):
    policy = None
    success_message = None
    error_message = None

    if request.method == "POST":
        if 'search' in request.POST:
            current_policy_id = request.POST.get("current_policy_id")
            try:
                policy = get_object_or_404(insurance_policy, policy_id=current_policy_id)
            except Http404:
                error_message = f"Policy with ID {current_policy_id} not found."
            except Exception as e:
                error_message = f"Error searching policy: {str(e)}"

        elif 'update' in request.POST:
            policy_id = request.POST.get("policy_id")
            new_policy_id = request.POST.get("new_policy_id")
            new_client_name = request.POST.get("new_client_name")
            new_client_id = request.POST.get("new_client_id")
            new_insurance_company = request.POST.get("new_insurance_company")

            try:
                policy = get_object_or_404(insurance_policy, id=policy_id)
                # Update fields
                policy.policy_id = new_policy_id
                policy.client_name = new_client_name
                policy.client_id = new_client_id
                policy.insurance_company = get_object_or_404(InsuranceCompany, name=new_insurance_company)
                policy.Update_by = request.user
                policy.save()
                success_message = "Policy updated successfully!"
            except Http404:
                error_message = "Policy not found during update."
            except Exception as e:
                error_message = f"Error updating policy: {str(e)}"

    return render(request, 'arcapp/crud/insurance_update.html', {
        'policy': policy,
        'success_message': success_message,
        'error_message': error_message,
    })

################### delete insurance ################ 

# def delete_insurance(request):
#     success_message = None
#     failure_message = None
#     if request.method == "POST" and "delete" in request.POST.get:
#         policyid = request.POST.get('PolicyId')
#         print(policyid)
#         if not policyid:
#             failure_message = f"the policy you encountered {policyid} net exist "
#             messages.error(request, failure_message)
#         else:
#             try: 
#                 Policy = get_object_or_404(insurance_policy, policy_id = policyid)
#                 Policy.delete()
#                 success_message = f"the insurance company with policy id = {policyid} had been ddeleted successfully"
#                 messages.success(request, success_message)
#             except Http404:
#                 failure_message = f"Box with number '{policyid}' not found."
#                 messages.error(request, failure_message)
#             except Exception as e:
                
#                 failure_message = f"An error occurred while deleting the Box: {str(e)}"
#                 messages.error(request, failure_message)

#     return render(request, 'arcapp/crud/insurance_delete.html', {
#        'success_message': messages.get_messages(request),
#         'error_message': messages.get_messages(request),
#     })

@permission_required('arcapp.can_delete_insurance', raise_exception=True)
def delete_insurance(request):
    success_message = None
    error_message = None

    if request.method == "POST" and "delete" in request.POST:
        policy_id = request.POST.get("PolicyId")
        
        if not policy_id:
            error_message = "Policy ID is required."
        else:
            try:
                policy = get_object_or_404(insurance_policy, policy_id=policy_id)
                policy.delete()
                success_message = f"Insurance policy {policy_id} deleted successfully."
            except Http404:
                error_message = f"Policy with ID {policy_id} not found."
            except Exception as e:
                error_message = f"Error deleting policy: {str(e)}"

    return render(request, 'arcapp/crud/insurance_delete.html', {
        'success_message': success_message,
        'error_message': error_message,
    })




########## insurance policy dataframe #############
@permission_required('arcapp.can_view_inurance', raise_exception=True)
def view_insurance_policies(request):

    all_policies = insurance_policy.objects.all()
    
    return render(request, 'arcapp/simple_policy_view.html', {
        'all_policies': all_policies,
    })





def export_policies_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="insurance_policiestest_fady.csv"'

    writer = csv.writer(response)
    
    writer.writerow(['Policy ID', 'Insurance Company', 'Client Name'])
    
    policies = insurance_policy.objects.all().select_related('insurance_company')
    
    for policy in policies:
        writer.writerow([
            policy.policy_id,
            policy.insurance_company,
            policy.client_name or '-'
        ])
    
    return response






######### the insurance company search #######


def insurance_company_search_request(request):
    term = request.GET.get('term', '')
    if not term:
        return JsonResponse({'results': []})
    
    # Use the correct field name - if your field is 'insurance_company' instead of 'name'
    insurance_companies = InsuranceCompany.objects.filter(name__icontains=term)[:20]
    
    # Make sure the field name in the results matches your model field
    results = [{'id': company.id, 'text': company.name} for company in insurance_companies]
    
    return JsonResponse({'results': results})
# views.py
def policyid_autocomplete(request):
    term = request.GET.get('term', '')
    if not term:
        return JsonResponse({'results': []})
    
    try:
        # Convert term to integer for exact match
        term_int = int(term)
        policies = insurance_policy.objects.filter(policy_id=term_int)[:20]
    except ValueError:
        # Handle non-integer input
        policies = insurance_policy.objects.none()
    
    results = [{'id': policy.id, 'text': str(policy.policy_id)} for policy in policies]
    return JsonResponse({'results': results})


############ claims per insurance policy ###########

def claims_report(request):
    insurance_companies = InsuranceCompany.objects.all().order_by('name')
    selected_company = request.GET.get('insurance_company')
    policy_number = request.GET.get('policy_number', '').strip()
    
    claims = ProviderSegregation.objects.none()
    error = None

    if selected_company:
        try:
            company = InsuranceCompany.objects.get(name=selected_company)
            boxes = Box.objects.filter(insurance_company=company)
            claims = ProviderSegregation.objects.filter(box_number__in=boxes)
            policy_info = insurance_policy.objects.filter(policy_id = str(claims.PolicyId))
            
            if policy_number:
                try:
                    policy_id = int(policy_number)
                    claims = claims.filter(PolicyId__policy_id=policy_id)
                except ValueError:
                    error = "Invalid policy number format - must be numeric"
            
        except InsuranceCompany.DoesNotExist:
            error = "Selected insurance company not found"
    else:
        if request.GET:
            error = "Insurance company selection is required"

    # Generate CSV if no errors and form is submitted
    if request.GET and not error:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="claims_report_{selected_company}.csv"'
        
        writer = csv.writer(response)
    
        writer.writerow([
            'Claim Code', 'Box Number', 'Insurance Company',
            'Policy ID', 'Member Name', 'Receive Date', 'Issuance Date'
        ])
        
        # Write data rows
        for claim in claims:
            writer.writerow([
                claim.ClaimCode,
                claim.box_number.box_number,
                claim.box_number.insurance_company.name,
                claim.PolicyId.policy_id,
                claim.Member_name,
                claim.Receive_date.strftime("%Y-%m-%d") if claim.Receive_date else '',
                claim.Issuance_date.strftime("%Y-%m-%d") if claim.Issuance_date else '',
            ])
        
        return response
    else:
        
        return render(request, 'arcapp/reports/claims_per_policy_ins.html', {
            'insurance_companies': insurance_companies,
            'selected_company': selected_company,
            'policy_number': policy_number,
            'error': error
        })


########## claims per box report ##############

from django.shortcuts import render, get_object_or_404

import csv
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from .models import Box, ProviderSegregation

import csv
from io import TextIOWrapper
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from .models import Box, ProviderSegregation

def claims_per_box(request):
    error = None
    context = {}
    
    if request.method == 'POST':
        if 'box_file' in request.FILES:
            csv_file = TextIOWrapper(request.FILES['box_file'].file, encoding=request.encoding)
            reader = csv.DictReader(csv_file)
            
            box_numbers = [row['box_number'] for row in reader if 'box_number' in row]
            
            if not box_numbers:
                error = "CSV file is empty or missing 'box_number' column"
            else:
                valid_boxes = Box.objects.filter(box_number__in=box_numbers)
                missing_boxes = set(box_numbers) - set(valid_boxes.values_list('box_number', flat=True))
                
                if missing_boxes:
                    error = f"Boxes not found: {', '.join(missing_boxes)}"
                
                # Collect all claims from valid boxes
                claims = ProviderSegregation.objects.filter(box_number__in=valid_boxes).select_related(
                    'box_number', 
                    'box_number__insurance_company',
                    'box_number__box_type',
                    'box_number__box_location',
                    'box_number__created_by',
                    'box_number__Update_by', # Corrected related_name if necessary or direct field name
                    'PolicyId',
                    'PolicyId__insurance_company',
                    'PolicyId__created_by',
                    'PolicyId__Update_by', # Corrected related_name if necessary or direct field name
                    'claim_status',
                    'created_by',
                    'Update_by' # Corrected related_name if necessary or direct field name
                )
                
                if claims.exists():
                    response = HttpResponse(content_type='text/csv')
                    filename = f"claims_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv" # Added seconds for more uniqueness
                    response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    
                    writer = csv.writer(response)
                    # Updated header row
                    writer.writerow([
                        # Box Data
                        'Box Number', 'Box Type', 'Box Location', 'Box Insurance Company', 
                        'Box Data Entrance Date', 'Box Created By', 'Box Updated By',
                        # ProviderSegregation Data
                        'Claim Code', 'Client ID (Segregation)', 'Member Name', 'Provider Name',
                        'Receive Date', 'Issuance Date', 'Segregation Date', 'Segregation By',
                        'Audit Date', 'Audit By', 'Batch ID', 'Request By', 'Request Date',
                        'Client Name (Segregation)', 'Claim Scan', 'Deducted Amount', 'Note',
                        'Retrieval Date', 'Return By', 'Return Date', 'Comment',
                        'Data Entrance Date (Segregation)', 'Claim Status', 
                        'Created By (Segregation)', 'Updated By (Segregation)',
                        # Insurance Policy Data (related to ProviderSegregation.PolicyId)
                        'Policy ID', 'Policy Insurance Company', 'Policy Client Name', 
                        'Policy Client ID', 'Policy Created By', 'Policy Updated By',
                    ])
                    
                    for claim in claims:
                        # Helper to safely get related user username
                        def get_username(user_field):
                            return user_field.username if user_field else ''

                        # Helper to safely get related field name or id
                        def get_related_name(related_obj, attr='name'):
                            return getattr(related_obj, attr) if related_obj else ''
                        
                        writer.writerow([
                            # Box Data
                            claim.box_number.box_number,
                            get_related_name(claim.box_number.box_type),
                            get_related_name(claim.box_number.box_location),
                            get_related_name(claim.box_number.insurance_company),
                            claim.box_number.data_entrance_date.strftime("%Y-%m-%d %H:%M:%S") if claim.box_number.data_entrance_date else '',
                            get_username(claim.box_number.created_by),
                            get_username(claim.box_number.Update_by), # Assumes 'Update_by' is the correct field name
                            
                            # ProviderSegregation Data
                            claim.ClaimCode,
                            claim.ClientId,
                            claim.Member_name,
                            claim.Provider_name,
                            claim.Receive_date.strftime("%Y-%m-%d") if claim.Receive_date else '',
                            claim.Issuance_date.strftime("%Y-%m-%d") if claim.Issuance_date else '',
                            claim.Segregation_date.strftime("%Y-%m-%d") if claim.Segregation_date else '',
                            claim.Segregation_by,
                            claim.Audit_date.strftime("%Y-%m-%d") if claim.Audit_date else '',
                            claim.Audit_by,
                            claim.batchID,
                            claim.request_by,
                            claim.request_date.strftime("%Y-%m-%d") if claim.request_date else '',
                            claim.Client_name, # This is ProviderSegregation.Client_name
                            claim.claimscan,
                            claim.DeductedAmount,
                            claim.note,
                            claim.retrieval_date.strftime("%Y-%m-%d") if claim.retrieval_date else '',
                            claim.return_by,
                            claim.return_date.strftime("%Y-%m-%d") if claim.return_date else '',
                            claim.comment,
                            claim.data_entrance_date.strftime("%Y-%m-%d %H:%M:%S") if claim.data_entrance_date else '',
                            get_related_name(claim.claim_status),
                            get_username(claim.created_by),
                            get_username(claim.Update_by), # Assumes 'Update_by' is the correct field name

                            # Insurance Policy Data
                            claim.PolicyId.policy_id if claim.PolicyId else '',
                            get_related_name(claim.PolicyId.insurance_company) if claim.PolicyId else '',
                            claim.PolicyId.client_name if claim.PolicyId else '',
                            claim.PolicyId.client_id if claim.PolicyId else '',
                            get_username(claim.PolicyId.created_by) if claim.PolicyId else '',
                            get_username(claim.PolicyId.Update_by) if claim.PolicyId else '' # Assumes 'Update_by' is the correct field name
                        ])
                    
                    return response
                else:
                    error = "No claims found for the provided box numbers"
        else:
            error = "No file uploaded." # Handle case where 'box_file' is not in request.FILES
            
    context['error'] = error
    
    return render(request, 'arcapp/reports/claims_per_box.html', context)
def generate_sample_sheet_claims_per_box(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bulk_claims_sample.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['box_number'])
    
    # Generate sample data from existing boxes
    sample_boxes = Box.objects.order_by('?')[:5]  # Random 5 boxes
    for box in sample_boxes:
        writer.writerow([box.box_number])
    
    return response







######## claims per batch report #########

# def claims_per_batch_report(request):
#     ClaimsBatch = ProviderSegregation.objects.none()
#     entered_batch = None
#     seg_start = None
#     seg_end = None
#     audit_start = None
#     audit_end = None
#     issue_start = None
#     issue_end = None
#     if request.method == 'POST':
#         entered_batch = request.POST.get("batchID")
#         seg_start = request.POST.get('seg_start')
#         seg_end = request.POST.get('seg_end')
#         audit_start = request.POST.get('audit_start')
#         audit_end = request.POST.get('audit_end')
#         issue_start = request.POST.get('issue_start')
#         issue_end = request.POST.get('issue_end')

#         ClaimsBatch = ProviderSegregation.objects.filter(batchID = entered_batch)

#         if seg_start and seg_end:
#             ClaimsBatch = ClaimsBatch.filter(Segregation_date__range=(seg_start, seg_end))
        
#         if audit_start and audit_end:
#             ClaimsBatch = ClaimsBatch.filter(Audit_date__range=(audit_start, audit_end))
        
#         if issue_start and issue_end:
#             ClaimsBatch = ClaimsBatch.filter(Issuance_date__range=(issue_start, issue_end))

#     return render(request, 'arcapp/reports/claims_per_batch.html', {
#     'ClaimsBatch': ClaimsBatch,
#     'entered_batch': entered_batch,
#     'seg_start': seg_start,
#     'seg_end': seg_end,
#     'audit_start': audit_start,
#     'audit_end': audit_end,
#     'issue_start': issue_start,
#     'issue_end': issue_end
# })



def claims_per_batch_report(request):
    context = {
        'entered_batch': None,
        'seg_start': None,
        'seg_end': None,
        'audit_start': None,
        'audit_end': None,
        'issue_start': None,
        'issue_end': None
    }

    if request.method == 'POST':
        
        entered_batch = request.POST.get("batchID")
        seg_start = request.POST.get('seg_start')
        seg_end = request.POST.get('seg_end')
        audit_start = request.POST.get('audit_start')
        audit_end = request.POST.get('audit_end')
        issue_start = request.POST.get('issue_start')
        issue_end = request.POST.get('issue_end')

        
        claims = ProviderSegregation.objects.filter(batchID=entered_batch)

        
        if seg_start and seg_end:
            claims = claims.filter(Segregation_date__range=(seg_start, seg_end))
        if audit_start and audit_end:
            claims = claims.filter(Audit_date__range=(audit_start, audit_end))
        if issue_start and issue_end:
            claims = claims.filter(Issuance_date__range=(issue_start, issue_end))

        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="batch_{entered_batch}_claims.csv"'
        
        writer = csv.writer(response)
        
        writer.writerow([
            'Claim Code', 'Batch ID', 'Member Name', 'Provider Name',
            'Segregation Date', 'Audit Date', 'Issuance Date', 'Box Number'
        ])
        
        
        for claim in claims:
            writer.writerow([
                claim.ClaimCode,
                claim.batchID,
                claim.Member_name,
                claim.Provider_name,
                claim.Segregation_date.strftime("%Y-%m-%d") if claim.Segregation_date else '',
                claim.Audit_date.strftime("%Y-%m-%d") if claim.Audit_date else '',
                claim.Issuance_date.strftime("%Y-%m-%d") if claim.Issuance_date else '',
                claim.box_number.box_number if claim.box_number else ''
            ])
        
        return response

    return render(request, 'arcapp/reports/claims_per_batch.html', context)


############ claims per request date ###########
# def claims_per_request_date_report(request):
#     context = {
#         'ClaimsRequest': ProviderSegregation.objects.none(),
#         'req_start': None,
#         'req_end': None
#     }
    
#     if request.method == "POST":
#         req_start = request.POST.get("request_start_date")
#         req_end = request.POST.get("request_end_date")
        
#         if req_start and req_end:
#             context['ClaimsRequest'] = ProviderSegregation.objects.filter(
#               request_date__range=(req_start, req_end))
            
#         context.update({
#             'req_start': req_start,
#             'req_end': req_end
#         })
    
#     return render(request, 'arcapp/reports/claims_per_request_date.html', context)

def claims_per_request_date_report(request):
    context = {
        'req_start': None,
        'req_end': None,
        'error': None
    }
    
    if request.method == "POST":
        req_start = request.POST.get("request_start_date")
        req_end = request.POST.get("request_end_date")
        
        if req_start and req_end:
            claims = ProviderSegregation.objects.filter(request_date__range=(req_start, req_end))
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            filename = f"claims_{req_start}_to_{req_end}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            writer = csv.writer(response)
            # Write headers
            writer.writerow([
                'Claim Code', 'Request Date', 'Member Name',
                'Provider Name', 'Box Number', 'Batch ID',
                'Segregation Date'
            ])
            
            # Write data rows
            for claim in claims:
                writer.writerow([
                    claim.ClaimCode,
                    claim.request_date.strftime("%Y-%m-%d") if claim.request_date else '',
                    claim.Member_name,
                    claim.Provider_name,
                    claim.box_number.box_number if claim.box_number else '',
                    claim.batchID,
                    claim.Segregation_date.strftime("%Y-%m-%d") if claim.Segregation_date else ''
                ])
            
            return response
        else:
            context['error'] = "Both start and end dates are required"
        
        context.update({
            'req_start': req_start,
            'req_end': req_end
        })
    
    return render(request, 'arcapp/reports/claims_per_request_date.html', context)


################## count of boc - claims - batches per inscom ##############

# def count_per_insurance_report(request):
#     companies = inscom = summary =None
#     if request.method == "POST":
#         inscom = request.POST.get("insurance_company")
#         companies = Box.objects.values_list("insurance_company", flat=True).distinct()
#         summary = {}
#         if inscom:
#             boxes = Box.objects.filter(insurance_company = inscom)
#             claims = ProviderSegregation.objects.filter(box_number__insurance_company = inscom)

#             summary = {
#                 'boxes': boxes.count(),
#                 'claims': claims.count(),
#                 'batches': claims.values('batchID').distinct().count()
#             }
#     return render(request, 'arcapp/reports/count_per_insurance_report.html', {
#         'companies': companies,
#         'selected': inscom,
#         'summary': summary
#     })

def count_per_insurance_report(request):
    # Get all insurance companies WITH boxes
    companies = InsuranceCompany.objects.filter(box__isnull=False).distinct()
    selected_company = None
    summary = None

    if request.method == "POST":
        selected_id = request.POST.get("insurance_company")
        
        if selected_id:
            try:
                selected_company = InsuranceCompany.objects.get(id=selected_id)
                boxes = Box.objects.filter(insurance_company=selected_company)
                claims = ProviderSegregation.objects.filter(box_number__insurance_company=selected_company)

                summary = {
                    'boxes': boxes.count(),
                    'claims': claims.count(),
                    'batches': claims.values('batchID').distinct().count()
                }
            except InsuranceCompany.DoesNotExist:
                pass  # Handle invalid selection if needed

    return render(request, 'arcapp/reports/count_per_insurance_report.html', {
        'companies': companies,
        'selected': selected_company.id if selected_company else None,
        'summary': summary
    })




############# box number by location filter insurance #############
from .forms import BoxStatsReportForm
def box_stats_report(request):
    form = BoxStatsReportForm(request.GET or None)
    boxes = Box.objects.all()

    if request.GET:
        # Apply filters
        locations = request.GET.getlist('locations')
        statuses = request.GET.getlist('statuses')
        types = request.GET.getlist('types')
        insurance_companies = request.GET.getlist('insurance_companies')

        if locations:
            boxes = boxes.filter(box_location__in=locations)
        if statuses:
            boxes = boxes.filter(box_status__in=statuses)
        if types:
            boxes = boxes.filter(box_type__in=types)
        if insurance_companies:
            boxes = boxes.filter(insurance_company__in=insurance_companies)
    

    context = {
        'form': form,
        'boxes': boxes,
    }
    return render(request, 'arcapp/reports/box_stats_report.html', context)

########### claims per date ###############

def export_data(segregations, date_field, format):
    """Generic export handler"""
    field_labels = {
        'Receive_date': 'Receive Date',
        'Issuance_date': 'Issuance Date',
        'Segregation_date': 'Segregation Date',
        'Audit_date': 'Audit Date',
        'request_date': 'Request Date',
    }

    # Generate headers
    headers = [
        'Claim Code', 'Provider Name', 'Box Number',
        field_labels[date_field],
        'Policy ID', 'Client Name', 'Audit Date', 'Request Date'
    ]

    # Generate rows
    rows = []
    for seg in segregations:
        rows.append([
            seg.ClaimCode,
            seg.Provider_name,
            seg.box_number.box_number,
            getattr(seg, date_field).strftime("%Y-%m-%d"),
            seg.PolicyId.policy_id,
            seg.Client_name,
            seg.Audit_date.strftime("%Y-%m-%d"),
            seg.request_date.strftime("%Y-%m-%d")
        ])

    # Create response based on format
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    elif format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Provider Segregations"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(response)
        return response

    return None
from .forms import DateFilterForm
def claims_per_date(request):
    form = DateFilterForm(request.GET or None)
    context = {'form': form}

    if form.is_valid():
        date_field = form.cleaned_data['date_field']
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        export_format = form.cleaned_data['export_format']

        # Filter data using actual model field names
        segregations = ProviderSegregation.objects.filter(
            **{f"{date_field}__range": [start_date, end_date]}
        ).select_related('box_number', 'PolicyId').order_by(f'-{date_field}')

        # Handle exports
        if export_format in ['csv', 'xlsx']:
            filename = f"segregations_{date_field}_{start_date}_{end_date}.{export_format}"
            response = export_data(segregations, date_field, export_format)
            if response:
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

        # Prepare HTML context
        context.update({
            'segregations': segregations,
            'start_date': start_date,
            'end_date': end_date,
            'selected_date_label': dict(form.fields['date_field'].choices)[date_field],
        })

    return render(request, 'arcapp/reports/claims_per_date.html', context)

################### claims by actor ################
from .forms  import ActorFilterForm
def claims_per_actor(request):
    form = ActorFilterForm(request.GET or None)
    segregations = []
    context = {'form': form}

    if form.is_valid():
        actor_type = form.cleaned_data['actor_type']
        date_type = form.cleaned_data['date_type']
        actor_query = form.cleaned_data['actor_query']
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        export_format = form.cleaned_data['export_format']

        # Build filters
        filters = {
            f'{date_type}__range': [start_date, end_date]
        }
        if actor_query:
            filters[f'{actor_type}__icontains'] = actor_query

        segregations = ProviderSegregation.objects.filter(**filters)\
                           .select_related('box_number', 'PolicyId')\
                           .order_by(f'-{date_type}')

        # Add temporary attributes for template
        for seg in segregations:
            seg.selected_actor = getattr(seg, actor_type)
            seg.selected_date = getattr(seg, date_type)

        # Handle exports
        if export_format in ['csv', 'xlsx']:
            filename = f"claims_{actor_type}_{date_type}_{start_date}_to_{end_date}.{export_format}"
            
            if export_format == 'csv':
                response = HttpResponse(content_type='text/csv')
                writer = csv.writer(response)
                writer.writerow(['Claim Code', 'Actor', 'Date', 'Provider', 'Box', 'Policy', 'Client'])
                for seg in segregations:
                    writer.writerow([
                        seg.ClaimCode,
                        seg.selected_actor,
                        seg.selected_date.strftime("%Y-%m-%d"),
                        seg.Provider_name,
                        seg.box_number.box_number,
                        seg.PolicyId.policy_id,
                        seg.Client_name
                    ])
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            
            elif export_format == 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(['Claim Code', 'Actor', 'Date', 'Provider', 'Box', 'Policy', 'Client'])
                for seg in segregations:
                    ws.append([
                        seg.ClaimCode,
                        seg.selected_actor,
                        seg.selected_date,
                        seg.Provider_name,
                        seg.box_number.box_number,
                        seg.PolicyId.policy_id,
                        seg.Client_name
                    ])
                wb.save(response)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

        context.update({
            'segregations': segregations,
            'actor_label': dict(form.fields['actor_type'].choices)[actor_type],
            'date_label': dict(form.fields['date_type'].choices)[date_type],
            'start_date': start_date,
            'end_date': end_date,
            'actor_query': actor_query,
        })

    return render(request, 'arcapp/reports/claims_per_actor.html', context)


##################### data statging stagging ############################

from django.core.paginator import Paginator
from django.db.models import Q
from .models import SegregationStep
from .forms import SegregationStepForm


def segregation_step_list(request):
    """List all segregation steps with search and pagination"""
    query = request.GET.get('q', '')
    segregation_steps = SegregationStep.objects.all()
    
    if query:
        segregation_steps = segregation_steps.filter(
            Q(claim_code__icontains=query) |
            Q(member_name__icontains=query) |
            Q(client_name__icontains=query) |
            Q(insurance_company__icontains=query)
        )
    
    paginator = Paginator(segregation_steps, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'total_count': segregation_steps.count()
    }
    return render(request, 'arcapp/segregation_step/list.html', context)

def segregation_step_create(request):
    """Create a new segregation step"""
    if request.method == 'POST':
        form = SegregationStepForm(request.POST)
        if form.is_valid():
            segregation_step = form.save()
            messages.success(request, f'Segregation 01146866415 test "{segregation_step.claim_code}" created successfully.')
            return redirect('segregation_step_detail', ms=segregation_step.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SegregationStepForm()
    
    fdii = {
        'form': form,
        'title': 'Create Segregation Step'
    }
    return render(request, 'arcapp/segregation_step/form.html', fdii)

def segregation_step_detail(request, pk):
    """Display segregation step details"""
    segregation_step = get_object_or_404(SegregationStep, pk=pk)
    fady = {
        'segregation_step': segregation_step
    }
    return render(request, 'arcapp/segregation_step/detail.html', fady)

def segregation_step_update(request, pk):
    """Update an existing segregation step"""
    segregation_step = get_object_or_404(SegregationStep, pk=pk)
    
    if request.method == 'POST':
        form = SegregationStepForm(request.POST, instance=segregation_step)
        if form.is_valid():
            form.save()
            messages.success(request, f'The Claim Coe "{segregation_step.claim_code}" updated successfully.')
            return redirect('segregation_step_detail', pk=segregation_step.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SegregationStepForm(instance=segregation_step)
    
    context = {
        'form': form,
        'segregation_step': segregation_step,
        'title': 'Update Segregation Step'
    }
    return render(request, 'arcapp/segregation_step/form.html', context)

def segregation_step_delete(request, pk):
    """Delete a segregation step"""
    segregation_step = get_object_or_404(SegregationStep, pk=pk)
    
    if request.method == 'POST':
        segregation_step.delete()
        messages.success(request, f'Segregation step "{segregation_step.claim_code}" deleted successfully.')
        return redirect('segregation_step_list')
    
    context = {
        'segregation_step': segregation_step
    }
    return render(request, 'arcapp/segregation_step/delete.html', context)
    














############### claim retrieval #################




















############ bulk insert segregtion staging ###########

def SegreStgIns(request):
    if request.method == 'POST':
        form = SegreStgCsvForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']

            if not csv_file.name.endswith('.csv'):
                messages.error("please enter valid csv i.e amc.csv (preferred utf-8)")
                return redirect('SegreStgIns')
            stage_vals = []
            decoding = csv_file.read().decode('utf-8')
            io_string = io.StringIO(decoding)
            reader = csv.DictReader(io_string)
            for row in reader:
                policy_id = row.get('Policy_id', '').strip()
                client_id = row.get('client_id', '').strip()
                Insurance_company_csv = row.get('insurance_company', '').strip()
                member_name = row.get('member_name', '').strip()
                card_number = row.get('card_number', '').strip()
                emp_code = row.get('emp_code', '').strip()
                batch_receiving = row.get('batch_receiving', '').strip()
                parent_bid = row.get('parent_bid', '').strip()
                batch_type = row.get('batch_type', '').strip()
                provider_type_name = row.get('provider_type_name', '').strip()
                pid = row.get('pid', '').strip()
                claim_code = row.get('claim_code', '').strip()
                service_date = row.get('service_date', '').strip()
                        ##### to be continu ^..^
            